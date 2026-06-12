from __future__ import annotations

import json
import re
import threading
import xml.etree.ElementTree as ET
from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

GRID_LABELS = {
    1: "1问题员工",
    2: "2差距员工",
    3: "3基本胜任",
    4: "4待发展者",
    5: "5中坚力量",
    6: "6熟练员工",
    7: "7潜力之星",
    8: "8绩效之星",
    9: "9超级明星",
}


class TalentReviewStoreMixin:
    def _parse_grid(self, value, fallback=5):
        if value is None or value == "":
            return fallback
        match = re.search(r"[1-9]", str(value))
        return int(match.group(0)) if match else fallback

    def _row_employee_id(self, row):
        explicit = row.get("员工ID") or row.get("工号") or row.get("employeeId")
        if explicit:
            return self._normalize_employee_id(explicit)
        name = row.get("姓名", "")
        department = row.get("组织全称", "")
        return f"{name}|{department}".strip()

    def _first_value(self, row, names):
        for name in names:
            value = row.get(name)
            if value not in (None, ""):
                return value
        normalized = {re.sub(r"\s+", "", str(key)).lower(): value for key, value in row.items()}
        for name in names:
            value = normalized.get(re.sub(r"\s+", "", str(name)).lower())
            if value not in (None, ""):
                return value
        return ""

    def _clean_compare_value(self, value):
        return re.sub(r"\s+", "", str(value or "").strip())

    def _ledger_adjustments(self, row, original_grid, current_grid):
        adjustments = []

        def add_adjustment(label, before, after, require_before=False):
            before_text = str(before or "").strip()
            after_text = str(after or "").strip()
            if require_before and not before_text:
                return
            if (before_text or after_text) and self._clean_compare_value(before_text) != self._clean_compare_value(after_text):
                adjustments.append({"label": label, "from": before_text or "空白", "to": after_text or "空白"})

        calibrated_grid_value = self._first_value(row, ["校准后九宫格位置"])
        if calibrated_grid_value not in (None, "") and original_grid and current_grid and original_grid != current_grid:
            add_adjustment("九宫格位置", GRID_LABELS.get(original_grid, str(original_grid)), GRID_LABELS.get(current_grid, str(current_grid)))
        add_adjustment(
            "AI人才标签",
            self._first_value(row, ["AI人才标签"]),
            self._first_value(row, ["校准后AI人才标签"]),
            require_before=True,
        )
        add_adjustment(
            "无成长预警",
            self._first_value(row, ["无成长预警"]),
            self._first_value(row, ["校准后无成长预警", "无成长预警"]),
        )
        return adjustments

    def _normalize_department(self, value):
        normalized = str(value or "").strip()
        normalized = normalized.replace(" > ", "/").replace(">", "/")
        normalized = re.sub(r"/+", "/", normalized)
        return normalized.strip("/")

    def _profile_department_values(self, profile):
        values = [
            profile.get("departmentPathRaw"),
            profile.get("departmentPath"),
            profile.get("组织全称"),
            profile.get("department"),
        ]
        return [self._normalize_department(value) for value in values if value]

    def _profile_matches_review_department(self, profile, review_department):
        review = self._normalize_department(review_department)
        if not review:
            return False
        for profile_department in self._profile_department_values(profile):
            if profile_department and (review == profile_department or review.endswith("/" + profile_department)):
                return True
        return False

    def _profile_for_result(self, result, profiles, profile_by_id):
        employee_id = result.get("employeeId")
        if employee_id in profile_by_id:
            return profile_by_id[employee_id]
        same_name = [profile for profile in profiles if profile.get("name") == result.get("name")]
        if len(same_name) == 1:
            return same_name[0]
        for profile in same_name:
            if self._profile_matches_review_department(profile, result.get("departmentPath")):
                return profile
        return {}

    def _department_levels(self, department_path):
        parts = [part.strip() for part in str(department_path or "").split("/") if part and part.strip()]
        levels = parts[:5] + [""] * max(0, 5 - len(parts))
        return {
            "一级组织": levels[0],
            "二级组织": levels[1],
            "三级组织": levels[2],
            "四级组织": levels[3],
            "五级组织": levels[4],
        }

    def _review_row_to_record(self, row):
        original_grid = self._parse_grid(self._first_value(row, ["九宫格位置", "原九宫格位置"]), None)
        calibrated_grid_value = self._first_value(row, ["校准后九宫格位置"])
        has_calibrated_grid = calibrated_grid_value not in (None, "")
        current_grid = self._parse_grid(calibrated_grid_value, original_grid)
        supervisor_adjustments = self._ledger_adjustments(row, original_grid, current_grid)
        department_path = self._first_value(row, ["组织全称"])
        department_levels = self._department_levels(department_path)
        history_value = self._first_value(row, ["近3年九宫格位置", "近3年九宫格", "近三年九宫格位置", "历史九宫格"])
        history = [part.strip() for part in re.split(r"[/,，;；\n\r]+", str(history_value or "")) if part and part.strip()]
        leadership_detail = [
            [row.get(f"领导力{index}", f"领导力{index}"), row.get(f"得分{index}", "")]
            for index in range(1, 5)
            if row.get(f"领导力{index}") or row.get(f"得分{index}")
        ]
        annual_manager_comment = self._first_value(
            row,
            [
                "2025年度上级评语",
                "2025 年度上级评语",
                "2025年度Manager Comment",
                "2025 Manager Comment",
                "2025 Manager Comment",
                "manager comment",
                "Manager Comment",
                "上级评语",
                "年度上级评语",
                "年度绩效评价",
            ],
        )
        no_growth_warning = self._first_value(row, ["无成长预警"])
        calibrated_no_growth_warning = self._first_value(row, ["校准后无成长预警", "无成长预警", "校准后无成长风险", "无成长风险"])
        return {
            "employeeId": self._row_employee_id(row),
            "name": self._first_value(row, ["姓名"]),
            "group": self._first_value(row, ["群体"]),
            "departmentPath": department_path,
            **department_levels,
            "level": self._first_value(row, ["职级"]),
            "age": self._first_value(row, ["年龄"]),
            "tenure": self._first_value(row, ["司龄"]),
            "performanceLatest": self._first_value(row, ["年度绩效"]),
            "performanceOriginal": self._first_value(row, ["年度绩效"]),
            "performanceCalibrated": self._first_value(row, ["校准后绩效", "年度绩效"]),
            "performanceBand": self._first_value(row, ["绩效等级"]),
            "potentialBand": self._first_value(row, ["潜能等级"]),
            "potentialScore": self._first_value(row, ["潜能综合得分"]),
            "importedOriginalGrid": original_grid,
            "meetingBaselineGrid": current_grid,
            "gridOriginal": original_grid,
            "gridCurrent": current_grid,
            "hasCalibratedGrid": has_calibrated_grid,
            "gridOriginalLabel": self._first_value(row, ["九宫格位置", "原九宫格位置"]),
            "gridCurrentLabel": calibrated_grid_value,
            "gridHistory": history_value,
            "grid2025": self._parse_grid(history[0], None) if len(history) > 0 else None,
            "grid2024": self._parse_grid(history[1], None) if len(history) > 1 else None,
            "grid2023": self._parse_grid(history[2], None) if len(history) > 2 else None,
            "aiThinking": self._first_value(row, ["AI思维"]),
            "aiApplication": self._first_value(row, ["AI应用"]),
            "aiTalentTag": self._first_value(row, ["AI人才标签"]),
            "aiTalentTagCalibrated": self._first_value(row, ["校准后AI人才标签"]),
            "aiAbilityOriginal": self._first_value(row, ["AI人才标签"]),
            "aiAbilityCalibrated": self._first_value(row, ["校准后AI人才标签", "AI人才标签"]),
            "noGrowthWarningOriginal": no_growth_warning,
            "noGrowthWarningCalibrated": calibrated_no_growth_warning,
            "noGrowthWarning": calibrated_no_growth_warning,
            "incentive": self._first_value(row, ["激励(可多选)", "激励"]),
            "incentives": self._first_value(row, ["激励(可多选)", "激励"]),
            "incentivesOriginal": self._first_value(row, ["激励(可多选)", "激励"]),
            "developmentAdvice": self._first_value(row, ["发展建议(可多选)", "发展建议"]),
            "developmentAdviceOriginal": self._first_value(row, ["发展建议(可多选)", "发展建议"]),
            "managerComment2025": annual_manager_comment,
            "annualPerformanceReview": annual_manager_comment,
            "supervisorAdjustments": supervisor_adjustments,
            "hasSupervisorAdjustment": bool(supervisor_adjustments),
            "reviewNote": self._first_value(row, ["综合备注"]),
            "workflowStep": self._first_value(row, ["当前步骤"]),
            "workflowOwner": self._first_value(row, ["当前执行人"]),
            "professionalAbility": {
                "total": self._first_value(row, ["知识及技能"]),
                "detail": [
                    ["知识及技能", self._first_value(row, ["知识及技能"])],
                    ["干部品质", self._first_value(row, ["干部品质"])],
                    ["创新", self._first_value(row, ["创新", "鍒涙柊"])],
                    ["进取", self._first_value(row, ["进取", "杩涘彇"])],
                    ["分享", self._first_value(row, ["分享", "鍒嗕韩"])],
                    ["尊重", self._first_value(row, ["尊重", "灏婇噸"])],
                ],
            },
            "growthMindset": {
                "total": "",
                "detail": [
                    ["成就欲", row.get("成就欲", "")],
                    ["韧性", row.get("韧性", "")],
                    ["谦逊好学", row.get("谦逊好学", "")],
                ],
            },
            "leadership": {
                "total": "",
                "detail": leadership_detail,
            },
            "raw": row,
        }

    def _excel_rows(self, excel_path):
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            rows = sheet.iter_rows(values_only=True)
            headers = [str(value).strip() if value is not None else "" for value in next(rows)]
            parsed = []
            for row in rows:
                item = {headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index]}
                if any(value not in (None, "") for value in item.values()):
                    parsed.append(item)
            return headers, parsed
        finally:
            workbook.close()

    def _pick_row_value(self, row, names):
        for name in names:
            value = row.get(name)
            if value not in (None, ""):
                return value
        normalized = {re.sub(r"[\s_\-]+", "", str(key or "")).lower(): value for key, value in row.items()}
        for name in names:
            value = normalized.get(re.sub(r"[\s_\-]+", "", str(name or "")).lower())
            if value not in (None, ""):
                return value
        return ""

    def _text(self, value):
        if value is None:
            return ""
        text = str(value).strip()
        if re.fullmatch(r"\d+\.0", text):
            return text[:-2]
        return text

    def _excel_column_index(self, cell_ref):
        match = re.match(r"([A-Z]+)", str(cell_ref or ""))
        if not match:
            return 0
        index = 0
        for char in match.group(1):
            index = index * 26 + ord(char) - ord("A") + 1
        return index

    def _xlsx_shared_strings(self, archive):
        if "xl/sharedStrings.xml" not in archive.namelist():
            return []
        root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
        strings = []
        for item in root:
            if item.tag.endswith("si"):
                strings.append("".join(item.itertext()))
        return strings

    def _xlsx_sheet_paths(self, archive):
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        rel_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        rels = {
            rel.attrib.get("Id"): rel.attrib.get("Target", "")
            for rel in rel_root
            if rel.tag.endswith("Relationship")
        }
        paths = []
        for sheet in workbook_root.iter():
            if not sheet.tag.endswith("sheet"):
                continue
            relationship_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            target = rels.get(relationship_id, "")
            if target:
                paths.append("xl/" + target.lstrip("/"))
        return paths or [name for name in archive.namelist() if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")]

    def _xlsx_cell_value(self, cell, shared_strings):
        cell_type = cell.attrib.get("t", "")
        value_node = next((child for child in cell if child.tag.endswith("v")), None)
        if cell_type == "inlineStr":
            return self._text("".join(cell.itertext()))
        if value_node is None or value_node.text is None:
            return ""
        value = value_node.text
        if cell_type == "s":
            try:
                return shared_strings[int(value)]
            except (ValueError, IndexError):
                return ""
        return value

    def _xlsx_rows_from_xml(self, source_path: Path):
        rows_by_sheet = []
        with ZipFile(source_path) as archive:
            shared_strings = self._xlsx_shared_strings(archive)
            for sheet_path in self._xlsx_sheet_paths(archive):
                if sheet_path not in archive.namelist():
                    continue
                root = ET.fromstring(archive.read(sheet_path))
                rows = []
                for row_node in root.iter():
                    if not row_node.tag.endswith("row"):
                        continue
                    cells = {}
                    for cell in row_node:
                        if not cell.tag.endswith("c"):
                            continue
                        column = self._excel_column_index(cell.attrib.get("r", ""))
                        if column:
                            cells[column] = self._xlsx_cell_value(cell, shared_strings)
                    if cells:
                        rows.append([cells.get(column, "") for column in range(1, max(cells) + 1)])
                if rows:
                    rows_by_sheet.append(rows)
        return rows_by_sheet

    def import_employee_roster_excel(self, source_path: Path, filename: str):
        return self.import_employee_roster_excels([(source_path, filename)])

    def _employee_roster_items(self, source_path: Path):
        items = []
        header_markers = {
            "直线经理", "职务", "姓名", "工号", "员工ID",
            "POIdEmpAdmin-ExportName", "POldEmpAdmin-ExportName", "OIdJobPost", "OIdJobPost-ExportName",
        }
        normalized_markers = {re.sub(r"[\s_\-]+", "", marker).lower() for marker in header_markers}

        def parse_sheet_rows(sheet_rows):
            sheet_items = []
            if not sheet_rows:
                return sheet_items
            header_row = None
            headers = []
            for row_index in range(min(len(sheet_rows), 20)):
                candidate = [self._text(value) for value in sheet_rows[row_index]]
                candidate_norm = {re.sub(r"[\s_\-]+", "", header).lower() for header in candidate if header}
                if candidate_norm.intersection(normalized_markers):
                    header_row = row_index
                    headers = candidate
                    if {"姓名", "工号"}.issubset(set(candidate)) or {"姓名", "直线经理"}.issubset(set(candidate)):
                        break
            if header_row is None:
                return sheet_items
            for row_values in sheet_rows[header_row + 1:]:
                row = {
                    headers[index]: row_values[index] if index < len(row_values) else ""
                    for index in range(len(headers))
                    if headers[index]
                }
                if not any(value not in (None, "") for value in row.values()):
                    continue
                name = self._text(self._pick_row_value(row, ["姓名", "Name", "ExportName", "EmpName", "OIdEmp-ExportName", "OIdEmpAdmin-ExportName", "FName", "员工姓名"]))
                employee_id = self._normalize_employee_id(self._text(self._pick_row_value(row, ["工号", "员工ID", "employeeId", "EmpId", "OIdEmp", "OIdEmpAdmin", "PersonId", "UserId", "Code", "EmployeeCode", "JobNumber"])))
                if not name and not employee_id:
                    continue
                levels = [self._text(self._pick_row_value(row, [key])) for key in ("一级组织", "二级组织", "三级组织", "四级组织", "五级组织")]
                levels = [value for value in levels if value]
                title = self._text(self._pick_row_value(row, ["职位", "岗位", "Title", "JobTitle", "岗位名称", "OIdJobPosition_Name"]))
                job_post = self._text(self._pick_row_value(row, ["职务", "OIdJobPost", "OIdJobPost-ExportName", "岗位", "职位"]))
                sheet_items.append(
                    {
                        "employeeId": employee_id,
                        "name": name,
                        "manager": self._text(self._pick_row_value(row, ["直线经理", "POIdEmpAdmin-ExportName", "POldEmpAdmin-ExportName", "直接上级", "上级姓名", "Manager", "ManagerName"])),
                        "managerEmail": self._text(self._pick_row_value(row, ["直线经理邮箱", "上级邮箱", "ManagerEmail", "POIdEmpAdmin-Email"])),
                        "title": title,
                        "level": self._text(self._pick_row_value(row, ["职级", "Level", "Rank", "OIdJobLevel_Name"])),
                        "sequence": job_post,
                        "departmentPath": self._text(self._pick_row_value(row, ["组织全称", "部门全称", "部门路径", "OrganizationFullName", "DeptFullName", "LookupPrefix_OIdDepartment_POIdOrgAdminNameTreePath", "parent_Name"])) or "/".join(levels),
                        "一级组织": self._text(self._pick_row_value(row, ["一级组织"])),
                        "二级组织": self._text(self._pick_row_value(row, ["二级组织"])),
                        "三级组织": self._text(self._pick_row_value(row, ["三级组织"])),
                        "四级组织": self._text(self._pick_row_value(row, ["四级组织"])),
                        "五级组织": self._text(self._pick_row_value(row, ["五级组织"])),
                        "age": self._text(self._pick_row_value(row, ["年龄", "Age"])),
                        "tenure": self._text(self._pick_row_value(row, ["累计司龄", "司龄", "Tenure"])),
                        "email": self._text(self._pick_row_value(row, ["企业邮箱", "邮箱", "Email"])),
                    }
                )
            return sheet_items

        try:
            for sheet_rows in self._xlsx_rows_from_xml(source_path):
                items.extend(parse_sheet_rows(sheet_rows))
        except Exception:
            items = []

        if not items:
            workbook = load_workbook(source_path, data_only=True)
            try:
                for sheet in workbook.worksheets:
                    sheet_rows = [
                        [sheet.cell(row=row_number, column=column).value for column in range(1, sheet.max_column + 1)]
                        for row_number in range(1, sheet.max_row + 1)
                    ]
                    items.extend(parse_sheet_rows(sheet_rows))
            finally:
                workbook.close()

        if not items:
            raise ValueError("未能从 Excel 中识别员工花名册，请确认文件包含姓名、工号、直线经理或组织字段。")
        return items

    def import_employee_roster_excels(self, files):
        items = []
        source_files = []
        for source_path, filename in files:
            items.extend(self._employee_roster_items(source_path))
            source_files.append(Path(filename).name or "employee_roster.xlsx")
        name_counts = {}
        for item in items:
            name_counts[item["name"]] = name_counts.get(item["name"], 0) + 1
        payload = {
            "sourceFile": "?".join(source_files),
            "updatedAt": datetime.now(timezone.utc).isoformat(),
            "rows": len(items),
            "byEmployeeId": {item["employeeId"]: item for item in items if item["employeeId"]},
            "byName": {item["name"]: item for item in items if item["name"] and name_counts[item["name"]] == 1},
        }
        self.employee_map_path.parent.mkdir(parents=True, exist_ok=True)
        self.employee_map_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        self._clear_cache()
        return {"rows": len(items), "files": source_files, "jsonPath": str(self.employee_map_path)}

    def import_review_excel(self, source_path: Path, filename: str):
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        safe_name = Path(filename).name or f"talent_review_{timestamp}.xlsx"
        source_bytes = Path(source_path).read_bytes()
        self._clear_imported_files(self.review_source_dir, ["*.xlsx", "*.json"])
        excel_path = self.review_source_dir / safe_name
        excel_path.write_bytes(source_bytes)
        headers, rows = self._excel_rows(excel_path)
        records = [self._review_row_to_record(row) for row in rows]
        json_path = self.review_source_dir / f"{excel_path.stem}.json"
        json_path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
        self.overrides_path.write_text(
            json.dumps(
                {
                    "version": "2026-talent-calibration-v1",
                    "updatedAt": datetime.now(timezone.utc).isoformat(),
                    "changes": [],
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        self._clear_cache()
        return {"rows": len(records), "excelPath": str(excel_path), "jsonPath": str(json_path), "headers": headers}

    def import_profiles_json(self, data, filename: str):
        profiles = data.get("profiles") if isinstance(data, dict) and "profiles" in data else data
        if not isinstance(profiles, list):
            raise ValueError("人才档案 JSON 应为数组，或包含 profiles 数组。")
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        safe_name = Path(filename).name or f"people_profiles_{timestamp}.json"
        self.profile_source_dir.mkdir(parents=True, exist_ok=True)
        target = self.profile_source_dir / safe_name
        target.write_text(json.dumps(profiles, ensure_ascii=False, indent=2), encoding="utf-8")
        self._clear_cache()
        return {"rows": len(profiles), "jsonPath": str(target)}

    def overrides(self):
        signature = self._path_signature(self.overrides_path)

        def load_overrides():
            data = self._read_json(self.overrides_path, {"changes": []})
            if isinstance(data, list):
                return {"changes": data}
            return {"changes": data.get("changes", [])}

        return self._cached("overrides", signature, load_overrides)

    def people(self):
        signature = (
            self._path_signature(self._review_source_path()),
            self._files_signature(self._profile_source_dirs(), "*.json", self.profile_path),
            self._path_signature(self.employee_map_path),
            self._path_signature(self.overrides_path),
        )

        def load_people():
            profiles = self.profiles()
            profile_by_id = {self._normalize_employee_id(item.get("employeeId")): item for item in profiles if item.get("employeeId")}
            change_by_id = {item.get("employeeId"): item for item in self.overrides()["changes"]}
            merged = []
            for result in self.review_results():
                result = self._enrich_with_employee_info(result)
                employee_id = self._normalize_employee_id(result.get("employeeId"))
                change = change_by_id.get(employee_id, {})
                grid_original = int(result.get("meetingBaselineGrid", result.get("gridCurrent", result.get("gridOriginal", 5))) or 5)
                grid_current = int(change.get("calibratedGrid", result.get("gridCurrent", grid_original)) or grid_original)
                ai_calibrated = change.get("aiAbilityCalibrated", result.get("aiAbilityCalibrated", result.get("aiAbilityOriginal", "")))
                no_growth_calibrated = change.get(
                    "noGrowthWarningCalibrated",
                    result.get("noGrowthWarningCalibrated", result.get("noGrowthWarningOriginal", "")),
                )
                incentives = change.get("incentives", result.get("incentives", result.get("incentivesOriginal", "")))
                development_advice = change.get("developmentAdvice", result.get("developmentAdvice", result.get("developmentAdviceOriginal", "")))
                profile = self._profile_for_result(result, profiles, profile_by_id)
                if profile:
                    result = {**profile, **result, "profile": profile}
                supervisor_adjustments = result.get("supervisorAdjustments", [])
                if not supervisor_adjustments and isinstance(result.get("raw"), dict):
                    supervisor_adjustments = self._ledger_adjustments(result["raw"], result.get("importedOriginalGrid"), result.get("meetingBaselineGrid"))
                merged.append(
                    {
                        **result,
                        "employeeId": employee_id,
                        "gridOriginal": grid_original,
                        "gridCurrent": grid_current,
                        "aiAbilityCalibrated": ai_calibrated,
                        "noGrowthWarningCalibrated": no_growth_calibrated,
                        "noGrowthWarning": no_growth_calibrated,
                        "incentives": incentives,
                        "developmentAdvice": development_advice,
                        "supervisorAdjustments": supervisor_adjustments,
                        "hasSupervisorAdjustment": bool(supervisor_adjustments),
                        "adjustment": change,
                        "profile": profile,
                    }
                )
            return merged

        return self._cached("people", signature, load_people)

    def _filter_by_department(self, items, department):
        """按一级组织或完整部门路径筛选人员。"""
        if not department:
            return items
        dept_lower = department.strip().lower()
        filtered = []
        for item in items:
            # 优先从 profile 中取 departmentPath。
            profile = item.get("profile") if isinstance(item, dict) else None
            if not profile and isinstance(item, dict):
                profile = item
            dp = (profile.get("departmentPath") or profile.get("departmentPathRaw") or "") if isinstance(profile, dict) else ""
            # 匹配一级部门或完整路径包含关系。
            top_dept = dp.split(">")[0].strip() if ">" in dp else dp.strip()
            if dept_lower in top_dept.lower() or dept_lower in dp.lower():
                filtered.append(item)
        return filtered

    @staticmethod
    def _slim_person(p):
        """精简单条人员数据，只保留 AI 分析需要的字段。"""
        return {
            "employeeId": p.get("employeeId", ""),
            "name": p.get("name", ""),
            "departmentPath": (p.get("profile") or {}).get("departmentPath", "") if isinstance(p.get("profile"), dict) else "",
            "gridOriginal": p.get("gridOriginal"),
            "gridCurrent": p.get("gridCurrent"),
            "gridOriginalLabel": p.get("gridOriginalLabel", ""),
            "gridCurrentLabel": p.get("gridCurrentLabel", ""),
            "nineBoxPosition": p.get("nineBoxPosition"),
            "performanceScore": p.get("performanceScore") or p.get("performance"),
            "potentialScore": p.get("potentialScore") or p.get("potential"),
            "calibratedPosition": p.get("calibratedPosition"),
            "reviewComment": (p.get("reviewResult") or {}).get("comment", "") if isinstance(p.get("reviewResult"), dict) else "",
        }

    def analysis_context(self, department=None):
        signature = (
            department or "",
            self._path_signature(self._review_source_path()),
            self._files_signature(self._profile_source_dirs(), "*.json", self.profile_path),
            self._path_signature(self.employee_map_path),
            self._path_signature(self.overrides_path),
        )
        cache_key = ("analysis_context", department or "")
        if not hasattr(self, "_cache"):
            self._cache = {}
            self._cache_lock = threading.RLock()
        with self._cache_lock:
            entry = self._cache.get(cache_key)
            if entry and entry["signature"] == signature:
                return entry["value"]
        people_list = self.people()
        total_people = len(people_list)
        review_results = self.review_results()
        profiles = self.profiles()
        if department:
            people_ids = {p["employeeId"] for p in self._filter_by_department(people_list, department) if p.get("employeeId")}
            people_list = [p for p in people_list if p.get("employeeId") in people_ids]
            review_results = [r for r in review_results if r.get("employeeId") in people_ids]
            profiles = [p for p in profiles if p.get("employeeId") in people_ids]
        # 精简数据：每条人员记录只保留关键字段，降低上下文体积。
        slim_people = [self._slim_person(p) for p in people_list]
        # 精简 profile：只保留关键 HR 字段。
        slim_profiles = []
        key_profile_keys = {"employeeId","name","departmentPath","positionName","workAge","talentType","riskFlag"}
        for p in profiles:
            if isinstance(p, dict):
                slim_profiles.append({k: p.get(k) for k in key_profile_keys if k in p})
        context = {
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "filteredDepartment": department or "（全部）",
            "totalPeopleBeforeFilter": total_people,
            "totalPeopleAfterFilter": len(slim_people),
            "reviewResults": review_results,
            "profiles": slim_profiles[:200],
            "overrides": self.overrides(),
            "people": slim_people,
        }
        with self._cache_lock:
            self._cache[cache_key] = {"signature": signature, "value": context}
        return context

    def talent_pools(self):
        payload = self._read_json(self.talent_pool_path, {"pools": []})
        pools = payload.get("pools", []) if isinstance(payload, dict) else []
        normalized = []
        seen = set()
        for pool in pools:
            if not isinstance(pool, dict):
                continue
            name = str(pool.get("name", "")).strip()
            if not name or name in seen:
                continue
            members = []
            member_seen = set()
            for member in pool.get("members", []):
                member_name = str(member).strip()
                if member_name and member_name not in member_seen:
                    members.append(member_name)
                    member_seen.add(member_name)
            normalized.append({"name": name, "members": members})
            seen.add(name)
        return {"pools": normalized}

    def save_talent_pools(self, pools):
        normalized = []
        seen = set()
        for pool in pools or []:
            if not isinstance(pool, dict):
                continue
            name = str(pool.get("name", "")).strip()
            if not name or name in seen:
                continue
            raw_members = pool.get("members", [])
            if isinstance(raw_members, str):
                raw_members = re.split(r"[;；,\n\r]+", raw_members)
            members = []
            member_seen = set()
            for member in raw_members:
                member_name = str(member).strip()
                if member_name and member_name not in member_seen:
                    members.append(member_name)
                    member_seen.add(member_name)
            normalized.append({"name": name, "members": members})
            seen.add(name)
        payload = {
            "updatedAt": datetime.now(timezone.utc).isoformat(),
            "pools": normalized,
        }
        self.talent_pool_path.parent.mkdir(parents=True, exist_ok=True)
        self.talent_pool_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return payload

    def save_overrides(self, changes):
        timestamp = datetime.now(timezone.utc).isoformat()
        normalized = []
        review_by_id = {item.get("employeeId"): item for item in self.review_results()}
        for change in changes:
            employee_id = change.get("employeeId")
            if employee_id not in review_by_id:
                continue
            source = review_by_id[employee_id]
            original_grid = self._parse_grid(source.get("meetingBaselineGrid", source.get("gridCurrent", source.get("gridOriginal"))), 5)
            calibrated_grid = int(change.get("calibratedGrid", original_grid))
            original_ai = source.get("aiAbilityCalibrated", source.get("aiAbilityOriginal", ""))
            calibrated_ai = change.get("aiAbilityCalibrated", original_ai)
            original_growth = source.get("noGrowthWarningCalibrated", source.get("noGrowthWarningOriginal", ""))
            calibrated_growth = change.get("noGrowthWarningCalibrated", original_growth)
            original_incentives = self._normalize_multi_value(source.get("incentives", source.get("incentivesOriginal", "")))
            calibrated_incentives = self._normalize_multi_value(change.get("incentives", original_incentives))
            original_development = self._normalize_multi_value(source.get("developmentAdvice", source.get("developmentAdviceOriginal", "")))
            calibrated_development = self._normalize_multi_value(change.get("developmentAdvice", original_development))
            if (
                calibrated_grid == original_grid
                and calibrated_ai == original_ai
                and calibrated_growth == original_growth
                and calibrated_incentives == original_incentives
                and calibrated_development == original_development
                and not change.get("reason")
            ):
                continue
            normalized.append(
                {
                    "employeeId": employee_id,
                    "name": source.get("name", change.get("name", "")),
                    "originalGrid": original_grid,
                    "calibratedGrid": calibrated_grid,
                    "aiAbilityOriginal": source.get("aiAbilityOriginal", ""),
                    "aiAbilityCalibrated": calibrated_ai,
                    "noGrowthWarningOriginal": source.get("noGrowthWarningOriginal", ""),
                    "noGrowthWarningCalibrated": calibrated_growth,
                    "incentivesOriginal": source.get("incentivesOriginal", source.get("incentives", "")),
                    "incentives": calibrated_incentives,
                    "developmentAdviceOriginal": source.get("developmentAdviceOriginal", source.get("developmentAdvice", "")),
                    "developmentAdvice": calibrated_development,
                    "reason": change.get("reason", ""),
                    "updatedBy": change.get("updatedBy", "local-user"),
                    "updatedAt": change.get("updatedAt", timestamp),
                }
            )
        payload = {
            "version": "2026-talent-calibration-v1",
            "updatedAt": timestamp,
            "changes": normalized,
        }
        self.overrides_path.parent.mkdir(parents=True, exist_ok=True)
        self.overrides_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        stale_calibrated_path = self.review_source_dir / "calibrated_latest.json"
        if stale_calibrated_path.exists():
            stale_calibrated_path.unlink()
        self._clear_cache()
        return payload

    def _has_calibration_difference(self, person):
        original_grid = person.get("gridOriginal", person.get("meetingBaselineGrid"))
        current_grid = person.get("gridCurrent", original_grid)
        original_grid = int(original_grid) if original_grid not in (None, "") else None
        current_grid = int(current_grid) if current_grid not in (None, "") else None
        original_ai = person.get("aiAbilityOriginal", "")
        current_ai = person.get("aiAbilityCalibrated", original_ai)
        original_growth = person.get("noGrowthWarningOriginal", "")
        current_growth = person.get("noGrowthWarningCalibrated", original_growth)
        original_incentives = self._normalize_multi_value(person.get("incentivesOriginal", ""))
        current_incentives = self._normalize_multi_value(person.get("incentives", original_incentives))
        original_development = self._normalize_multi_value(person.get("developmentAdviceOriginal", ""))
        current_development = self._normalize_multi_value(person.get("developmentAdvice", original_development))
        reason = person.get("adjustment", {}).get("reason", "")
        return (
            current_grid != original_grid
            or current_ai != original_ai
            or current_growth != original_growth
            or current_incentives != original_incentives
            or current_development != original_development
            or bool(reason)
        )

    def _grid_export_label(self, value):
        grid = self._parse_grid(value, None)
        return GRID_LABELS.get(grid, str(value or ""))

    def _score_detail_value(self, person, section, label):
        details = person.get(section, {}).get("detail", []) if isinstance(person.get(section), dict) else []
        for item in details:
            if isinstance(item, (list, tuple)) and len(item) >= 2 and str(item[0]) == label:
                return item[1]
        return ""

    def _leadership_score_values(self, person):
        details = person.get("leadership", {}).get("detail", []) if isinstance(person.get("leadership"), dict) else []
        values = []
        for item in details:
            if isinstance(item, (list, tuple)) and len(item) >= 2:
                values.append(item[1])
        return values[:4] + [""] * max(0, 4 - len(values))

    def _employee_number_for_export(self, person):
        profile = person.get("profile") if isinstance(person.get("profile"), dict) else {}
        for value in (
            profile.get("employeeId"),
            person.get("employeeNumber"),
            person.get("workNumber"),
            person.get("工号"),
            person.get("raw", {}).get("工号") if isinstance(person.get("raw"), dict) else "",
        ):
            normalized = self._normalize_employee_id(value)
            if normalized:
                return normalized
        return self._normalize_employee_id(person.get("employeeId"))

    def _level_number(self, person):
        for value in (
            person.get("level"),
            person.get("profile", {}).get("level") if isinstance(person.get("profile"), dict) else "",
            person.get("raw", {}).get("职级") if isinstance(person.get("raw"), dict) else "",
        ):
            match = re.search(r"(\d+(?:\.\d+)?)", str(value or ""))
            if match:
                try:
                    return float(match.group(1))
                except ValueError:
                    return None
        return None

    def _leader_scores_allowed(self, person):
        level = self._level_number(person)
        return level is not None and level > 6

    def _performance_export_value(self, person):
        raw = person.get("raw") if isinstance(person.get("raw"), dict) else {}
        return person.get("performanceCalibrated") or person.get("performanceLatest") or person.get("performanceOriginal") or raw.get("年度绩效", "")

    def _required_export_reason(self, person):
        reason = str(person.get("adjustment", {}).get("reason", "")).strip()
        if reason:
            return reason
        return "无校准调整，沿用原盘点结果"

    def _template_export_row(self, person):
        raw = person.get("raw") if isinstance(person.get("raw"), dict) else {}
        leadership_values = self._leadership_score_values(person) if self._leader_scores_allowed(person) else ["", "", "", ""]
        original_grid = person.get("gridOriginal", person.get("meetingBaselineGrid", raw.get("九宫格位置", "")))
        current_grid = person.get("gridCurrent", original_grid)
        return {
            "*工号": self._employee_number_for_export(person),
            "*姓名": person.get("name", ""),
            "群体": person.get("group", raw.get("群体", "")),
            "年度绩效": person.get("performanceLatest", raw.get("年度绩效", "")),
            "*校准后绩效": self._performance_export_value(person),
            "创新": self._score_detail_value(person, "professionalAbility", "创新") or raw.get("创新", ""),
            "进取": self._score_detail_value(person, "professionalAbility", "进取") or raw.get("进取", ""),
            "分享": self._score_detail_value(person, "professionalAbility", "分享") or raw.get("分享", ""),
            "尊重": self._score_detail_value(person, "professionalAbility", "尊重") or raw.get("尊重", ""),
            "*知识及技能": self._score_detail_value(person, "professionalAbility", "知识及技能") or raw.get("知识及技能", ""),
            "干部品质": self._score_detail_value(person, "professionalAbility", "干部品质") or raw.get("干部品质", ""),
            "领导力1得分": leadership_values[0],
            "领导力2得分": leadership_values[1],
            "领导力3得分": leadership_values[2],
            "领导力4得分": leadership_values[3],
            "*成就欲": self._score_detail_value(person, "growthMindset", "成就欲") or raw.get("成就欲", ""),
            "*韧性": self._score_detail_value(person, "growthMindset", "韧性") or raw.get("韧性", ""),
            "*谦逊好学": self._score_detail_value(person, "growthMindset", "谦逊好学") or raw.get("谦逊好学", ""),
            "*潜能综合得分": person.get("potentialScore", raw.get("潜能综合得分", "")),
            "*潜能等级": person.get("potentialBand", raw.get("潜能等级", "")),
            "*绩效等级": person.get("performanceBand", raw.get("绩效等级", "")),
            "*九宫格位置": self._grid_export_label(original_grid),
            "*校准后九宫格位置": self._grid_export_label(current_grid),
            "*校准后AI人才标签": person.get("aiAbilityCalibrated", person.get("aiAbilityOriginal", "")),
            "*校准后无成长预警": person.get("noGrowthWarningCalibrated", person.get("noGrowthWarningOriginal", "")),
            "*激励": person.get("incentives", raw.get("激励(可多选)", "")),
            "*发展建议": person.get("developmentAdvice", raw.get("发展建议(可多选)", "")),
            "综合备注": person.get("reviewNote", raw.get("综合备注", "")),
            "*校准原因": self._required_export_reason(person),
        }

    def export_calibrated_excel(self, differences_only=False):
        template_path = self.data_dir / "templates" / "盘点结果导入模板.xlsx"
        if not template_path.exists():
            raise FileNotFoundError("未找到盘点结果导入模板，请确认 data/templates/盘点结果导入模板.xlsx 存在。")
        workbook = load_workbook(template_path)
        sheet = workbook[workbook.sheetnames[0]]
        headers = [str(cell.value or "").strip() for cell in sheet[2]]
        header_columns = {header: index + 1 for index, header in enumerate(headers) if header}
        people = [person for person in self.people() if not differences_only or self._has_calibration_difference(person)]
        template_row = 3
        if sheet.max_row < template_row:
            sheet.insert_rows(template_row)
        template_styles = {}
        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=template_row, column=column)
            template_styles[column] = {
                "style": copy(cell._style),
                "number_format": cell.number_format,
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
                "protection": copy(cell.protection),
            }
        target_max_row = max(sheet.max_row, len(people) + 2)
        for row_index in range(3, target_max_row + 1):
            for column in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_index, column=column)
                cell.value = None
                styles = template_styles[column]
                cell._style = copy(styles["style"])
                cell.number_format = styles["number_format"]
                cell.font = copy(styles["font"])
                cell.fill = copy(styles["fill"])
                cell.border = copy(styles["border"])
                cell.alignment = copy(styles["alignment"])
                cell.protection = copy(styles["protection"])
        for row_index, person in enumerate(people, start=3):
            row = self._template_export_row(person)
            for header, value in row.items():
                column = header_columns.get(header)
                if column:
                    sheet.cell(row=row_index, column=column, value=value)

        export_dir = self.data_dir / "exports"
        export_dir.mkdir(parents=True, exist_ok=True)
        prefix = "calibration-differences-template" if differences_only else "calibrated-review-template"
        output_path = export_dir / f"{prefix}-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        workbook.save(output_path)
        return output_path

    def export_calibration_differences(self):
        return self.export_calibrated_excel(differences_only=True)
