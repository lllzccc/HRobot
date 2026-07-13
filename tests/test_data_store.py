import base64
import json
import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch
from urllib.request import urlopen
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook

from server import DataStore, Handler, is_newer_version


class DataStoreTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.data_dir = Path(self.tmp.name)
        (self.data_dir / "talent_review_2026.json").write_text(
            json.dumps(
                [
                    {
                        "employeeId": "E001",
                        "name": "张三",
                        "gridOriginal": 5,
                        "performanceBand": "中",
                        "potentialBand": "中",
                    },
                    {
                        "employeeId": "E002",
                        "name": "李四",
                        "gridOriginal": 8,
                        "performanceBand": "高",
                        "potentialBand": "中",
                    },
                ],
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        (self.data_dir / "people_profiles.json").write_text(
            json.dumps(
                [
                    {
                        "employeeId": "E001",
                        "departmentPath": "运营部 > 数据分析组",
                        "level": "P7",
                        "title": "数据分析师",
                        "manager": "王经理",
                    }
                ],
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        (self.data_dir / "calibration_overrides.json").write_text(
            json.dumps({"changes": []}, ensure_ascii=False),
            encoding="utf-8",
        )

    def tearDown(self):
        self.tmp.cleanup()

    def test_people_merge_review_result_with_read_only_profile_detail(self):
        store = DataStore(self.data_dir)

        people = store.people()

        self.assertEqual(people[0]["employeeId"], "E001")
        self.assertEqual(people[0]["gridOriginal"], 5)
        self.assertEqual(people[0]["gridCurrent"], 5)
        self.assertEqual(people[0]["profile"]["level"], "P7")
        self.assertEqual(people[1]["profile"], {})

    def test_people_data_layer_combines_current_profile_and_project_record(self):
        store = DataStore(self.data_dir)

        context = store.people_data().analysis_context()
        records = {item["employeeId"]: item for item in context["people"] if item.get("employeeId")}

        self.assertEqual(context["model"], "employee-centered-v1")
        self.assertIn("E001", records)
        self.assertEqual(records["E001"]["current"]["level"], "P7")
        self.assertEqual(records["E001"]["projects"]["talentReview"]["gridOriginal"], 5)
        self.assertEqual(context["dataSources"]["currentProfiles"]["kind"], "daily_snapshot")
        self.assertEqual(context["dataSources"]["talentReview"]["kind"], "project_snapshot")

    def test_overrides_change_current_grid_without_mutating_sources(self):
        store = DataStore(self.data_dir)

        store.save_overrides(
            [
                {
                    "employeeId": "E001",
                    "calibratedGrid": 9,
                    "reason": "校准会确认为高潜核心人才",
                    "updatedBy": "local-user",
                }
            ]
        )

        people = store.people()
        self.assertEqual(people[0]["gridOriginal"], 5)
        self.assertEqual(people[0]["gridCurrent"], 9)
        self.assertFalse((self.data_dir / "review_results" / "calibrated_latest.json").exists())
        source = json.loads((self.data_dir / "talent_review_2026.json").read_text(encoding="utf-8"))
        self.assertEqual(source[0]["gridOriginal"], 5)

    def test_review_results_ignore_stale_calibrated_snapshot_without_fallback(self):
        review_dir = self.data_dir / "review_results"
        review_dir.mkdir()
        (review_dir / "calibrated_latest.json").write_text(
            json.dumps(
                [
                    {
                        "employeeId": "STALE",
                        "name": "旧快照",
                        "gridOriginal": 9,
                    }
                ],
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        store = DataStore(self.data_dir)

        reviews = store.review_results()

        self.assertEqual(reviews, [])

    def test_empty_review_source_folder_does_not_fallback_to_legacy_review_json(self):
        (self.data_dir / "review_results").mkdir()
        store = DataStore(self.data_dir)

        self.assertEqual(store.review_results(), [])
        self.assertEqual(store.people(), [])

    def test_empty_profile_source_folder_does_not_fallback_to_legacy_profiles(self):
        (self.data_dir / "talent_profiles").mkdir()
        store = DataStore(self.data_dir)

        self.assertEqual(store.profiles(), [])

    def test_overrides_save_ai_and_growth_calibration(self):
        store = DataStore(self.data_dir)

        payload = store.save_overrides(
            [
                {
                    "employeeId": "E001",
                    "calibratedGrid": 5,
                    "aiAbilityCalibrated": "AI KOL",
                    "noGrowthWarningCalibrated": "是",
                    "updatedBy": "local-user",
                }
            ]
        )
        people = store.people()

        self.assertEqual(payload["changes"][0]["aiAbilityCalibrated"], "AI KOL")
        self.assertEqual(payload["changes"][0]["noGrowthWarningCalibrated"], "是")
        self.assertEqual(people[0]["aiAbilityCalibrated"], "AI KOL")
        self.assertEqual(people[0]["noGrowthWarningCalibrated"], "是")

    def test_reads_json_files_with_utf8_bom(self):
        (self.data_dir / "calibration_overrides.json").write_text(
            "\ufeff" + json.dumps(
                {
                    "changes": [
                        {
                            "employeeId": "E002",
                            "calibratedGrid": 9,
                            "reason": "PowerShell wrote a BOM",
                        }
                    ]
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        store = DataStore(self.data_dir)

        people = store.people()

        self.assertEqual(people[1]["gridCurrent"], 9)

    def test_version_compare_handles_semver_values(self):
        self.assertTrue(is_newer_version("0.1.1", "0.1.0"))
        self.assertTrue(is_newer_version("v0.2.0", "0.1.9"))
        self.assertFalse(is_newer_version("0.1.0", "0.1.0"))
        self.assertFalse(is_newer_version("0.1.0", "0.1.1"))

    def test_update_config_is_local_user_setting(self):
        store = DataStore(self.data_dir)

        payload = store.save_update_config({"sourcePath": r"D:\WeCom\HROBOT-release"})

        self.assertEqual(payload["sourcePath"], r"D:\WeCom\HROBOT-release")
        saved = json.loads((self.data_dir / "update_config.json").read_text(encoding="utf-8"))
        self.assertEqual(saved["sourcePath"], r"D:\WeCom\HROBOT-release")

    def test_update_status_reads_shared_drive_release_manifest(self):
        release_dir = self.data_dir / "shared-release"
        release_dir.mkdir()
        (release_dir / "HrobotSetup.exe").write_bytes(b"stub")
        (release_dir / "release.json").write_text(
            json.dumps(
                {
                    "app": "Hrobot",
                    "version": "9.9.9",
                    "installer": "HrobotSetup.exe",
                    "notes": "Test release",
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        store = DataStore(self.data_dir)

        status = store.update_status(str(release_dir))

        self.assertTrue(status["updateAvailable"])
        self.assertEqual(status["latest"]["installer"], "HrobotSetup.exe")
        self.assertEqual(status["latest"]["notes"], "Test release")

    def test_update_status_reads_release_manifest_url(self):
        class FakeResponse:
            headers = {"Content-Type": "application/json"}

            def __enter__(self):
                return self

            def __exit__(self, *args):
                return False

            def read(self, *_args):
                return json.dumps(
                    {
                        "version": "9.9.9",
                        "installer": "HrobotSetup.exe",
                        "notes": "URL release",
                    }
                ).encode("utf-8")

        store = DataStore(self.data_dir)
        with patch("server.urlopen", return_value=FakeResponse()):
            status = store.update_status("https://example.com/hrobot/release.json")

        self.assertTrue(status["updateAvailable"])
        self.assertEqual(status["sourceType"], "url")
        self.assertEqual(status["latest"]["installerPath"], "https://example.com/hrobot/HrobotSetup.exe")

    def test_update_status_reads_github_latest_release_assets(self):
        class FakeResponse:
            headers = {"Content-Type": "application/json"}

            def __enter__(self):
                return self

            def __exit__(self, *args):
                return False

            def read(self, *_args):
                return json.dumps(
                    {
                        "tag_name": "v9.9.9",
                        "html_url": "https://github.com/lllzccc/HRobot/releases/tag/v9.9.9",
                        "body": "GitHub release notes",
                        "published_at": "2026-07-02T10:00:00Z",
                        "assets": [
                            {
                                "name": "hrobot-mac-source-9.9.9.zip",
                                "browser_download_url": "https://github.com/lllzccc/HRobot/releases/download/v9.9.9/hrobot-mac-source-9.9.9.zip",
                            },
                            {
                                "name": "hrobot-win-9.9.9.exe",
                                "browser_download_url": "https://github.com/lllzccc/HRobot/releases/download/v9.9.9/hrobot-win-9.9.9.exe",
                            },
                        ],
                    }
                ).encode("utf-8")

        store = DataStore(self.data_dir)
        with patch("server.urlopen", return_value=FakeResponse()):
            status = store.update_status()

        self.assertTrue(status["updateAvailable"])
        self.assertEqual(status["sourceType"], "github")
        self.assertEqual(status["latest"]["installer"], "hrobot-win-9.9.9.exe")
        self.assertEqual(status["latest"]["notes"], "GitHub release notes")
        self.assertTrue(status["latest"]["canAutoInstall"])

    def test_update_status_selects_matching_macos_dmg_architecture(self):
        class FakeResponse:
            headers = {"Content-Type": "application/json"}

            def __enter__(self):
                return self

            def __exit__(self, *args):
                return False

            def read(self, *_args):
                return json.dumps(
                    {
                        "tag_name": "v9.9.9",
                        "html_url": "https://github.com/lllzccc/HRobot/releases/tag/v9.9.9",
                        "body": "macOS release notes",
                        "published_at": "2026-07-13T10:00:00Z",
                        "assets": [
                            {
                                "name": "hrobot-mac-x64-9.9.9.dmg",
                                "browser_download_url": "https://example.com/hrobot-mac-x64-9.9.9.dmg",
                            },
                            {
                                "name": "hrobot-mac-arm64-9.9.9.dmg",
                                "browser_download_url": "https://example.com/hrobot-mac-arm64-9.9.9.dmg",
                            },
                            {
                                "name": "hrobot-win-9.9.9.exe",
                                "browser_download_url": "https://example.com/hrobot-win-9.9.9.exe",
                            },
                        ],
                    }
                ).encode("utf-8")

        store = DataStore(self.data_dir)
        with (
            patch("server.urlopen", return_value=FakeResponse()),
            patch("server.app_platform", return_value="mac"),
            patch("server.app_architecture", return_value="arm64"),
            patch("server.sys.frozen", True, create=True),
        ):
            status = store.update_status()

        self.assertEqual(status["latest"]["installer"], "hrobot-mac-arm64-9.9.9.dmg")
        self.assertTrue(status["latest"]["canAutoInstall"])
        self.assertEqual(status["packages"]["mac-arm64"]["name"], "hrobot-mac-arm64-9.9.9.dmg")

    def test_update_status_rejects_html_share_page_url(self):
        class FakeHtmlResponse:
            headers = {"Content-Type": "text/html"}

            def __enter__(self):
                return self

            def __exit__(self, *args):
                return False

            def read(self, *_args):
                return b"<!DOCTYPE html><html></html>"

        store = DataStore(self.data_dir)
        with patch("server.urlopen", return_value=FakeHtmlResponse()):
            with self.assertRaisesRegex(ValueError, "release.json"):
                store.update_status("https://drive.weixin.qq.com/s?k=test")

    def test_review_results_prefer_review_result_folder(self):
        source_dir = self.data_dir / "review_results"
        source_dir.mkdir()
        (source_dir / "2026人才盘点结果.json").write_text(
            json.dumps(
                [
                    {
                        "employeeId": "R001",
                        "name": "真实盘点员工",
                        "gridOriginal": 6,
                    }
                ],
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        store = DataStore(self.data_dir)

        reviews = store.review_results()

        self.assertEqual(reviews[0]["employeeId"], "R001")

    def test_profiles_prefer_talent_profile_folder(self):
        source_dir = self.data_dir / "talent_profiles"
        source_dir.mkdir()
        (source_dir / "运营部人才档案快照.json").write_text(
            json.dumps(
                [
                    {
                        "employeeId": "P001",
                        "name": "真实档案员工",
                        "level": "P7",
                    }
                ],
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        store = DataStore(self.data_dir)

        profiles = store.profiles()

        self.assertEqual(profiles[0]["employeeId"], "P001")

    def test_import_profiles_json_keeps_multiple_department_profile_files(self):
        store = DataStore(self.data_dir)
        store.import_profiles_json([{"employeeId": "TECH", "name": "Tech"}], "tech.json")
        store.import_profiles_json([{"employeeId": "GAME", "name": "Game"}], "studio.json")

        profiles = store.profiles()
        imported_files = list((self.data_dir / "talent_profiles").glob("*.json"))

        self.assertCountEqual([item["employeeId"] for item in profiles], ["TECH", "GAME"])
        self.assertEqual(len(imported_files), 2)

    def test_profiles_merge_multiple_files_and_newer_duplicate_wins(self):
        store = DataStore(self.data_dir)
        store.import_profiles_json(
            [
                {"employeeId": "E001", "name": "Old Name", "level": "P6"},
                {"employeeId": "E002", "name": "Tech", "level": "P7"},
            ],
            "tech.json",
        )
        store.import_profiles_json(
            [
                {"employeeId": "E001", "name": "New Name", "level": "P8"},
                {"employeeId": "E003", "name": "Studio", "level": "P5"},
            ],
            "studio.json",
        )
        os.utime(self.data_dir / "talent_profiles" / "tech.json", (1, 1))
        os.utime(self.data_dir / "talent_profiles" / "studio.json", (2, 2))

        profiles = store.profiles()

        self.assertCountEqual([item["employeeId"] for item in profiles], ["E001", "E002", "E003"])
        self.assertEqual(next(item for item in profiles if item["employeeId"] == "E001")["level"], "P8")

    def test_analysis_context_uses_review_profiles_and_overrides(self):
        store = DataStore(self.data_dir)
        store.save_overrides(
            [
                {
                    "employeeId": "E001",
                    "calibratedGrid": 9,
                    "reason": "calibration reason",
                    "updatedBy": "local-user",
                }
            ]
        )

        context = store.analysis_context()

        self.assertEqual(len(context["reviewResults"]), 2)
        self.assertEqual(len(context["profiles"]), 1)
        self.assertEqual(context["overrides"]["changes"][0]["employeeId"], "E001")
        self.assertEqual(context["people"][0]["gridCurrent"], 9)

    def test_report_assets_and_generated_report_are_stored(self):
        store = DataStore(self.data_dir)

        skill = store.import_report_asset("skill", "framework.md", "按组织分析".encode("utf-8"))
        material = store.import_report_asset("material", "meeting.txt", "补充材料".encode("utf-8"))
        assets = store.report_assets()
        report = store.save_generated_report("## 人才盘点结论\n内容", "生成报告")

        self.assertEqual(skill["filename"], "framework.md")
        self.assertEqual(material["filename"], "meeting.txt")
        self.assertEqual(assets["skills"][0]["filename"], "framework.md")
        self.assertEqual(assets["materials"][0]["filename"], "meeting.txt")
        self.assertIn("人才盘点结论", store.generated_report()["content"])
        self.assertEqual(report["source"], "ai")
        self.assertTrue(Path(report["mdPath"]).exists())
        self.assertIn("人才盘点结论", Path(report["mdPath"]).read_text(encoding="utf-8"))
        history = json.loads((self.data_dir / "report_generation" / "generated_reports.json").read_text(encoding="utf-8"))
        self.assertFalse(Path(history[0]["mdPath"]).is_absolute())

    def test_home_memos_are_saved_in_json_file(self):
        store = DataStore(self.data_dir)

        payload = store.save_home_memo("2026-06-10", "先确认校准差异人员")

        self.assertEqual(payload["records"][0]["date"], "2026-06-10")
        self.assertEqual(payload["records"][0]["text"], "先确认校准差异人员")
        raw = json.loads((self.data_dir / "home_memo.json").read_text(encoding="utf-8"))
        self.assertEqual(raw["records"][0]["text"], "先确认校准差异人员")

    def test_design_prompt_config_has_defaults_and_reference_folder(self):
        store = DataStore(self.data_dir)

        config = store.design_prompt_config()

        self.assertIn("template", config)
        self.assertTrue((self.data_dir / "design_center" / "references").exists())
        self.assertTrue((self.data_dir / "design_center" / "references" / "README.md").exists())

    def test_design_prompt_config_can_be_saved(self):
        store = DataStore(self.data_dir)

        config = store.save_design_prompt_config(
            {
                "basePrompt": "生成招聘海报",
                "brandRequirements": "使用公司标准色",
                "customRequirements": "必须预留 logo 区域",
                "referenceInstructions": "优先参考 references 目录",
                "template": "{basePrompt}\n{customRequirements}\n需求：{requirement}",
            }
        )

        self.assertEqual(config["basePrompt"], "生成招聘海报")
        self.assertEqual(config["customRequirements"], "必须预留 logo 区域")
        raw = json.loads((self.data_dir / "design_center" / "poster_prompt_config.json").read_text(encoding="utf-8"))
        self.assertEqual(raw["referenceInstructions"], "优先参考 references 目录")

    def test_agent_center_imports_web_project_zip_as_card(self):
        store = DataStore(self.data_dir)
        zip_path = self.data_dir / "demo-agent.zip"
        with ZipFile(zip_path, "w") as archive:
            archive.writestr("demo/index.html", "<!doctype html><title>Demo Web Agent</title><h1>Hello</h1>")
            archive.writestr("demo/README.md", "# Demo Web Agent\n用于查看导入后的独立 Web 功能。")
            archive.writestr("demo/app.js", "console.log('demo')")

        payload = store.import_agent_project_zip("demo-agent.zip", zip_path.read_bytes())
        project = payload["project"]
        listed = store.agent_projects()

        self.assertEqual(project["name"], "Demo Web Agent")
        self.assertIn("用于查看导入后的独立 Web 功能", project["description"])
        self.assertTrue(project["entryUrl"].endswith("/demo/index.html"))
        self.assertEqual(project["runtime"], "static-web")
        self.assertEqual(project["serverEntry"], "")
        self.assertEqual(project["analysis"]["runtime"], "static-web")
        self.assertGreaterEqual(project["analysis"]["confidence"], 0.7)
        self.assertTrue(project["environment"]["canOpen"])
        self.assertEqual(project["environment"]["label"], "无需运行时")
        self.assertEqual(project["fileCount"], 3)
        self.assertEqual(listed["count"], 1)
        self.assertTrue(listed["projects"][0]["environment"]["canOpen"])
        self.assertTrue((self.data_dir / "agent_center" / "manifest.json").exists())
        self.assertFalse((self.data_dir / "agent_center" / "zips" / project["sourceZip"]).exists())

    def test_agent_center_detects_vite_project_and_requires_dependencies(self):
        store = DataStore(self.data_dir)
        zip_path = self.data_dir / "vite-agent.zip"
        with ZipFile(zip_path, "w") as archive:
            archive.writestr("web/index.html", "<!doctype html><title>Vite Agent</title><div id='root'></div>")
            archive.writestr(
                "web/package.json",
                json.dumps(
                    {
                        "scripts": {"dev": "vite"},
                        "dependencies": {"@vitejs/plugin-react": "latest"},
                        "devDependencies": {"vite": "latest"},
                    }
                ),
            )
            archive.writestr("web/src/main.js", "console.log('vite')")

        project = store.import_agent_project_zip("vite-agent.zip", zip_path.read_bytes())["project"]

        self.assertEqual(project["runtime"], "node-vite")
        self.assertEqual(project["analysis"]["startCwd"], "web")
        self.assertIn("npm run dev", project["analysis"]["startCommand"])
        self.assertTrue(project["analysis"]["requiresInstall"])
        self.assertFalse(project["environment"]["canOpen"])
        self.assertIn(project["environment"]["label"], {"缺 Node.js", "依赖未安装"})

    def test_agent_center_blocks_node_project_when_runtime_missing(self):
        store = DataStore(self.data_dir)
        zip_path = self.data_dir / "node-ready-agent.zip"
        with ZipFile(zip_path, "w") as archive:
            archive.writestr("web/index.html", "<!doctype html><title>Node Ready Agent</title><div id='root'></div>")
            archive.writestr(
                "web/package.json",
                json.dumps({"scripts": {"dev": "vite"}, "devDependencies": {"vite": "latest"}}),
            )
            archive.writestr("web/node_modules/.package-lock.json", "{}")

        with patch.object(DataStore, "_portable_runtime_path", return_value=""), patch(
            "app.modules.agent_center.store.shutil.which",
            return_value=None,
        ):
            project = store.import_agent_project_zip("node-ready-agent.zip", zip_path.read_bytes())["project"]
            with self.assertRaisesRegex(ValueError, "Node"):
                store.open_agent_project(project["id"])

        self.assertEqual(project["runtime"], "node-vite")
        self.assertFalse(project["analysis"]["requiresInstall"])
        self.assertFalse(project["environment"]["canOpen"])
        self.assertEqual(project["environment"]["label"], "缺 Node.js")

    def test_agent_center_uses_ai_fallback_for_uncertain_node_project(self):
        store = DataStore(self.data_dir)
        zip_path = self.data_dir / "uncertain-agent.zip"
        with ZipFile(zip_path, "w") as archive:
            archive.writestr("client/index.html", "<!doctype html><title>Hybrid Agent</title>")
            archive.writestr("client/package.json", json.dumps({"scripts": {"start": "node server.js"}}))
            archive.writestr("client/server.js", "console.log('server')")

        def fake_ai(messages):
            return {
                "configured": True,
                "message": json.dumps(
                    {
                        "runtime": "node-server",
                        "entry": "client/index.html",
                        "startCwd": "client",
                        "startCommand": "npm run start",
                        "requiresInstall": True,
                        "confidence": 0.9,
                        "description": "AI 判断为前后端混合 Node Web 功能。",
                        "reason": "package.json 的 start 脚本指向 server.js。",
                        "flags": ["需要确认依赖安装。"],
                    },
                    ensure_ascii=False,
                ),
            }

        project = store.import_agent_project_zip("uncertain-agent.zip", zip_path.read_bytes(), ai_provider=fake_ai)["project"]

        self.assertEqual(project["runtime"], "node-server")
        self.assertEqual(project["analysis"]["source"], "rules+ai")
        self.assertEqual(project["analysis"]["ai"]["status"], "ok")
        self.assertIn("混合 Node Web", project["description"])

    def test_agent_center_delete_removes_project_folder_and_source_zip(self):
        store = DataStore(self.data_dir)
        zip_path = self.data_dir / "delete-agent.zip"
        with ZipFile(zip_path, "w") as archive:
            archive.writestr("index.html", "<!doctype html><title>Delete Agent</title>")
            archive.writestr("data/app.db", "db")
        payload = store.import_agent_project_zip("delete-agent.zip", zip_path.read_bytes())
        project = payload["project"]
        project_dir = Path(project["folderPath"])
        source_zip = self.data_dir / "agent_center" / "zips" / project["sourceZip"]

        result = store.delete_agent_project(project["id"])

        self.assertEqual(result["count"], 0)
        self.assertFalse(project_dir.exists())
        self.assertFalse(source_zip.exists())

    def test_agent_center_auto_imports_zip_drop_folder(self):
        store = DataStore(self.data_dir)
        zip_dir = self.data_dir / "agent_center" / "zips"
        zip_dir.mkdir(parents=True)
        with ZipFile(zip_dir / "drop-agent.zip", "w") as archive:
            archive.writestr("site/index.html", "<!doctype html><title>Drop Agent</title>")

        payload = store.agent_projects()

        self.assertEqual(payload["count"], 1)
        self.assertEqual(payload["projects"][0]["name"], "Drop Agent")
        self.assertTrue(payload["projects"][0]["entryUrl"].endswith("/site/index.html"))
        self.assertFalse((zip_dir / "drop-agent.zip").exists())

    def test_agent_center_opens_server_project_on_isolated_port(self):
        store = DataStore(self.data_dir)
        zip_path = self.data_dir / "server-agent.zip"
        server_code = """
import argparse
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer

parser = argparse.ArgumentParser()
parser.add_argument("--host", default="127.0.0.1")
parser.add_argument("--port", type=int, required=True)
args = parser.parse_args()
ThreadingHTTPServer((args.host, args.port), SimpleHTTPRequestHandler).serve_forever()
""".strip()
        with ZipFile(zip_path, "w") as archive:
            archive.writestr("app/index.html", "<!doctype html><title>Server Agent</title><h1>isolated</h1>")
            archive.writestr("app/server.py", server_code)

        project = store.import_agent_project_zip("server-agent.zip", zip_path.read_bytes())["project"]
        opened = store.open_agent_project(project["id"])

        try:
            self.assertEqual(project["runtime"], "python-server")
            self.assertEqual(project["serverEntry"], "app/server.py")
            self.assertGreaterEqual(opened["port"], 8768)
            self.assertTrue(opened["url"].endswith("/index.html"))
            body = urlopen(opened["url"], timeout=3).read().decode("utf-8")
            self.assertIn("isolated", body)
        finally:
            store.delete_agent_project(project["id"])

    def test_generated_report_can_be_deleted_with_markdown_file(self):
        store = DataStore(self.data_dir)
        report = store.save_generated_report("## 待删除报告\n内容", "生成报告")

        result = store.delete_generated_report(report["id"])

        self.assertTrue(result["deleted"])
        self.assertEqual(store.generated_report_list(), [])
        self.assertFalse(Path(report["mdPath"]).exists())

    def test_generated_report_html_is_created_on_demand(self):
        store = DataStore(self.data_dir)
        report = store.save_generated_report("## Demo Report\nBody text", "generate report")

        self.assertEqual(report["contentFormat"], "markdown")
        self.assertFalse(report.get("htmlPath"))

        html_report = store.generate_report_html(report["id"])
        html_path = Path(html_report["htmlPath"])

        self.assertTrue(html_path.exists())
        self.assertIn("<!doctype html>", html_report["htmlContent"].lower())
        self.assertIn("Demo Report", html_report["htmlContent"])
        self.assertTrue(store.generated_report_list()[0]["hasHtml"])

        store.delete_generated_report(report["id"])
        self.assertFalse(html_path.exists())

    def test_360_generated_report_html_uses_fixed_dashboard_template(self):
        store = DataStore(self.data_dir)
        report = store.save_generated_report(
            "# 莫绵 360°评估通俗解读分析\n\n# 一句话总结\n\n整体表现稳定。\n\n# 开放性反馈的关键信息\n\n| 来源 | 内容 |\n|---|---|\n| 上级 | 结果导向 |\n",
            "生成莫绵360报告",
            "360",
        )

        html_report = store.generate_report_html(report["id"])

        self.assertIn("quadrant-panel", html_report["htmlContent"])
        self.assertIn("角色评分对比", html_report["htmlContent"])
        self.assertIn("metric-card", html_report["htmlContent"])
        self.assertIn("<table>", html_report["htmlContent"])

    def test_delete_imported_report_and_data_files(self):
        store = DataStore(self.data_dir)
        store.import_report_asset("skill", "framework.md", "按组织分析".encode("utf-8"))
        store.import_profiles_json([{"employeeId": "P001", "name": "Profile"}], "profiles.json")
        (self.data_dir / "employee_manager_map.json").write_text("{}", encoding="utf-8")

        skill_result = store.delete_imported_file("report-skill", "framework.md")
        profile_result = store.delete_imported_file("profile", "profiles.json")
        roster_result = store.delete_imported_file("employee-roster", "employee_manager_map.json")

        self.assertTrue(skill_result["deleted"])
        self.assertTrue(profile_result["deleted"])
        self.assertTrue(roster_result["deleted"])
        self.assertEqual(store.report_assets()["skills"], [])
        self.assertFalse((self.data_dir / "talent_profiles" / "profiles.json").exists())
        self.assertFalse((self.data_dir / "employee_manager_map.json").exists())

    def test_delete_imported_file_rejects_path_traversal(self):
        store = DataStore(self.data_dir)

        with self.assertRaises(ValueError):
            store.delete_imported_file("profile", "../people_profiles.json")

    def test_html_generated_report_is_detected_and_cleaned(self):
        store = DataStore(self.data_dir)

        report = store.save_generated_report(
            "```html\n<!doctype html><html><head><title>HTML报告</title></head><body><h1>HTML报告</h1><p>内容</p></body></html>\n```",
            "生成 HTML 报告",
            "talent-review",
        )

        self.assertEqual(report["contentFormat"], "html")
        self.assertEqual(report["title"], "HTML报告")
        self.assertNotIn("```", store.generated_report()["content"])
        self.assertTrue(Path(report["mdPath"]).exists())
        self.assertIn("HTML报告", Path(report["mdPath"]).read_text(encoding="utf-8"))

    def test_legacy_html_report_history_is_renderable(self):
        store = DataStore(self.data_dir)
        history_path = self.data_dir / "report_generation" / "generated_reports.json"
        history_path.parent.mkdir(parents=True, exist_ok=True)
        history_path.write_text(
            json.dumps(
                [
                    {
                        "id": "legacy-html",
                        "title": "旧HTML报告",
                        "content": "```html\n<!doctype html><html><body><h1>旧HTML报告</h1></body></html>\n```",
                        "updatedAt": "2026-06-09T12:00:00+08:00",
                        "source": "ai",
                    }
                ],
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

        report = store.generated_report("legacy-html")

        self.assertEqual(report["contentFormat"], "html")
        self.assertTrue(report["content"].lstrip().lower().startswith("<!doctype html>"))
        self.assertNotIn("```", report["content"])

    def test_report_settings_are_exposed_as_builtin_abilities(self):
        setting_dir = self.data_dir / "report_generation" / "settings"
        setting_dir.mkdir(parents=True, exist_ok=True)
        (setting_dir / "360报告设定说明.md").write_text("卡片标题必须为 xxx360报告解读", encoding="utf-8")
        store = DataStore(self.data_dir)

        abilities = store.report_abilities()["abilities"]
        ability = next(item for item in abilities if item.get("presetId") == "360")

        self.assertEqual(ability["sourceType"], "builtin")
        self.assertEqual(ability["settingFile"], "360报告设定说明.md")
        self.assertIn("卡片标题必须为 xxx360报告解读", ability["descriptionMd"])

        context = store.report_asset_context("360", selected_abilities=[ability["id"]])
        self.assertNotIn("报告设定说明", [item["type"] for item in context])
        ability_context = next(item for item in context if item["type"] == "能力中心")
        self.assertIn("卡片标题必须为 xxx360报告解读", ability_context["content"])

    def test_empty_selected_abilities_do_not_include_all_abilities(self):
        store = DataStore(self.data_dir)

        context = store.report_asset_context("talent-review", selected_abilities=[])

        self.assertNotIn("能力中心", [item["type"] for item in context])

    def test_saving_builtin_ability_returns_full_ability_list(self):
        store = DataStore(self.data_dir)

        payload = store.save_report_ability(
            "360报告能力",
            description_md="# 自定义 360 能力说明",
            ability_id="report-preset-ability-360",
        )

        self.assertEqual(len(payload["abilities"]), 3)
        self.assertEqual(payload["ability"]["id"], "report-preset-ability-360")
        self.assertIn("自定义 360", payload["ability"]["descriptionMd"])
        self.assertIn("report-preset-ability-org-diagnosis", [item["id"] for item in payload["abilities"]])
        self.assertIn("report-preset-ability-talent-review", [item["id"] for item in payload["abilities"]])

    def test_report_abilities_can_exceed_five_items(self):
        store = DataStore(self.data_dir)

        payload = {}
        for index in range(1, 7):
            payload = store.save_report_ability(
                f"自定义能力 {index}",
                description_md=f"第 {index} 个自定义能力说明",
            )

        custom_abilities = [
            item for item in payload["abilities"]
            if item.get("sourceType") == "custom"
        ]
        self.assertEqual(len(custom_abilities), 6)
        self.assertEqual(len(payload["abilities"]), 9)
        self.assertIn("自定义能力 6", [item["name"] for item in custom_abilities])

    def test_360_report_context_rejects_unreadable_pdf_material(self):
        store = DataStore(self.data_dir)
        store.import_report_asset("material", "person-360.pdf", b"%PDF-1.4\ninvalid pdf bytes")

        with self.assertRaises(ValueError) as error:
            store.report_asset_context("360")

        self.assertIn("360 PDF", str(error.exception))

    def test_360_report_context_filters_materials_by_target_name(self):
        store = DataStore(self.data_dir)
        store.import_report_asset("material", "莫绵-个人报告核心版.md", "莫绵材料".encode("utf-8"))
        store.import_report_asset("material", "管炜-个人报告核心版.md", "管炜核心版".encode("utf-8"))
        store.import_report_asset("material", "管炜-执行层领导力360评估-个人报告标准版.md", "管炜标准版".encode("utf-8"))

        context = store.report_asset_context("360", target_name="管炜")
        filenames = [item["filename"] for item in context if item["type"] == "其他分析材料"]

        self.assertCountEqual(
            filenames,
            ["管炜-个人报告核心版.md", "管炜-执行层领导力360评估-个人报告标准版.md"],
        )

    def test_report_context_uses_only_explicitly_selected_assets(self):
        store = DataStore(self.data_dir)
        store.import_report_asset("skill", "framework-a.md", "框架 A".encode("utf-8"))
        store.import_report_asset("skill", "framework-b.md", "框架 B".encode("utf-8"))
        store.import_report_asset("material", "管炜-核心版.md", "管炜材料".encode("utf-8"))
        store.import_report_asset("material", "莫绵-核心版.md", "莫绵材料".encode("utf-8"))

        context = store.report_asset_context(
            "360",
            target_name="管炜",
            selected_skills=["framework-b.md"],
            selected_materials=["莫绵-核心版.md"],
        )
        filenames = [item["filename"] for item in context]

        self.assertIn("framework-b.md", filenames)
        self.assertIn("莫绵-核心版.md", filenames)
        self.assertNotIn("framework-a.md", filenames)
        self.assertNotIn("管炜-核心版.md", filenames)

    def test_report_context_rejects_missing_selected_asset(self):
        store = DataStore(self.data_dir)

        with self.assertRaises(FileNotFoundError):
            store.report_asset_context(selected_materials=["missing.pdf"])

    def test_report_messages_forward_explicit_asset_selection(self):
        store = DataStore(self.data_dir)
        store.import_report_asset("skill", "selected-skill.md", "选中框架".encode("utf-8"))
        store.import_report_asset("skill", "ignored-skill.md", "未选框架".encode("utf-8"))
        store.import_report_asset("material", "selected-material.md", "选中材料".encode("utf-8"))
        store.import_report_asset("material", "ignored-material.md", "未选材料".encode("utf-8"))
        handler = Handler.__new__(Handler)
        handler.store = store

        messages = handler._build_report_messages(
            "生成报告",
            report_type="talent-review",
            selected_assets={
                "skills": ["selected-skill.md"],
                "materials": ["selected-material.md"],
            },
        )
        asset_message = next(message["content"] for message in messages if message["content"].startswith("能力中心、导入 skill 和分析材料："))

        self.assertIn("selected-skill.md", asset_message)
        self.assertIn("selected-material.md", asset_message)
        self.assertNotIn("ignored-skill.md", asset_message)
        self.assertNotIn("ignored-material.md", asset_message)

    def test_360_pdf_summary_extracts_core_metrics(self):
        store = DataStore(self.data_dir)

        summary = store._extract_360_pdf_summary(
            "管炜-个人报告核心版.pdf",
            "排名：13 / 249\n他评：4.62\n自评：4.95\n核心结果：执行层领导力360评估\n上级：4.42  |  同事：4.70  |  下级：4.81\n文本反馈\n",
        )

        self.assertIn("排名：13 / 249", summary)
        self.assertIn("他评：4.62", summary)
        self.assertIn("自评：4.95", summary)
        self.assertIn("角色评分：上级：4.42；同事：4.70；下级：4.81", summary)
        self.assertIn("含开放性反馈/文本反馈原文", summary)

    def test_360_pdf_summary_extracts_standard_report_capabilities(self):
        store = DataStore(self.data_dir)

        summary = store._extract_360_pdf_summary(
            "管炜-个人报告标准版.pdf",
            "概况\n您在这次评估中得到的总分为：4.62分。\n他评分较高的能力：目标导向要事第一、务实正直、向外学习并转化；\n他评分较低的能力：路径清晰有根据、引入和培养高潜；\n各角色分差异较大的能力：路径清晰有根据、小步快跑快速迭代。\n开放性反馈\n",
        )

        self.assertIn("他评：4.62", summary)
        self.assertIn("高分能力：目标导向要事第一、务实正直、向外学习并转化", summary)
        self.assertIn("待发展能力：路径清晰有根据、引入和培养高潜", summary)
        self.assertIn("差异较大能力：路径清晰有根据、小步快跑快速迭代", summary)

    def test_360_report_card_title_and_intro_are_normalized(self):
        store = DataStore(self.data_dir)

        report = store.save_generated_report(
            "<!doctype html><html><head><title>莫绵360°评估通俗解读分析</title></head><body><h1>莫绵360°评估</h1></body></html>",
            "生成莫绵的360报告",
            "360",
        )

        self.assertEqual(report["title"], "莫绵360报告解读")
        self.assertEqual(report["intro"], "参考材料来源")

    def test_360_generated_report_html_extracts_flexible_summary_metrics(self):
        store = DataStore(self.data_dir)
        report = store.save_generated_report(
            "# 兰海涵360报告解读\n\n## 一句话总结\n\n整体表现积极。\n\n- 总体他评：4.56，自评：4.35，排名：21 / 249。\n- 上级评分：4.28。\n- 同事评分：4.53。\n- 下级评分：4.98。\n",
            "生成兰海涵的360报告",
            "360",
        )

        html_report = store.generate_report_html(report["id"])

        self.assertIn("4.56", html_report["htmlContent"])
        self.assertIn("21 / 249", html_report["htmlContent"])
        self.assertIn("4.28", html_report["htmlContent"])
        self.assertIn("4.53", html_report["htmlContent"])
        self.assertIn("4.98", html_report["htmlContent"])

    def test_360_report_person_name_strips_instruction_verbs(self):
        store = DataStore(self.data_dir)

        name = store._extract_360_report_person_name("", "生成管炜360报告解读")

        self.assertEqual(name, "管炜")

    def test_ai_config_groups_status_does_not_expose_api_keys(self):
        store = DataStore(self.data_dir)

        saved = store.save_ai_config(
            {
                "multimodal": {
                    "apiKey": "text-secret",
                    "baseUrl": "https://text.example.test/v1",
                    "model": "talent-model",
                },
                "image": {
                    "apiKey": "image-secret",
                    "baseUrl": "https://image.example.test/v1",
                    "model": "poster-model",
                },
            }
        )
        status = store.ai_config_status()
        raw = json.loads((self.data_dir / "ai_config.json").read_text(encoding="utf-8"))
        secrets = json.loads((self.data_dir / "local_ai_secrets.json").read_text(encoding="utf-8"))
        restarted_store = DataStore(self.data_dir)

        self.assertTrue(saved["multimodal"]["configured"])
        self.assertTrue(saved["image"]["configured"])
        self.assertEqual(status["multimodal"]["baseUrl"], "https://text.example.test/v1")
        self.assertEqual(status["multimodal"]["model"], "talent-model")
        self.assertEqual(status["image"]["baseUrl"], "https://image.example.test/v1")
        self.assertEqual(status["image"]["model"], "poster-model")
        self.assertNotIn("apiKey", status["multimodal"])
        self.assertNotIn("apiKey", status["image"])
        self.assertEqual(raw["multimodal"]["apiKey"], "")
        self.assertEqual(raw["image"]["apiKey"], "")
        self.assertEqual(secrets["multimodal"]["apiKey"], "text-secret")
        self.assertEqual(secrets["image"]["apiKey"], "image-secret")
        self.assertEqual(store.ai_config()["multimodal"]["apiKey"], "text-secret")
        self.assertEqual(store.ai_config()["image"]["apiKey"], "image-secret")
        self.assertEqual(restarted_store.ai_config()["multimodal"]["apiKey"], "text-secret")
        self.assertEqual(restarted_store.ai_config()["image"]["apiKey"], "image-secret")

    def test_data_source_config_status_does_not_expose_mcp_secret(self):
        store = DataStore(self.data_dir)

        status = store.save_mcp_config(
            {
                "url": "http://127.0.0.1:8000/mcp",
                "authHeaderName": "Authorization",
                "apiKey": "Bearer secret-token",
            }
        )

        self.assertTrue(status["authConfigured"])
        self.assertNotIn("secret-token", json.dumps(status, ensure_ascii=False))
        self.assertEqual(store.mcp_config()["headers"]["Authorization"], "Bearer secret-token")

    def test_local_data_source_scan_detects_talent_review_file(self):
        store = DataStore(self.data_dir)
        folder = self.data_dir / "local_sources"
        folder.mkdir()
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["工号", "姓名", "九宫格位置", "校准后九宫格位置", "年度绩效"])
        sheet.append(["E001", "张三", "5中坚力量", "7潜力之星", "A"])
        workbook.save(folder / "2025人才盘点结果.xlsx")

        payload = store.scan_local_data_sources(str(folder))

        self.assertEqual(payload["fileCount"], 1)
        self.assertEqual(payload["files"][0]["detectedType"], "talent_review")
        self.assertIn("九宫格位置", payload["files"][0]["headers"])
        self.assertIn("summary", payload)
        self.assertIn("九宫格", " ".join(payload["summary"]["analysisOpportunities"]))
        self.assertTrue(payload["summary"]["recommendations"])

    def test_ai_config_reads_legacy_flat_file_as_multimodal_config(self):
        (self.data_dir / "ai_config.json").write_text(
            json.dumps(
                {
                    "apiKey": "legacy-secret",
                    "baseUrl": "https://legacy.example.test/v1",
                    "model": "legacy-model",
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        store = DataStore(self.data_dir)

        config = store.ai_config()
        status = store.ai_config_status()

        self.assertEqual(config["multimodal"]["apiKey"], "legacy-secret")
        self.assertEqual(config["multimodal"]["baseUrl"], "https://legacy.example.test/v1")
        self.assertEqual(config["multimodal"]["model"], "legacy-model")
        self.assertEqual(config["image"]["apiKey"], "")
        self.assertTrue(status["multimodal"]["configured"])
        self.assertFalse(status["image"]["configured"])

        store.save_ai_config({"multimodal": {"baseUrl": "https://legacy.example.test/v1", "model": "legacy-model"}})
        raw = json.loads((self.data_dir / "ai_config.json").read_text(encoding="utf-8"))
        secrets = json.loads((self.data_dir / "local_ai_secrets.json").read_text(encoding="utf-8"))
        restarted_store = DataStore(self.data_dir)

        self.assertEqual(raw["multimodal"]["apiKey"], "")
        self.assertEqual(secrets["multimodal"]["apiKey"], "legacy-secret")
        self.assertEqual(restarted_store.ai_config()["multimodal"]["apiKey"], "legacy-secret")

    def test_intelligence_filters_current_items_by_channel_and_search(self):
        (self.data_dir / "intelligence.json").write_text(
            json.dumps(
                {
                    "updated_at": "2026-06-04T09:00:00+08:00",
                    "items": [
                        {
                            "title": "HR AI tool launch",
                            "summary": "Recruiting workflow agent",
                            "source": "News A",
                            "channel": "ai_hr",
                            "category": "tool",
                            "published_at": "2026-06-04",
                            "keywords": ["agent"],
                        },
                        {
                            "title": "Game studio org change",
                            "summary": "Producer team adjusted",
                            "source": "News B",
                            "channel": "game_org",
                            "category": "org_change",
                            "published_at": "2026-06-04",
                            "keywords": ["studio"],
                        },
                    ],
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        store = DataStore(self.data_dir)

        result = store.intelligence({"channel": ["ai_hr"], "search": ["agent"]})

        self.assertEqual(result["scope"], "current")
        self.assertEqual(result["updated_at"], "2026-06-04T09:00:00+08:00")
        self.assertEqual(result["count"], 1)
        self.assertEqual(result["items"][0]["title"], "HR AI tool launch")

    def test_intelligence_filters_history_by_date(self):
        (self.data_dir / "intelligence_history.json").write_text(
            json.dumps(
                {
                    "updated_at": "2026-06-04T10:00:00+08:00",
                    "items": [
                        {"title": "June item", "channel": "ai_hr", "category": "tool", "published_at": "2026-06-04"},
                        {"title": "May item", "channel": "ai_hr", "category": "tool", "published_at": "2026-05-30"},
                    ],
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        store = DataStore(self.data_dir)

        result = store.intelligence({"date": ["2026-05-30"]})

        self.assertEqual(result["scope"], "history")
        self.assertEqual(result["total"], 2)
        self.assertEqual(result["count"], 1)
        self.assertEqual(result["items"][0]["title"], "May item")

    def test_intelligence_config_defaults_and_save(self):
        store = DataStore(self.data_dir)

        default_config = store.intelligence_config()
        saved_config = store.save_intelligence_config(
            {
                "autoEnabled": False,
                "runAt": "10:00",
                "channel": "game_org",
                "source": "bing",
                "maxPerQuery": 5,
                "wechatFulltextLimit": 0,
                "allowUnverifiedWechat": True,
            }
        )

        self.assertTrue(default_config["autoEnabled"])
        self.assertEqual(default_config["runAt"], "10:00")
        self.assertFalse(saved_config["autoEnabled"])
        self.assertEqual(saved_config["channel"], "game_org")
        self.assertEqual(saved_config["source"], "bing")
        self.assertEqual(saved_config["maxPerQuery"], 5)
        self.assertEqual(saved_config["wechatFulltextLimit"], 0)
        self.assertTrue(saved_config["allowUnverifiedWechat"])

    def test_design_posters_empty_history_and_save_base64_image(self):
        store = DataStore(self.data_dir)
        image_b64 = base64.b64encode(b"fake-png-data").decode("ascii")

        self.assertEqual(store.design_posters()["items"], [])
        saved = store.save_design_poster(
            {
                "posterType": "referral",
                "style": "modern",
                "size": "1024x1024",
                "prompt": "Hire a TA expert",
                "model": "poster-model",
                "b64_json": image_b64,
            }
        )
        history = store.design_posters()
        image_path = self.data_dir / "design_center" / "posters" / Path(saved["imagePath"]).name

        self.assertEqual(saved["posterType"], "referral")
        self.assertTrue(image_path.exists())
        self.assertEqual(image_path.read_bytes(), b"fake-png-data")
        self.assertEqual(history["items"][0]["id"], saved["id"])
        self.assertEqual(history["items"][0]["imagePath"], saved["imagePath"])

    def test_talent_pools_save_normalizes_names_and_members(self):
        store = DataStore(self.data_dir)

        payload = store.save_talent_pools(
            [
                {"name": "核心策划", "members": "张三;李四;张三;"},
                {"name": "核心策划", "members": ["王五"]},
                {"name": "高潜技术", "members": [" 王五 ", "", "赵六"]},
            ]
        )

        self.assertEqual(
            payload["pools"],
            [
                {"name": "核心策划", "members": ["张三", "李四"]},
                {"name": "高潜技术", "members": ["王五", "赵六"]},
            ],
        )
        self.assertEqual(store.talent_pools()["pools"][0]["members"], ["张三", "李四"])

    def test_ai_request_payload_does_not_force_temperature(self):
        payload = Handler._ai_request_payload("gpt-5.5", [{"role": "user", "content": "hello"}])

        self.assertEqual(payload["model"], "gpt-5.5")
        self.assertEqual(payload["messages"][0]["content"], "hello")
        self.assertNotIn("temperature", payload)

    def test_import_review_excel_converts_to_json_for_app_usage(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.append(["群体", "姓名", "组织全称", "职级", "年度绩效", "潜能等级", "九宫格位置", "校准后九宫格位置"])
        sheet.append(["干部", "Alice", "示例集团/运营部", "P7", "B", "高", "7潜力之星", "9超级明星"])
        excel_path = self.data_dir / "review.xlsx"
        workbook.save(excel_path)
        store = DataStore(self.data_dir)

        result = store.import_review_excel(excel_path, "review.xlsx")
        reviews = store.review_results()

        self.assertEqual(result["rows"], 1)
        self.assertTrue(Path(result["jsonPath"]).exists())
        self.assertEqual(reviews[0]["name"], "Alice")
        self.assertEqual(reviews[0]["gridOriginal"], 7)
        self.assertEqual(reviews[0]["gridCurrent"], 9)
        self.assertEqual(reviews[0]["raw"]["校准后九宫格位置"], "9超级明星")

    def test_import_review_excel_overwrites_previous_review_import(self):
        store = DataStore(self.data_dir)
        for name, person in [("old.xlsx", "Old"), ("new.xlsx", "New")]:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["群体", "姓名", "组织全称", "职级", "九宫格位置", "校准后九宫格位置"])
            sheet.append(["干部", person, "示例集团-研发事业群/运营部", "P7", "7潜力之星", "7潜力之星"])
            excel_path = self.data_dir / name
            workbook.save(excel_path)
            store.import_review_excel(excel_path, name)

        reviews = store.review_results()
        imported_excels = list((self.data_dir / "review_results").glob("*.xlsx"))
        imported_json = list((self.data_dir / "review_results").glob("*.json"))

        self.assertEqual([item["name"] for item in reviews], ["New"])
        self.assertEqual(len(imported_excels), 1)
        self.assertEqual(len(imported_json), 1)

    def test_import_review_excel_resets_previous_calibration_overrides(self):
        store = DataStore(self.data_dir)
        store.save_overrides(
            [
                {
                    "employeeId": "E001",
                    "calibratedGrid": 9,
                    "reason": "previous cycle",
                }
            ]
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["群体", "姓名", "组织全称", "职级", "九宫格位置", "校准后九宫格位置"])
        sheet.append(["干部", "Alice", "示例集团/运营部", "P7", "7潜力之星", "7潜力之星"])
        excel_path = self.data_dir / "review.xlsx"
        workbook.save(excel_path)

        store.import_review_excel(excel_path, "review.xlsx")

        self.assertEqual(store.overrides()["changes"], [])

    def test_people_match_imported_profiles_by_name_and_department_when_ids_differ(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["群体", "姓名", "组织全称", "职级", "九宫格位置", "校准后九宫格位置"])
        sheet.append(["干部", "赵振飞", "示例集团-研发事业群/运营部", "M4", "8绩效之星", "8绩效之星"])
        excel_path = self.data_dir / "review.xlsx"
        workbook.save(excel_path)
        store = DataStore(self.data_dir)
        store.import_review_excel(excel_path, "review.xlsx")
        store.import_profiles_json(
            [
                {
                    "employeeId": "2789",
                    "name": "赵振飞",
                    "departmentPath": "运营部",
                    "departmentPathRaw": "运营部",
                    "title": "运营总监",
                }
            ],
            "profiles.json",
        )

        person = store.people()[0]

        self.assertEqual(person["employeeId"], "赵振飞|示例集团-研发事业群/运营部")
        self.assertEqual(person["profile"]["employeeId"], "2789")
        self.assertEqual(person["profile"]["title"], "运营总监")

    def test_import_review_excel_splits_department_levels(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["群体", "姓名", "组织全称", "职级", "九宫格位置", "校准后九宫格位置"])
        sheet.append(["干部", "Alice", "示例集团-研发事业群/孵化二部/技术架构组", "P7", "7潜力之星", "7潜力之星"])
        excel_path = self.data_dir / "review.xlsx"
        workbook.save(excel_path)
        store = DataStore(self.data_dir)

        store.import_review_excel(excel_path, "review.xlsx")
        review = store.review_results()[0]

        self.assertEqual(review["一级组织"], "示例集团-研发事业群")
        self.assertEqual(review["二级组织"], "孵化二部")
        self.assertEqual(review["三级组织"], "技术架构组")
        self.assertEqual(review["四级组织"], "")
        self.assertEqual(review["五级组织"], "")

    def test_export_calibrated_excel_updates_calibrated_grid_column(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["群体", "姓名", "组织全称", "职级", "九宫格位置", "校准后九宫格位置"])
        sheet.append(["干部", "Alice", "示例集团/运营部", "P7", "7潜力之星", "7潜力之星"])
        excel_path = self.data_dir / "review.xlsx"
        workbook.save(excel_path)
        store = DataStore(self.data_dir)
        store.import_review_excel(excel_path, "review.xlsx")
        employee_id = store.review_results()[0]["employeeId"]
        store.save_overrides(
            [
                {
                    "employeeId": employee_id,
                    "calibratedGrid": 9,
                    "reason": "calibrated in meeting",
                }
            ]
        )

        template_dir = self.data_dir / "templates"
        template_dir.mkdir()
        template = Workbook()
        template_sheet = template.active
        template_sheet.append(["盘点结果导入模板"])
        template_sheet.append(["群体", "*姓名", "*九宫格位置", "*校准后九宫格位置"])
        template_sheet.append(["", "", "", ""])
        template.save(template_dir / "盘点结果导入模板.xlsx")

        output_path = store.export_calibrated_excel()
        exported = load_workbook(output_path, data_only=True)
        row = [cell.value for cell in next(exported.active.iter_rows(min_row=3, max_row=3))]

        self.assertEqual(row[3], "9超级明星")

    def test_export_calibrated_excel_uses_roster_employee_id_for_synthetic_review_key(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["群体", "姓名", "组织全称", "职级", "年度绩效", "九宫格位置", "校准后九宫格位置"])
        sheet.append(["干部", "Alice", "示例集团/运营部", "P7", "B", "7潜力之星", "7潜力之星"])
        excel_path = self.data_dir / "review.xlsx"
        workbook.save(excel_path)
        store = DataStore(self.data_dir)
        store.import_review_excel(excel_path, "review.xlsx")
        (self.data_dir / "employee_manager_map.json").write_text(
            json.dumps(
                {
                    "byEmployeeId": {"1001": {"employeeId": "1001", "name": "Alice"}},
                    "byName": {"Alice": {"employeeId": "1001", "name": "Alice"}},
                }
            ),
            encoding="utf-8",
        )

        template_dir = self.data_dir / "templates"
        template_dir.mkdir()
        template = Workbook()
        template_sheet = template.active
        template_sheet.append(["盘点结果导入模板"])
        template_sheet.append(["*工号", "*姓名", "群体", "年度绩效", "*九宫格位置", "*校准后九宫格位置"])
        template_sheet.append(["", "", "", "", "", ""])
        template.save(template_dir / "盘点结果导入模板.xlsx")

        output_path = store.export_calibrated_excel()
        exported = load_workbook(output_path, data_only=True)
        row = [cell.value for cell in next(exported.active.iter_rows(min_row=3, max_row=3))]

        self.assertEqual(row[0], "1001")
        self.assertNotIn("|", row[0])

    def test_export_calibrated_excel_does_not_export_synthetic_review_key_as_employee_id(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["群体", "姓名", "组织全称", "职级", "年度绩效", "九宫格位置", "校准后九宫格位置"])
        sheet.append(["干部", "Alice", "示例集团/运营部", "P7", "B", "7潜力之星", "7潜力之星"])
        excel_path = self.data_dir / "review.xlsx"
        workbook.save(excel_path)
        store = DataStore(self.data_dir)
        store.import_review_excel(excel_path, "review.xlsx")

        template_dir = self.data_dir / "templates"
        template_dir.mkdir()
        template = Workbook()
        template_sheet = template.active
        template_sheet.append(["盘点结果导入模板"])
        template_sheet.append(["*工号", "*姓名", "群体", "年度绩效", "*九宫格位置", "*校准后九宫格位置"])
        template_sheet.append(["", "", "", "", "", ""])
        template.save(template_dir / "盘点结果导入模板.xlsx")

        output_path = store.export_calibrated_excel()
        exported = load_workbook(output_path, data_only=True)
        row = [cell.value for cell in next(exported.active.iter_rows(min_row=3, max_row=3))]

        self.assertIn(row[0], (None, ""))


class ProductionDataTests(unittest.TestCase):
    def test_runtime_data_uses_current_review_source_folder(self):
        data_dir = Path(__file__).resolve().parents[1] / "data"
        store = DataStore(data_dir)

        runtime_reviews = store.review_results()
        runtime_people = store.people()
        current_reviews = json.loads(store._review_source_path().read_text(encoding="utf-8-sig"))
        legacy_reviews = json.loads((data_dir / "talent_review_2026.json").read_text(encoding="utf-8-sig"))

        self.assertEqual(len(runtime_reviews), len(current_reviews))
        self.assertEqual(len(runtime_people), len(runtime_reviews))
        self.assertNotEqual(len(runtime_reviews), len(legacy_reviews))


class ReportShellTests(unittest.TestCase):
    def _report_shell_source(self):
        root = Path(__file__).resolve().parents[1]
        return "\n".join(
            [
                (root / "index.html").read_text(encoding="utf-8"),
                (root / "static" / "js" / "app.js").read_text(encoding="utf-8"),
                (root / "static" / "css" / "app.css").read_text(encoding="utf-8"),
                (root / "static" / "modules" / "talent-review" / "talent-review.js").read_text(encoding="utf-8"),
            ]
        )

    def test_index_has_platform_pages_query_generator_import_intelligence_design_and_settings(self):
        html = self._report_shell_source()

        self.assertIn("HR一站式AI工作台", html)
        self.assertIn("HRobot", html)
        self.assertIn('id="page-3"', html)
        self.assertIn('id="page-4"', html)
        self.assertIn('id="page-5"', html)
        self.assertIn('id="page-6"', html)
        self.assertIn('id="page-7"', html)
        self.assertIn('id="page-8"', html)
        self.assertIn('id="page-9"', html)
        self.assertIn('id="page-10"', html)
        self.assertIn('id="generatedReport"', html)
        self.assertIn("/api/report", html)
        self.assertIn('id="reportDeleteBtn"', html)
        self.assertIn('id="aiChatForm"', html)
        self.assertIn('id="aiChatInput"', html)
        self.assertIn('id="multimodalConfigForm"', html)
        self.assertIn('id="imageConfigForm"', html)
        self.assertIn('id="intelligenceSearch"', html)
        self.assertIn('id="designPosterForm"', html)
        self.assertIn('id="agentProjectGrid"', html)
        self.assertIn('id="agentProjectDropCard"', html)
        self.assertIn('id="agentProjectZipInput"', html)
        self.assertIn("repeat(3, minmax(0, 1fr))", html)
        self.assertIn("uploadAgentProjectFile", html)
        self.assertIn("dragover", html)
        self.assertIn("agent-project-desc", html)
        self.assertIn("Python 服务", html)
        self.assertIn("/api/agent-projects/analyze", html)
        self.assertNotIn("agent-project-preview", html)
        self.assertIn('id="posterHistoryGrid"', html)
        self.assertIn('id="posterPreviewDialog"', html)
        self.assertIn("openPosterPreview", html)
        self.assertIn("data-poster-preview", html)
        self.assertIn('id="abilityCenter"', html)
        self.assertIn('id="reportAbilityForm"', html)
        self.assertIn('id="materialImportForm"', html)
        self.assertIn('id="reportGenerateForm"', html)
        self.assertIn('id="reviewImportForm"', html)
        self.assertIn('id="profileImportForm"', html)
        self.assertIn("数据配置", html)
        self.assertIn('id="talentPoolNameInput"', html)
        self.assertIn('id="talentPoolMembersInput"', html)
        self.assertIn('id="talentPoolFilterPanel"', html)
        self.assertIn("/api/report/generate", html)
        self.assertIn("/api/ai/chat", html)
        self.assertIn("/api/talent-pools", html)
        self.assertIn("/api/report/upload-skill", html)
        self.assertIn("/api/report/upload-material", html)
        self.assertIn("/api/import/file", html)
        self.assertIn("data-file-delete", html)
        self.assertIn("/api/import/review-excel", html)
        self.assertIn("/api/import/profiles-json", html)
        self.assertIn("/api/import/sources", html)
        self.assertIn("/api/intelligence", html)
        self.assertIn("/api/design/posters", html)
        self.assertIn("/api/design/posters/generate", html)
        self.assertIn("/api/agent-projects", html)
        self.assertIn("/api/agent-projects/upload", html)
        self.assertIn("/api/agent-projects/open", html)
        self.assertIn("openAgentProject", html)
        self.assertIn("/api/ai/image/test", html)
        self.assertIn('id="refreshImportSourcesBtn"', html)
        self.assertIn('id="reviewSourceList"', html)
        self.assertIn('id="profileSourceList"', html)
        self.assertNotIn('id="profileSnapshotList"', html)
        self.assertNotIn("profileSnapshots", html)
        self.assertIn('id="refreshReportAssetsBtn"', html)
        self.assertIn('id="refreshReportAssetsInlineBtn"', html)
        self.assertIn('id="reportAbilityList"', html)
        self.assertIn("/api/report/ability", html)
        self.assertIn('id="reportMaterialAssetList"', html)
        self.assertNotIn('id="aiConfigForm"', html)
        self.assertIn("AI问答", html)
        self.assertIn("人才盘点", html)
        self.assertIn("情报中心", html)
        self.assertIn("设计中心", html)
        self.assertIn("Agent中心", html)

    def test_index_supports_hiding_sidebar_and_page_navigation(self):
        html = self._report_shell_source()

        self.assertIn('id="sidebarToggle"', html)
        self.assertIn("toggleSidebar", html)
        self.assertIn(".shell.sidebar-collapsed", html)
        self.assertIn("switchPage(item.dataset.page)", html)
        self.assertIn(".report-page:not(.active) { display: none; }", html)
        self.assertIn(".report-page.active.cover-page { display: grid; }", html)

    def test_index_filters_departments_by_review_org_levels(self):
        html = self._report_shell_source()

        self.assertIn('id="groupFilterPanel"', html)
        self.assertIn('id="departmentFilterPanel"', html)
        self.assertIn('id="growthWarningFilterPanel"', html)
        self.assertIn('id="aiAbilityFilterPanel"', html)
        self.assertIn('id="talentPoolFilterPanel"', html)
        self.assertIn('id="aiSelect"', html)
        self.assertIn('id="growthSelect"', html)
        self.assertIn("renderPotentialScores", html)
        self.assertIn('data-filter-trigger="department"', html)
        self.assertIn("buildDepartmentTree", html)
        self.assertIn("department-toggle", html)
        self.assertIn("expandedDepartments", html)
        self.assertIn("personOrgPath", html)
        self.assertIn("personTalentPools", html)
        self.assertIn('profileValue(person, "sequence")', html)
        self.assertIn("person.aiAbilityCalibrated || person.aiAbilityOriginal", html)
        self.assertIn("person.noGrowthWarningCalibrated || person.noGrowthWarningOriginal", html)
        self.assertIn("profile-collapsed", html)
        self.assertIn("profileExpanded", html)
        self.assertIn("annualPerformanceReviewValue", html)
        self.assertIn("profilePerformanceHistory", html)
        self.assertIn('period.includes("2025") && period.includes("年度")', html)
        self.assertIn("annual?.managerComment", html)
        self.assertIn("recentYearPerformanceValue", html)
        self.assertIn("if (person.performanceLatest) return person.performanceLatest;", html)
        self.assertIn("if (person.performanceOriginal) return person.performanceOriginal;", html)
        self.assertIn("profileTalentReviewHistory", html)
        self.assertIn("historicalGridValue(person, 2025, person.gridOriginal)", html)
        self.assertIn("review?.value", html)
        self.assertNotIn('id="secondDepartmentFilterPanel"', html)
        self.assertNotIn('id="thirdDepartmentFilterPanel"', html)

    def test_placeholder_report_pages_are_blank(self):
        html = self._report_shell_source()

        self.assertNotIn("page-head", html)
        self.assertNotIn("blank-panel", html)
        self.assertNotIn("后续补充报告内容", html)


    def test_intelligence_design_and_settings_are_not_placeholder_pages(self):
        html = self._report_shell_source()

        self.assertNotIn('id="page-7">\n        <div class="module-placeholder-inner"', html)
        self.assertNotIn('id="page-8">\n        <div class="module-placeholder-inner"', html)
        self.assertIn("renderIntelligence", html)
        self.assertIn("generateDesignPoster", html)
        self.assertIn("saveSettingsConfig", html)
        self.assertIn('id="intelligenceConfigForm"', html)
        self.assertIn("intelligence-settings-card", html)
        self.assertIn('id="intelligenceUpdateBtn"', html)
        self.assertIn("/api/intelligence/update", html)
        self.assertIn("/api/intelligence/config", html)
        self.assertIn("hrbp-takeaway", html)


if __name__ == "__main__":
    unittest.main()
