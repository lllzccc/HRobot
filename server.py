from __future__ import annotations

import argparse
import base64
import contextlib
import csv
import html as html_lib
import importlib.util
import io
import json
import os
import platform
import shlex
import subprocess
import re
import shutil
import sys
import tempfile
import threading
import time
import uuid
import webbrowser
from datetime import datetime, timezone, timedelta
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, urljoin, urlparse
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen
from zipfile import ZipFile

from app.modules.agent_center import AgentCenterStoreMixin
from app.modules.talent_review import TalentReviewStoreMixin
from app.core.people import PeopleDataService
from app.shared.constants import GRID_LABELS

if getattr(sys, "frozen", False):
    if sys.stdout is None:
        sys.stdout = open(os.devnull, "w", encoding="utf-8")
    if sys.stderr is None:
        sys.stderr = open(os.devnull, "w", encoding="utf-8")


def packaged_resource_root():
    if getattr(sys, "frozen", False):
        return Path(getattr(sys, "_MEIPASS", Path(sys.executable).resolve().parent))
    return Path(__file__).resolve().parent


def app_root():
    if getattr(sys, "frozen", False) and sys.platform == "darwin":
        return Path.home() / "Library" / "Application Support" / "Hrobot"
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


PACKAGED_RESOURCE_ROOT = packaged_resource_root()
ROOT = app_root()
DATA_DIR = ROOT / "data"
DEBUG_LOG_PATH = ROOT / "server-debug.log"


def sync_packaged_macos_runtime():
    if not (getattr(sys, "frozen", False) and sys.platform == "darwin"):
        return
    ROOT.mkdir(parents=True, exist_ok=True)
    for filename in ("index.html", "app_version.json"):
        source = PACKAGED_RESOURCE_ROOT / filename
        if source.exists():
            shutil.copy2(source, ROOT / filename)
    for dirname in ("static", "assets", "scripts"):
        source = PACKAGED_RESOURCE_ROOT / dirname
        if source.exists():
            shutil.copytree(source, ROOT / dirname, dirs_exist_ok=True)


sync_packaged_macos_runtime()


def talent_snapshot_root(data_dir: Path):
    configured = str(os.environ.get("HROBOT_TALENT_SNAPSHOT_ROOT") or "").strip()
    if configured:
        return Path(configured).expanduser().resolve()
    external = Path(data_dir).resolve().parent.parent / "HRobot talent snapshots"
    return external if external.exists() else None


def _agent_debug_log(location, message, data=None, hypothesis_id=""):
    if os.environ.get("HROBOT_DEBUG") != "1":
        return
    try:
        entry = {
            "location": location,
            "message": message,
            "data": data or {},
            "hypothesisId": hypothesis_id,
            "timestamp": int(time.time() * 1000),
        }
        with open(DEBUG_LOG_PATH, "a", encoding="utf-8") as log_file:
            log_file.write(json.dumps(entry, ensure_ascii=False) + "\n")
    except Exception:
        pass
LOCAL_TZ = timezone(timedelta(hours=8))
INTELLIGENCE_UPDATE_LOCK = threading.Lock()
SERVER_STARTED_AT = datetime.now(LOCAL_TZ)
SERVER_HOST = "127.0.0.1"
SERVER_PORT = 8767
SERVER_ALLOW_REMOTE_CLIENTS = False
DEFAULT_APP_VERSION = "0.1.0"
APP_VERSION_PATH = ROOT / "app_version.json"
GITHUB_RELEASE_PAGE_URL = "https://github.com/lllzccc/HRobot/releases/latest"
GITHUB_RELEASE_API_URL = "https://api.github.com/repos/lllzccc/HRobot/releases/latest"


def app_platform():
    if sys.platform == "darwin":
        return "mac"
    if os.name == "nt":
        return "windows"
    return "linux"


def app_architecture():
    machine = platform.machine().lower()
    if machine in {"arm64", "aarch64"}:
        return "arm64"
    if machine in {"x86_64", "amd64"}:
        return "x64"
    return machine or "unknown"


def app_version_payload():
    saved = {}
    if APP_VERSION_PATH.exists():
        try:
            saved = json.loads(APP_VERSION_PATH.read_text(encoding="utf-8-sig"))
        except (OSError, json.JSONDecodeError, UnicodeDecodeError):
            saved = {}
    if not isinstance(saved, dict):
        saved = {}
    app_name = str(saved.get("name") or "Hrobot")
    default_install_dir = Path(os.environ.get("LOCALAPPDATA") or Path.home()) / app_name
    return {
        "name": app_name,
        "version": str(saved.get("version") or DEFAULT_APP_VERSION),
        "installDir": str(default_install_dir),
    }


def _version_parts(value):
    text = str(value or "").strip().lstrip("vV")
    parts = []
    for item in re.split(r"[.\-+_]", text):
        if item.isdigit():
            parts.append(int(item))
        elif item:
            match = re.match(r"^(\d+)", item)
            parts.append(int(match.group(1)) if match else 0)
    return tuple(parts or [0])


def is_newer_version(candidate, current):
    left = list(_version_parts(candidate))
    right = list(_version_parts(current))
    length = max(len(left), len(right))
    left.extend([0] * (length - len(left)))
    right.extend([0] * (length - len(right)))
    return tuple(left) > tuple(right)

REPORT_PRESETS = {
    "360": {
        "name": "360报告",
        "description": "围绕多方反馈、关键行为、协作影响和发展建议生成个人或团队 360 反馈报告。",
        "keywords": ["360", "360报告", "360反馈", "multi-rater", "feedback"],
        "settingFile": "360报告设定说明.md",
        "prompt": (
            "请按 360 报告口径生成：正文必须严格使用导入的 360 报告 skill 架构，"
            "不得自行新增 skill 之外的章节、表格、说明块或模板字段；"
            "重点呈现优势行为、待发展行为、典型证据、风险提醒和行动建议。"
        ),
    },
    "org-diagnosis": {
        "name": "组织诊断报告",
        "description": "聚焦组织结构、角色分工、关键流程、协作效率、风险和改进路径。",
        "keywords": ["组织诊断", "org-diagnosis", "diagnosis", "organization", "组织分析"],
        "settingFile": "组织诊断报告设定说明.md",
        "prompt": (
            "请按组织诊断报告口径生成：覆盖组织现状、关键问题、根因判断、影响范围、"
            "优先级和可执行改进方案；结论要区分数据事实、访谈/材料线索和推断。"
        ),
    },
    "talent-review": {
        "name": "人才盘点报告",
        "description": "基于九宫格、绩效、潜力、职级、序列和校准记录生成盘点分析。",
        "keywords": ["人才盘点", "九宫格", "talent-review", "ninebox", "talent"],
        "settingFile": "人才盘点报告设定说明.md",
        "prompt": (
            "请按人才盘点报告口径生成：覆盖人才结构、九宫格分布、关键人才、风险人员、"
            "历史变化、校准差异、培养、激励、保留建议和后续跟进动作。"
        ),
    },
}

REPORT_PRESET_ABILITY_ID_PREFIX = "report-preset-ability-"


def report_preset_ability_id(preset_id):
    return f"{REPORT_PRESET_ABILITY_ID_PREFIX}{preset_id}"

DEFAULT_HOME_MEMO = {
    "updatedAt": "",
    "records": [],
}

DEFAULT_HOME_AVATAR_CONFIG = {
    "updatedAt": "",
    "selectedSrc": "assets/brand/hrobot-buddy-avatar.svg",
}

HOME_AVATAR_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".svg"}
HOME_AVATAR_TRIMMABLE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp"}

DEFAULT_DESIGN_PROMPT_CONFIG = {
    "basePrompt": "为 HRobot HRBP 工作台生成一张可直接使用的本地海报图片。",
    "brandRequirements": "优先使用HRobot 通用视觉：主蓝 #2f64e9、辅助青 #22d3ee、中性深色 #0f172a；整体保持专业、清晰、可信赖的 HR 场景表达。",
    "customRequirements": "",
    "referenceInstructions": "如果 references 目录中有可用素材，请结合素材文件名、说明和用户需求，预留合适的 logo/吉祥物/模板位置。",
    "template": (
        "{basePrompt}\n"
        "{brandRequirements}\n"
        "{customRequirements}\n"
        "{referenceInstructions}\n"
        "可用参考素材：{referenceSummary}\n"
        "海报类型：{posterType}\n"
        "风格：{style}\n"
        "尺寸：{size}\n"
        "需求：{requirement}"
    ),
}

def server_status_payload():
    now = datetime.now(LOCAL_TZ)
    return {
        "connected": True,
        "status": "connected",
        "pid": os.getpid(),
        "host": SERVER_HOST,
        "port": SERVER_PORT,
        "root": str(ROOT),
        "startedAt": SERVER_STARTED_AT.isoformat(timespec="seconds"),
        "checkedAt": now.isoformat(timespec="seconds"),
        "uptimeSeconds": max(0, int((now - SERVER_STARTED_AT).total_seconds())),
    }


def schedule_server_restart(delay_seconds=0.8):
    command = [sys.executable]
    if not getattr(sys, "frozen", False):
        command.append(str(ROOT / "server.py"))
    command.extend(["--host", SERVER_HOST, "--port", str(SERVER_PORT)])
    if SERVER_ALLOW_REMOTE_CLIENTS:
        command.append("--allow-remote-clients")
    probe_host = "127.0.0.1" if SERVER_HOST in {"0.0.0.0", "::", ""} else SERVER_HOST

    launcher_script = (
        "import os, socket, subprocess, time\n"
        f"time.sleep({float(delay_seconds)!r})\n"
        f"command = {command!r}\n"
        f"cwd = {str(ROOT)!r}\n"
        f"host = {probe_host!r}\n"
        f"port = {int(SERVER_PORT)!r}\n"
        "deadline = time.time() + 20\n"
        "while time.time() < deadline:\n"
        "    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:\n"
        "        sock.settimeout(0.25)\n"
        "        if sock.connect_ex((host, port)) != 0:\n"
        "            break\n"
        "    time.sleep(0.25)\n"
        "kwargs = {'cwd': cwd, 'stdin': subprocess.DEVNULL, 'stdout': subprocess.DEVNULL, 'stderr': subprocess.DEVNULL}\n"
        "if os.name == 'nt':\n"
        "    kwargs['creationflags'] = subprocess.CREATE_NO_WINDOW\n"
        "subprocess.Popen(command, **kwargs)\n"
    )
    kwargs = {
        "cwd": str(ROOT),
        "stdin": subprocess.DEVNULL,
        "stdout": subprocess.DEVNULL,
        "stderr": subprocess.DEVNULL,
    }
    if os.name == "nt":
        kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
    subprocess.Popen([sys.executable, "-c", launcher_script], **kwargs)
    threading.Timer(0.2, lambda: os._exit(0)).start()


class ReusableThreadingHTTPServer(ThreadingHTTPServer):
    allow_reuse_address = True


class DataStore(AgentCenterStoreMixin, TalentReviewStoreMixin):
    def __init__(self, data_dir: Path):
        self.data_dir = Path(data_dir)
        self.talent_snapshot_root = talent_snapshot_root(self.data_dir)
        self.review_source_dir = self.data_dir / "review_results"
        self.profile_source_dir = self.data_dir / "talent_profiles"
        self.profile_snapshot_dir = self.data_dir / "talent_profile_snapshots"
        self.hrbp_profile_split_dir = (
            self.talent_snapshot_root / "hrbp_profile_splits"
            if self.talent_snapshot_root
            else self.data_dir / "hrbp_profile_splits"
        )
        self.permission_dir = (
            self.talent_snapshot_root / "permissions"
            if self.talent_snapshot_root
            else self.data_dir / "permissions"
        )
        self.report_dir = self.data_dir / "report_generation"
        self.legacy_review_source_dirs = [
            self.data_dir / "盘点结果",
        ]
        self.legacy_profile_source_dirs = [
            self.data_dir / "人才档案",
        ]
        self.legacy_profile_snapshot_dirs = [self.data_dir / "人才档案快照_hrobot"]
        self.legacy_hrbp_profile_split_dirs = [self.data_dir / "人才档案_HRBP拆分"]
        self.legacy_permission_dirs = [self.data_dir / "权限配置"]
        self.legacy_report_dirs = [self.data_dir / "报告生成"]
        self.review_path = self.data_dir / "talent_review_2026.json"
        self.profile_path = self.data_dir / "people_profiles.json"
        self.overrides_path = self.data_dir / "calibration_overrides.json"
        self.profile_notes_path = self.data_dir / "profile_notes.json"
        self.employee_map_path = self.data_dir / "employee_manager_map.json"
        self.ai_config_path = self.data_dir / "ai_config.json"
        self.ai_secrets_path = self.data_dir / "local_ai_secrets.json"
        self.mcp_config_path = self.data_dir / "mcp_config.json"
        self.data_source_config_path = self.data_dir / "data_sources.json"
        self.update_config_path = self.data_dir / "update_config.json"
        self.talent_pool_path = self.data_dir / "talent_pools.json"
        self.intelligence_path = self.data_dir / "intelligence.json"
        self.intelligence_history_path = self.data_dir / "intelligence_history.json"
        self.intelligence_config_path = self.data_dir / "intelligence_config.json"
        self.intelligence_update_status_path = self.data_dir / "intelligence_update_status.json"
        self.design_dir = self.data_dir / "design_center"
        self.design_poster_dir = self.design_dir / "posters"
        self.design_history_path = self.design_dir / "poster_history.json"
        self.design_prompt_config_path = self.design_dir / "poster_prompt_config.json"
        self.design_reference_dir = self.design_dir / "references"
        self.design_reference_readme_path = self.design_reference_dir / "README.md"
        self.agent_center_dir = self.data_dir / "agent_center"
        self.agent_zip_dir = self.agent_center_dir / "zips"
        self.agent_project_dir = self.agent_center_dir / "projects"
        self.agent_manifest_path = self.agent_center_dir / "manifest.json"
        self._agent_processes = {}
        self._agent_process_lock = threading.RLock()
        self.home_memo_path = self.data_dir / "home_memo.json"
        self.home_avatar_config_path = self.data_dir / "home_avatar_config.json"
        self.home_avatar_dir = ROOT / "assets" / "avatars"
        self.home_avatar_trim_dir = self.home_avatar_dir / "_trimmed"
        self.report_skill_dir = self.report_dir / "skills"
        self.report_material_dir = self.report_dir / "materials"
        self.report_ability_dir = self.report_dir / "abilities"
        self.report_ability_file_dir = self.report_ability_dir / "files"
        self.report_ability_manifest_path = self.report_ability_dir / "manifest.json"
        self.report_setting_dir = self.report_dir / "settings"
        self.report_markdown_dir = self.report_dir / "reports_md"
        self.report_html_dir = self.report_dir / "reports_html"
        self.generated_report_path = self.report_dir / "generated_report.json"
        self.report_history_path = self.report_dir / "generated_reports.json"
        self._people_data_service = PeopleDataService(self)

    def people_data(self):
        return self._people_data_service

    def _review_source_dirs(self):
        return [self.review_source_dir, *self.legacy_review_source_dirs]

    def _profile_source_dirs(self):
        return [self.profile_source_dir, *self.legacy_profile_source_dirs]

    def _report_dirs(self):
        return [self.report_dir, *self.legacy_report_dirs]

    def _first_existing_path(self, paths):
        return next((path for path in paths if path.exists()), paths[0])

    def _path_signature(self, path: Path):
        if not path.exists() or not path.is_file():
            return (str(path), None, 0)
        stat = path.stat()
        return (str(path), stat.st_mtime_ns, stat.st_size)

    def _files_signature(self, folders, pattern, fallback_path=None):
        files = []
        has_source_dir = False
        for folder in folders:
            if folder.exists():
                has_source_dir = True
                files.extend(path for path in folder.glob(pattern) if path.is_file())
        if not files and fallback_path is not None and not has_source_dir:
            files = [fallback_path]
        return tuple(self._path_signature(path) for path in sorted(set(files), key=lambda item: str(item)))

    def _has_source_dir(self, folders):
        return any(folder.exists() for folder in folders)

    def _source_signature(self, folders, pattern, fallback_path: Path):
        source_path = self._latest_file_in(folders, pattern) or fallback_path
        return self._path_signature(source_path)

    def _cached(self, key, signature, factory):
        if not hasattr(self, "_cache"):
            self._cache = {}
            self._cache_lock = threading.RLock()
        with self._cache_lock:
            entry = self._cache.get(key)
            if entry and entry["signature"] == signature:
                return entry["value"]
            value = factory()
            self._cache[key] = {"signature": signature, "value": value}
            return value

    def _clear_cache(self):
        if hasattr(self, "_cache_lock"):
            with self._cache_lock:
                self._cache.clear()

    def _read_json(self, path: Path, fallback):
        if not path.exists():
            return fallback
        try:
            return json.loads(path.read_text(encoding="utf-8-sig"))
        except UnicodeDecodeError as error:
            _agent_debug_log(
                "server.py:_read_json",
                "UnicodeDecodeError reading JSON file",
                {"path": str(path), "error": str(error)},
                "H1",
            )
            raise
        except json.JSONDecodeError as error:
            _agent_debug_log(
                "server.py:_read_json",
                "JSONDecodeError parsing file",
                {"path": str(path), "error": str(error)},
                "H1",
            )
            raise

    def _latest_json_in(self, folder: Path):
        if not folder.exists():
            return None
        files = [path for path in folder.glob("*.json") if path.is_file()]
        if not files:
            return None
        return max(files, key=lambda path: (path.stat().st_mtime, path.name))

    def _latest_file_in(self, folders, pattern):
        files = []
        for folder in folders:
            if folder.exists():
                files.extend(path for path in folder.glob(pattern) if path.is_file())
        if not files:
            return None
        return max(files, key=lambda path: (path.stat().st_mtime, path.name))

    def _latest_review_json(self):
        files = []
        for folder in self._review_source_dirs():
            if folder.exists():
                files.extend(
                    path
                    for path in folder.glob("*.json")
                    if path.is_file() and path.name != "calibrated_latest.json"
                )
        if not files:
            return None
        return max(files, key=lambda path: (path.stat().st_mtime, path.name))

    def _read_source_json(self, folders, fallback_path: Path, fallback):
        source_path = self._latest_file_in(folders, "*.json") or fallback_path
        return self._read_json(source_path, fallback)

    def _review_source_path(self):
        source_path = self._latest_review_json()
        if source_path:
            return source_path
        if self._has_source_dir(self._review_source_dirs()):
            return self.review_source_dir / "__empty_review_results.json"
        return self.review_path

    def _read_profile_source_json(self):
        files = []
        for folder in self._profile_source_dirs():
            if folder.exists():
                files.extend(path for path in folder.glob("*.json") if path.is_file())
        if not files:
            if self._has_source_dir(self._profile_source_dirs()):
                return []
            return self._read_json(self.profile_path, [])

        merged = {}
        anonymous = []
        for path in sorted(files, key=lambda item: (item.stat().st_mtime_ns, item.name)):
            data = self._read_json(path, [])
            profiles = data.get("profiles") if isinstance(data, dict) and "profiles" in data else data
            if not isinstance(profiles, list):
                continue
            for profile in profiles:
                if not isinstance(profile, dict):
                    continue
                employee_id = profile.get("employeeId")
                if employee_id:
                    merged[str(employee_id)] = profile
                else:
                    anonymous.append(profile)
        return anonymous + list(merged.values())

    def _clear_imported_files(self, folder: Path, patterns):
        folder.mkdir(parents=True, exist_ok=True)
        for pattern in patterns:
            for path in folder.glob(pattern):
                if path.is_file():
                    path.unlink()

    def review_results(self):
        source_path = self._review_source_path()
        signature = self._path_signature(source_path)
        return self._cached(
            "review_results",
            signature,
            lambda: self._read_json(source_path, []),
        )

    def profiles(self):
        signature = (
            self._files_signature(self._profile_source_dirs(), "*.json", self.profile_path),
            self._path_signature(self.employee_map_path),
        )

        def load_profiles():
            profiles = self._read_profile_source_json()
            if not isinstance(profiles, list):
                return []
            normalized = []
            for profile in profiles:
                if not isinstance(profile, dict):
                    continue
                item = dict(profile)
                if item.get("employeeId") not in (None, ""):
                    item["employeeId"] = self._normalize_employee_id(item.get("employeeId"))
                normalized.append(self._enrich_with_employee_info(item))
            return normalized

        return self._cached("profiles", signature, load_profiles)

    def employee_map(self):
        signature = self._path_signature(self.employee_map_path)
        return self._cached("employee_map", signature, lambda: self._read_json(self.employee_map_path, {"byEmployeeId": {}, "byName": {}}))

    def _normalize_employee_id(self, value):
        text = str(value or "").strip()
        if re.fullmatch(r"\d+\.0", text):
            return text[:-2]
        return text

    def _employee_info_for(self, record):
        mapping = self.employee_map()
        employee_id = self._normalize_employee_id(record.get("employeeId") or record.get("员工ID") or record.get("工号") or "")
        name = str(record.get("name") or record.get("姓名") or "").strip()
        return mapping.get("byEmployeeId", {}).get(employee_id) or mapping.get("byName", {}).get(name) or {}

    def _normalize_period(self, value):
        text = re.sub(r"\s+", "", str(value or "").strip())
        return text.lstrip(":：")

    def _normalize_multi_value(self, value):
        parts = [part.strip() for part in re.split(r"[;,，、；\n\r]+", str(value or "")) if part and part.strip()]
        seen = []
        for part in parts:
            if part not in seen:
                seen.append(part)
        return ",".join(seen)

    def _profile_annual_manager_comment(self, record):
        for item in record.get("performanceHistory", []) or []:
            if not isinstance(item, dict):
                continue
            period = self._normalize_period(item.get("period", ""))
            comment = item.get("managerComment") or item.get("manager comment") or item.get("Manager Comment")
            if period == "2025年度" and comment not in (None, ""):
                return comment
        return ""

    def _enrich_with_employee_info(self, record):
        enriched = dict(record)
        annual_comment = self._profile_annual_manager_comment(enriched)
        if annual_comment:
            enriched["managerComment2025"] = annual_comment
            enriched["annualPerformanceReview"] = annual_comment
        info = self._employee_info_for(record)
        if not info:
            return enriched
        for target, source in (
            ("manager", "manager"),
            ("managerEmail", "managerEmail"),
            ("title", "title"),
            ("level", "level"),
            ("sequence", "sequence"),
            ("departmentPath", "departmentPath"),
            ("age", "age"),
            ("tenure", "tenure"),
            ("email", "email"),
            ("一级组织", "一级组织"),
            ("二级组织", "二级组织"),
            ("三级组织", "三级组织"),
            ("四级组织", "四级组织"),
            ("五级组织", "五级组织"),
        ):
            if info.get(source) and not enriched.get(target):
                enriched[target] = info.get(source)
        return enriched

    def _safe_upload_name(self, filename: str, fallback: str):
        safe_name = Path(filename or fallback).name
        return safe_name or fallback

    def _read_text_file(self, path: Path, limit=60000):
        if path.suffix.lower() == ".zip":
            snippets = []
            with ZipFile(path) as archive:
                names = [
                    name for name in archive.namelist()
                    if not name.endswith("/") and Path(name).suffix.lower() in {".md", ".txt", ".json", ".yaml", ".yml", ".xml", ".html"}
                ]
                for name in names[:12]:
                    raw = archive.read(name)
                    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
                        try:
                            snippets.append(f"--- {name} ---\n{raw.decode(encoding)[:12000]}")
                            break
                        except UnicodeDecodeError:
                            continue
                    if sum(len(item) for item in snippets) >= limit:
                        break
            return ("\n\n".join(snippets) or f"[压缩包内未发现可读取文本：{path.name}]")[:limit]
        data = path.read_bytes()
        for encoding in ("utf-8-sig", "utf-8", "gb18030"):
            try:
                text = data.decode(encoding)
                return text[:limit]
            except UnicodeDecodeError:
                continue
        return f"[二进制文件，仅记录文件名：{path.name}]"

    def _read_pdf_text_file(self, path: Path, limit=60000):
        try:
            from pypdf import PdfReader
        except ImportError as error:
            raise ValueError("当前环境缺少 PDF 解析依赖 pypdf，无法读取 360 PDF 材料。请先安装依赖或上传可读取文本材料。") from error
        try:
            reader = PdfReader(str(path))
            pages = []
            for page in reader.pages:
                text = page.extract_text() or ""
                if text.strip():
                    pages.append(text.strip())
                if sum(len(item) for item in pages) >= limit:
                    break
            content = "\n\n".join(pages).strip()
        except Exception as error:
            raise ValueError(f"360 PDF 材料解析失败：{path.name}。请上传可解析 PDF 或先转换为文本/Markdown。") from error
        if not content:
            raise ValueError(f"360 PDF 材料未解析出可读取正文：{path.name}。请上传可解析 PDF 或先转换为文本/Markdown。")
        return content[:limit]

    def _extract_360_pdf_summary(self, filename: str, content: str):
        text = str(content or "")
        is_core_report = "核心版" in filename
        summary = []
        rank_match = re.search(r"排名[：:\s]*([0-9]+\s*/\s*[0-9]+)", text)
        score_pattern = r"([0-5]\.\d{1,2})"
        total_match = re.search(rf"(?:他评[：:\s]*|总分为[：:\s]*){score_pattern}", text)
        self_match = re.search(rf"自评[：:\s]*{score_pattern}", text) if is_core_report else None
        role_pairs = []
        if is_core_report:
            for label in ("上级", "同事", "下级"):
                match = re.search(rf"{re.escape(label)}[：:\s]*{score_pattern}", text)
                if match:
                    role_pairs.append(f"{label}：{match.group(1)}")
        if rank_match:
            summary.append(f"- 排名：{rank_match.group(1)}")
        if total_match:
            summary.append(f"- 他评：{total_match.group(1)}")
        if self_match:
            summary.append(f"- 自评：{self_match.group(1)}")
        if role_pairs:
            summary.append(f"- 角色评分：{'；'.join(role_pairs)}")

        capability_patterns = (
            ("高分能力", r"他评分较高的能力[：:]\s*([^；。\n]+)"),
            ("待发展能力", r"他评分较低的能力[：:]\s*([^；。\n]+)"),
            ("差异较大能力", r"各角色分差异较大的能力[：:]\s*([^；。\n]+)"),
        )
        for label, pattern in capability_patterns:
            match = re.search(pattern, text)
            if match:
                summary.append(f"- {label}：{self._clean_metric_value(match.group(1), limit=80)}")

        if "开放性反馈" in text or "文本反馈" in text:
            summary.append("- 含开放性反馈/文本反馈原文")

        kind = "核心版" if "核心版" in filename else "标准版" if "标准版" in filename else "PDF"
        header = f"### 360材料结构化摘要 - {Path(filename).name}（{kind}）"
        if not summary:
            summary.append("- 暂未从正文稳定提取到结构化指标，请结合下方原始解析正文阅读。")
        return f"{header}\n" + "\n".join(summary)

    def _read_360_pdf_context(self, path: Path, limit=60000):
        content = self._read_pdf_text_file(path, limit=limit)
        summary = self._extract_360_pdf_summary(path.name, content)
        return f"{summary}\n\n### 原始解析正文 - {path.name}\n{content}"[:limit]

    def _read_report_asset_text(self, path: Path, preset_id=None, label=None):
        if preset_id == "360" and label == "其他分析材料" and path.suffix.lower() == ".pdf":
            return self._read_360_pdf_context(path)
        return self._read_text_file(path)

    def import_report_asset(self, kind: str, filename: str, content: bytes):
        folder = self.report_skill_dir if kind == "skill" else self.report_material_dir
        folder.mkdir(parents=True, exist_ok=True)
        safe_name = self._safe_upload_name(filename, f"{kind}_{datetime.now().strftime('%Y%m%d%H%M%S')}.txt")
        target = folder / safe_name
        target.write_bytes(content)
        return {"filename": target.name, "path": str(target), "size": target.stat().st_size}

    def _report_ability_manifest(self):
        payload = self._read_json(self.report_ability_manifest_path, {"abilities": [], "updatedAt": ""})
        if not isinstance(payload, dict):
            payload = {"abilities": [], "updatedAt": ""}
        abilities = payload.get("abilities", [])
        if not isinstance(abilities, list):
            abilities = []
        payload["abilities"] = [item for item in abilities if isinstance(item, dict)]
        return payload

    def _save_report_ability_manifest(self, abilities):
        self.report_ability_manifest_path.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "abilities": abilities,
            "updatedAt": datetime.now(LOCAL_TZ).isoformat(timespec="seconds"),
        }
        self.report_ability_manifest_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        self._clear_cache()
        return payload

    def _report_preset_setting_path(self, preset):
        setting_file = preset.get("settingFile", "")
        if not setting_file:
            return Path("")
        setting_paths = [
            self.report_setting_dir / setting_file,
            *[folder / "设定说明" / setting_file for folder in self.legacy_report_dirs],
        ]
        return self._first_existing_path(setting_paths)

    def _report_preset_abilities(self, payload=None):
        payload = payload or self._report_ability_manifest()
        overrides = {item.get("id"): item for item in payload.get("abilities", []) if item.get("id")}
        abilities = []
        for preset_id, preset in REPORT_PRESETS.items():
            ability_id = report_preset_ability_id(preset_id)
            override = overrides.get(ability_id, {})
            setting_path = self._report_preset_setting_path(preset)
            default_description = self._read_text_file(setting_path) if setting_path.exists() else preset.get("prompt", "")
            description_md = override.get("descriptionMd") if override else default_description
            updated_at = override.get("updatedAt", "")
            if not updated_at and setting_path.exists():
                updated_at = datetime.fromtimestamp(setting_path.stat().st_mtime, LOCAL_TZ).isoformat(timespec="seconds")
            abilities.append({
                "id": ability_id,
                "name": override.get("name") or f"{preset['name']}能力",
                "descriptionMd": description_md,
                "summary": self._report_ability_summary(description_md),
                "zipFilename": override.get("zipFilename", ""),
                "zipAsset": override.get("zipAsset", ""),
                "zipSize": int(override.get("zipSize", 0) or 0),
                "zipFileCount": int(override.get("zipFileCount", 0) or 0),
                "createdAt": override.get("createdAt") or updated_at or datetime.now(LOCAL_TZ).isoformat(timespec="seconds"),
                "updatedAt": updated_at or datetime.now(LOCAL_TZ).isoformat(timespec="seconds"),
                "sourceType": "builtin",
                "presetId": preset_id,
                "settingFile": preset.get("settingFile", ""),
            })
        return abilities

    def report_abilities(self):
        payload = self._report_ability_manifest()
        builtin_abilities = self._report_preset_abilities(payload)
        builtin_ids = {item["id"] for item in builtin_abilities}
        custom_abilities = [
            item for item in payload.get("abilities", [])
            if item.get("id") not in builtin_ids
        ]
        return {
            **payload,
            "abilities": [
                *builtin_abilities,
                *sorted(custom_abilities, key=lambda item: item.get("updatedAt", ""), reverse=True),
            ],
        }

    @staticmethod
    def _report_ability_summary(markdown):
        text = re.sub(r"[*_`>#|\\-]+", "", str(markdown or ""))
        text = re.sub(r"\s+", " ", text).strip()
        return text[:120]

    def _validate_report_ability_zip(self, content):
        if not content:
            return []
        try:
            with ZipFile(io.BytesIO(content)) as archive:
                names = [name for name in archive.namelist() if name and not name.endswith("/")]
        except Exception as error:
            raise ValueError(f"能力 ZIP 无法读取：{error}")
        if not names:
            raise ValueError("能力 ZIP 为空。")
        return names

    def save_report_ability(self, name, description_md="", zip_filename="", zip_content=None, ability_id=""):
        title = str(name or "").strip()[:80]
        if not title:
            raise ValueError("请填写能力名称。")
        description_md = str(description_md or "").strip()
        ability_id = str(ability_id or "").strip()
        payload = self._report_ability_manifest()
        abilities = payload.get("abilities", [])
        builtin_existing = next((item for item in self._report_preset_abilities(payload) if item.get("id") == ability_id), None) if ability_id else None
        existing = next((item for item in abilities if item.get("id") == ability_id), None) if ability_id else None
        existing = existing or builtin_existing
        if not ability_id:
            slug = re.sub(r"[^a-zA-Z0-9]+", "-", title).strip("-").lower()[:32] or "ability"
            ability_id = f"{slug}-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6]}"
        self.report_ability_file_dir.mkdir(parents=True, exist_ok=True)
        zip_asset = existing.get("zipAsset", "") if existing else ""
        zip_original_name = existing.get("zipFilename", "") if existing else ""
        zip_size = int(existing.get("zipSize", 0) or 0) if existing else 0
        zip_file_count = int(existing.get("zipFileCount", 0) or 0) if existing else 0
        if zip_content:
            names = self._validate_report_ability_zip(zip_content)
            safe_zip_name = self._safe_upload_name(zip_filename or f"{ability_id}.zip", f"{ability_id}.zip")
            if not safe_zip_name.lower().endswith(".zip"):
                safe_zip_name = f"{Path(safe_zip_name).stem}.zip"
            zip_asset = f"{ability_id}-{safe_zip_name}"
            target = self.report_ability_file_dir / zip_asset
            target.write_bytes(zip_content)
            zip_original_name = Path(zip_filename or safe_zip_name).name
            zip_size = target.stat().st_size
            zip_file_count = len(names)
            if existing and existing.get("zipAsset") and existing.get("zipAsset") != zip_asset:
                old_path = self.report_ability_file_dir / Path(existing.get("zipAsset", "")).name
                if old_path.exists():
                    old_path.unlink()
        ability = {
            "id": ability_id,
            "name": title,
            "descriptionMd": description_md,
            "summary": self._report_ability_summary(description_md),
            "zipFilename": zip_original_name,
            "zipAsset": zip_asset,
            "zipSize": zip_size,
            "zipFileCount": zip_file_count,
            "createdAt": existing.get("createdAt") if existing else datetime.now(LOCAL_TZ).isoformat(timespec="seconds"),
            "updatedAt": datetime.now(LOCAL_TZ).isoformat(timespec="seconds"),
            "sourceType": "builtin" if builtin_existing or (existing and existing.get("sourceType") == "builtin") else "custom",
        }
        if ability["sourceType"] == "builtin":
            ability["presetId"] = (builtin_existing or existing).get("presetId", "")
            ability["settingFile"] = (builtin_existing or existing).get("settingFile", "")
        abilities = [ability, *[item for item in abilities if item.get("id") != ability_id]]
        self._save_report_ability_manifest(abilities)
        refreshed = self.report_abilities()
        refreshed["ability"] = next(
            (item for item in refreshed.get("abilities", []) if item.get("id") == ability_id),
            ability,
        )
        return refreshed

    def delete_report_ability(self, ability_id):
        ability_id = str(ability_id or "").strip()
        if not ability_id:
            raise ValueError("能力 ID 不能为空。")
        if ability_id in {report_preset_ability_id(preset_id) for preset_id in REPORT_PRESETS}:
            raise ValueError("内置报告能力不能删除，可以编辑说明。")
        payload = self._report_ability_manifest()
        abilities = payload.get("abilities", [])
        target = next((item for item in abilities if item.get("id") == ability_id), None)
        if not target:
            raise FileNotFoundError("未找到要删除的能力。")
        zip_asset = Path(target.get("zipAsset", "")).name
        if zip_asset:
            zip_path = self.report_ability_file_dir / zip_asset
            if zip_path.exists():
                zip_path.unlink()
        self._save_report_ability_manifest([item for item in abilities if item.get("id") != ability_id])
        return self.report_abilities()

    def _read_report_ability_zip_context(self, ability):
        zip_asset = Path(str(ability.get("zipAsset") or "")).name
        if not zip_asset:
            return ""
        zip_path = self.report_ability_file_dir / zip_asset
        if not zip_path.exists():
            return ""
        chunks = []
        total = 0
        text_suffixes = {".md", ".txt", ".json", ".yaml", ".yml", ".py", ".js", ".ts", ".html", ".css"}
        with ZipFile(zip_path) as archive:
            for info in archive.infolist():
                if info.is_dir() or len(chunks) >= 40:
                    continue
                suffix = Path(info.filename).suffix.lower()
                if suffix not in text_suffixes:
                    continue
                if info.file_size > 300000:
                    continue
                raw = archive.read(info)
                text = raw.decode("utf-8-sig", errors="replace").strip()
                if not text:
                    continue
                excerpt = text[:12000]
                chunks.append(f"### ZIP 文件：{info.filename}\n{excerpt}")
                total += len(excerpt)
                if total > 60000:
                    break
        return "\n\n".join(chunks)

    def report_ability_context(self, selected_abilities=None):
        if selected_abilities is None:
            return []
        requested_ids = [str(item or "").strip() for item in (selected_abilities or []) if str(item or "").strip()]
        if not requested_ids:
            return []
        if any(Path(item).name != item for item in requested_ids):
            raise ValueError("所选能力 ID 不合法。")
        abilities = self.report_abilities().get("abilities", [])
        if requested_ids:
            ability_by_id = {item.get("id"): item for item in abilities}
            missing = [item for item in requested_ids if item not in ability_by_id]
            if missing:
                raise FileNotFoundError(f"未找到所选能力：{'、'.join(missing)}")
            abilities = [ability_by_id[item] for item in requested_ids]
        sections = []
        for ability in abilities:
            content = [
                f"能力名称：{ability.get('name', '')}",
                f"能力说明 Markdown：\n{ability.get('descriptionMd', '') or '无'}",
            ]
            zip_context = self._read_report_ability_zip_context(ability)
            if zip_context:
                content.append(f"能力 ZIP 内容摘录：\n{zip_context}")
            sections.append({
                "type": "能力中心",
                "filename": ability.get("zipFilename") or ability.get("name", ""),
                "abilityId": ability.get("id", ""),
                "content": "\n\n".join(content),
            })
        return sections

    def _delete_named_file(self, folders, filename, exclude_names=None):
        safe_name = Path(filename or "").name
        if not safe_name or safe_name != filename:
            raise ValueError("文件名不合法。")
        if safe_name in set(exclude_names or []):
            raise ValueError("该文件不允许删除。")
        for folder in folders:
            path = folder / safe_name
            if path.exists() and path.is_file():
                path.unlink()
                self._clear_cache()
                return {"deleted": True, "filename": safe_name}
        raise FileNotFoundError("未找到要删除的文件。")

    def delete_imported_file(self, kind: str, filename: str):
        targets = {
            "report-skill": ([self.report_skill_dir, *[folder / "skills" for folder in self.legacy_report_dirs]], set()),
            "report-material": ([self.report_material_dir, *[folder / "materials" for folder in self.legacy_report_dirs]], set()),
            "review-result": (self._review_source_dirs(), {"calibrated_latest.json"}),
            "profile": (self._profile_source_dirs(), set()),
            "employee-roster": ([self.data_dir], set()),
        }
        if kind not in targets:
            raise ValueError("不支持的文件类型。")
        folders, exclude_names = targets[kind]
        if kind == "employee-roster" and filename != "employee_manager_map.json":
            raise ValueError("不支持删除该员工关系文件。")
        return self._delete_named_file(folders, filename, exclude_names)

    def report_assets(self):
        def list_files(folders):
            files = []
            for folder in folders:
                if folder.exists():
                    files.extend(path for path in folder.iterdir() if path.is_file())
            return [
                {"filename": path.name, "size": path.stat().st_size, "updatedAt": datetime.fromtimestamp(path.stat().st_mtime).isoformat()}
                for path in sorted(set(files), key=lambda item: item.stat().st_mtime, reverse=True)
            ]
        return {
            "abilities": self.report_abilities().get("abilities", []),
            "skills": list_files([self.report_skill_dir, *[folder / "skills" for folder in self.legacy_report_dirs]]),
            "materials": list_files([self.report_material_dir, *[folder / "materials" for folder in self.legacy_report_dirs]]),
        }

    def report_presets(self):
        return [
            {
                "id": preset_id,
                "name": preset["name"],
                "description": preset["description"],
                "settingFile": preset.get("settingFile", ""),
            }
            for preset_id, preset in REPORT_PRESETS.items()
        ]

    def import_sources(self):
        def list_files(folder, patterns, exclude_names=None):
            if not folder.exists():
                return []
            exclude_names = set(exclude_names or [])
            files = []
            for pattern in patterns:
                files.extend(path for path in folder.glob(pattern) if path.is_file() and path.name not in exclude_names)
            unique_files = sorted(set(files), key=lambda item: (item.stat().st_mtime, item.name), reverse=True)
            return [
                {
                    "filename": path.name,
                    "size": path.stat().st_size,
                    "updatedAt": datetime.fromtimestamp(path.stat().st_mtime).isoformat(),
                }
                for path in unique_files
            ]

        return {
            "reviewResults": sum((list_files(folder, ["*.xlsx", "*.json"], {"calibrated_latest.json"}) for folder in self._review_source_dirs()), []),
            "profiles": sum((list_files(folder, ["*.json"]) for folder in self._profile_source_dirs()), []),
            "employeeRoster": list_files(self.data_dir, ["employee_manager_map.json"]),
        }

    def report_asset_context(self, preset_id=None, target_name="", selected_skills=None, selected_materials=None, selected_abilities=None):
        sections = []
        preset = REPORT_PRESETS.get(preset_id or "")
        if preset:
            sections.append({
                "type": "报告预设",
                "filename": preset["name"],
                "content": preset["prompt"],
            })
        if selected_abilities is not None:
            sections.extend(self.report_ability_context(selected_abilities))
        asset_groups = [
            ("skill框架与分析逻辑", [self.report_skill_dir, *[folder / "skills" for folder in self.legacy_report_dirs]], True, selected_skills),
            ("其他分析材料", [self.report_material_dir, *[folder / "materials" for folder in self.legacy_report_dirs]], False, selected_materials),
        ]
        for label, folders, is_skill_group, selected_names in asset_groups:
            paths = []
            for folder in folders:
                if folder.exists():
                    paths.extend(path for path in folder.iterdir() if path.is_file())
            if selected_names is not None:
                requested_names = list(dict.fromkeys(str(name or "").strip() for name in selected_names if str(name or "").strip()))
                if any(Path(name).name != name for name in requested_names):
                    raise ValueError("所选报告资料文件名不合法。")
                available_paths = {}
                for path in paths:
                    available_paths.setdefault(path.name, path)
                missing_names = [name for name in requested_names if name not in available_paths]
                if missing_names:
                    raise FileNotFoundError(f"未找到所选报告资料：{'、'.join(missing_names)}")
                paths = [available_paths[name] for name in requested_names]
            elif preset_id == "360" and label == "其他分析材料" and target_name:
                matched_paths = [path for path in paths if target_name in path.name]
                if matched_paths:
                    paths = matched_paths
            if selected_names is not None:
                pass
            elif preset and is_skill_group:
                keywords = [keyword.lower() for keyword in preset.get("keywords", [])]

                def priority(path):
                    return 0 if any(keyword in path.name.lower() for keyword in keywords) else 1
                paths = sorted(paths, key=lambda item: (priority(item), -item.stat().st_mtime))
            else:
                paths = sorted(paths, key=lambda item: item.stat().st_mtime, reverse=True)
            for path in paths:
                if path.is_file():
                    sections.append({"type": label, "filename": path.name, "content": self._read_report_asset_text(path, preset_id, label)})
        return sections

    def _report_title(self, content: str, fallback: str):
        content = self._clean_report_content(content)
        html_title = self._html_report_title(content)
        if html_title:
            return html_title[:80]
        for line in str(content or "").splitlines():
            text = line.strip()
            if text.startswith("#"):
                return text.lstrip("#").strip()[:80] or fallback
        return fallback

    def _report_intro(self, content: str):
        text = self._plain_report_text(content)
        text = re.sub(r"[*_>`|#-]+", "", text)
        text = re.sub(r"\s+", " ", text).strip()
        return (text[:118] + "...") if len(text) > 118 else text

    def _clean_report_content(self, content: str):
        text = str(content or "").strip()
        fence = re.match(r"^```(?:html|HTML)?\s*(.*?)\s*```$", text, re.S)
        return fence.group(1).strip() if fence else text

    def _report_content_format(self, content: str):
        text = self._clean_report_content(content).lstrip()
        return "html" if re.search(r"<!doctype\s+html|<html[\s>]|<body[\s>]|<section[\s>]|<article[\s>]|<div[\s>]", text, re.I) else "markdown"

    def _html_report_title(self, content: str):
        text = self._clean_report_content(content)
        for pattern in (r"<title[^>]*>(.*?)</title>", r"<h1[^>]*>(.*?)</h1>"):
            match = re.search(pattern, text, re.I | re.S)
            if match:
                title = re.sub(r"<[^>]+>", "", match.group(1))
                return html_lib.unescape(re.sub(r"\s+", " ", title).strip())
        return ""

    def _plain_report_text(self, content: str):
        text = self._clean_report_content(content)
        if self._report_content_format(text) == "html":
            text = re.sub(r"<(script|style)[^>]*>.*?</\1>", " ", text, flags=re.I | re.S)
            text = re.sub(r"<[^>]+>", " ", text)
            return html_lib.unescape(text)
        return re.sub(r"#+\s*", "", text)

    def _extract_360_report_person_name(self, content: str, instruction: str = ""):
        text = f"{self._plain_report_text(content)}\n{instruction or ''}"
        text = re.sub(r"\s+", " ", text)
        blocked = {"个人", "员工", "人才", "领导力", "执行层", "战术层", "管理层", "报告", "评估", "通俗", "解读"}
        prefix_pattern = r"^(?:生成|输出|撰写|查看|打开|分析|制作|形成|产出)"
        patterns = [
            r"(?:姓名|对象|被评估人|人员|报告对象)[:：]\s*([\u4e00-\u9fa5]{2,4})",
            r"([\u4e00-\u9fa5]{2,4})\s*(?:-|_|\s)?(?:执行层|战术层|管理层)?(?:领导力)?\s*360",
            r"([\u4e00-\u9fa5]{2,4})的?\s*360[°度]?",
        ]
        for pattern in patterns:
            for match in re.finditer(pattern, text, re.I):
                name = match.group(1).strip()
                name = re.sub(prefix_pattern, "", name)
                if name and name not in blocked and not any(word in name for word in blocked):
                    return name
        return ""

    def _report_record(self, content: str, instruction: str, report_type: str):
        content = self._clean_report_content(content)
        preset = REPORT_PRESETS.get(report_type or "") or REPORT_PRESETS["talent-review"]
        title = self._report_title(content, preset["name"])
        return self._normalized_report_record({
            "id": uuid.uuid4().hex,
            "title": title,
            "intro": self._report_intro(content) or preset["description"],
            "content": content,
            "contentFormat": self._report_content_format(content),
            "instruction": instruction,
            "reportType": report_type or "talent-review",
            "reportTypeName": preset["name"],
            "updatedAt": datetime.now(LOCAL_TZ).isoformat(timespec="seconds"),
            "source": "ai",
        })

    def _safe_report_filename(self, report):
        title = re.sub(r"[\\/:*?\"<>|]+", "_", str(report.get("title") or "report")).strip(" ._")
        title = re.sub(r"\s+", "_", title)[:60] or "report"
        timestamp = re.sub(r"\D+", "", str(report.get("updatedAt") or ""))[:14] or datetime.now(LOCAL_TZ).strftime("%Y%m%d%H%M%S")
        report_id = str(report.get("id") or uuid.uuid4().hex)[:12]
        return f"{timestamp}-{title}-{report_id}.md"

    def _safe_report_html_filename(self, report):
        return re.sub(r"\.md$", ".html", self._safe_report_filename(report), flags=re.I)

    def _html_to_markdown_text(self, content: str):
        text = self._clean_report_content(content)
        text = re.sub(r"<(script|style)[^>]*>.*?</\1>", "", text, flags=re.I | re.S)

        def heading(match):
            level = int(match.group(1))
            body = html_lib.unescape(re.sub(r"<[^>]+>", "", match.group(2))).strip()
            return f"\n{'#' * min(level, 6)} {body}\n"

        text = re.sub(r"<h([1-6])[^>]*>(.*?)</h\1>", heading, text, flags=re.I | re.S)
        text = re.sub(r"<li[^>]*>", "\n- ", text, flags=re.I)
        text = re.sub(r"</(p|div|section|article|tr|ul|ol|table)>", "\n", text, flags=re.I)
        text = re.sub(r"<br\s*/?>", "\n", text, flags=re.I)
        text = re.sub(r"</t[dh]>\s*<t[dh][^>]*>", " | ", text, flags=re.I)
        text = re.sub(r"<[^>]+>", "", text)
        text = html_lib.unescape(text)
        text = re.sub(r"[ \t]+\n", "\n", text)
        text = re.sub(r"\n{3,}", "\n\n", text)
        return text.strip()

    def _report_markdown_document(self, report):
        content = report.get("content", "")
        body = self._html_to_markdown_text(content) if report.get("contentFormat") == "html" else self._clean_report_content(content)
        title = report.get("title") or report.get("reportTypeName") or "生成报告"
        meta = [
            f"# {title}",
            "",
            f"- 报告类型：{report.get('reportTypeName', '')}",
            f"- 生成时间：{report.get('updatedAt', '')}",
            f"- 材料来源：{report.get('intro', '')}",
            "",
            "---",
            "",
        ]
        return "\n".join(meta) + body.strip() + "\n"

    def _markdown_to_html_fragment(self, markdown: str):
        lines = str(markdown or "").splitlines()
        html = []
        paragraph = []
        list_open = False

        def close_list():
            nonlocal list_open
            if list_open:
                html.append("</ul>")
                list_open = False

        def flush_paragraph():
            nonlocal paragraph
            if paragraph:
                close_list()
                html.append(f"<p>{html_lib.escape(' '.join(paragraph))}</p>")
                paragraph = []

        for line in lines:
            text = line.strip()
            if not text:
                flush_paragraph()
                close_list()
                continue
            heading = re.match(r"^(#{1,6})\s+(.+)$", text)
            if heading:
                flush_paragraph()
                close_list()
                level = min(len(heading.group(1)), 4)
                html.append(f"<h{level}>{html_lib.escape(heading.group(2).strip())}</h{level}>")
                continue
            if re.match(r"^[-*]\s+", text):
                flush_paragraph()
                if not list_open:
                    html.append("<ul>")
                    list_open = True
                item_text = re.sub(r"^[-*]\s+", "", text)
                html.append(f"<li>{html_lib.escape(item_text)}</li>")
                continue
            if text.startswith(">"):
                flush_paragraph()
                close_list()
                html.append(f"<blockquote>{html_lib.escape(text.lstrip('> ').strip())}</blockquote>")
                continue
            paragraph.append(text)
        flush_paragraph()
        close_list()
        return "\n".join(html)

    def _inline_markdown_to_html(self, text: str):
        escaped = html_lib.escape(str(text or ""))
        escaped = re.sub(r"`([^`]+)`", r"<code>\1</code>", escaped)
        escaped = re.sub(r"\*\*([^*]+)\*\*", r"<strong>\1</strong>", escaped)
        return escaped

    def _markdown_table_cells(self, line: str):
        return [cell.strip() for cell in line.strip().strip("|").split("|")]

    def _is_markdown_table_separator(self, line: str):
        return bool(re.match(r"^\s*\|?\s*:?-{3,}:?\s*(\|\s*:?-{3,}:?\s*)+\|?\s*$", line))

    def _report_markdown_body_html(self, markdown: str, skip_first_title: str = ""):
        lines = str(markdown or "").splitlines()
        html = []
        paragraph = []
        list_stack = []
        section_open = False
        index = 0

        def close_lists():
            while list_stack:
                html.append(f"</{list_stack.pop()}>")

        def flush_paragraph():
            if paragraph:
                html.append(f"<p>{self._inline_markdown_to_html(' '.join(paragraph).strip())}</p>")
                paragraph.clear()

        while index < len(lines):
            text = lines[index].strip()
            if not text:
                flush_paragraph()
                close_lists()
                index += 1
                continue
            if text == "---":
                flush_paragraph()
                close_lists()
                html.append("<hr>")
                index += 1
                continue
            if text.startswith("|") and index + 1 < len(lines) and self._is_markdown_table_separator(lines[index + 1]):
                flush_paragraph()
                close_lists()
                header = self._markdown_table_cells(text)
                index += 2
                rows = []
                while index < len(lines) and lines[index].strip().startswith("|"):
                    rows.append(self._markdown_table_cells(lines[index].strip()))
                    index += 1
                html.append("<div class=\"table-wrap\"><table><thead><tr>")
                html.append("".join(f"<th>{self._inline_markdown_to_html(cell)}</th>" for cell in header))
                html.append("</tr></thead><tbody>")
                for row in rows:
                    html.append("<tr>")
                    html.append("".join(f"<td>{self._inline_markdown_to_html(cell)}</td>" for cell in row))
                    html.append("</tr>")
                html.append("</tbody></table></div>")
                continue
            heading = re.match(r"^(#{1,3})\s+(.+)$", text)
            if heading:
                flush_paragraph()
                close_lists()
                level = len(heading.group(1))
                title = heading.group(2).strip()
                if skip_first_title and title == skip_first_title:
                    index += 1
                    continue
                if level == 1:
                    if section_open:
                        html.append("</section>")
                    html.append(f"<section class=\"report-section\"><h2>{self._inline_markdown_to_html(title)}</h2>")
                    section_open = True
                elif level == 2:
                    html.append(f"<h3>{self._inline_markdown_to_html(title)}</h3>")
                else:
                    html.append(f"<h4>{self._inline_markdown_to_html(title)}</h4>")
                index += 1
                continue
            if text.startswith(">"):
                flush_paragraph()
                close_lists()
                html.append(f"<blockquote>{self._inline_markdown_to_html(text.lstrip('> ').strip())}</blockquote>")
                index += 1
                continue
            bullet = re.match(r"^[-*]\s+(.+)$", text)
            if bullet:
                flush_paragraph()
                if list_stack and list_stack[-1] != "ul":
                    close_lists()
                if not list_stack:
                    html.append("<ul>")
                    list_stack.append("ul")
                html.append(f"<li>{self._inline_markdown_to_html(bullet.group(1))}</li>")
                index += 1
                continue
            ordered = re.match(r"^\d+[.)]\s+(.+)$", text)
            if ordered:
                flush_paragraph()
                if list_stack and list_stack[-1] != "ol":
                    close_lists()
                if not list_stack:
                    html.append("<ol>")
                    list_stack.append("ol")
                html.append(f"<li>{self._inline_markdown_to_html(ordered.group(1))}</li>")
                index += 1
                continue
            paragraph.append(text)
            index += 1
        flush_paragraph()
        close_lists()
        if section_open:
            html.append("</section>")
        return "\n".join(html)

    def _first_markdown_section_text(self, markdown: str, heading: str):
        pattern = rf"^#\s+{re.escape(heading)}\s*$"
        lines = str(markdown or "").splitlines()
        start = -1
        for idx, line in enumerate(lines):
            if re.match(pattern, line.strip()):
                start = idx + 1
                break
        if start < 0:
            return ""
        chunk = []
        for line in lines[start:]:
            if line.startswith("# "):
                break
            if line.strip():
                chunk.append(line.strip())
        return re.sub(r"\s+", " ", " ".join(chunk)).strip()

    def _clean_metric_value(self, value: str, limit=40):
        text = re.sub(r"[*`]+", "", str(value or "")).strip()
        text = re.split(r"\s{2,}|[。；;]\s*", text, 1)[0].strip()
        return text[:limit]

    def _extract_360_text_metric(self, markdown: str, label: str, limit=40):
        pattern = rf"(?m)^\s*(?:[-*]\s*)?(?:\*\*)?\s*{re.escape(label)}\s*(?:\*\*)?\s*[：:]\s*(.+?)\s*$"
        match = re.search(pattern, markdown or "")
        if match:
            return self._clean_metric_value(match.group(1), limit)
        if label == "排名":
            rank_match = re.search(r"(?:排名|项目中排名)\s*[：:]?\s*([0-9,]+\s*/\s*[0-9,]+)", markdown or "")
            if rank_match:
                return self._clean_metric_value(rank_match.group(1), limit)
        return ""

    def _extract_360_score_metric(self, markdown: str, label: str):
        value = self._extract_360_text_metric(markdown, label, limit=16)
        match = re.search(r"([0-5](?:\.\d{1,2})?)", value)
        if match:
            return float(match.group(1))
        alias_patterns = {
            "他评总分": [
                r"(?:总体他评|整体他评|他评总分|你的他评(?:总分)?|他评)\s*[：:]\s*\**([0-5](?:\.\d{1,2})?)",
                r"他评(?:总分)?是\s*\**([0-5](?:\.\d{1,2})?)",
            ],
            "自评": [
                r"自评\s*[：:]\s*\**([0-5](?:\.\d{1,2})?)",
                r"自评(?:总分)?是\s*\**([0-5](?:\.\d{1,2})?)",
            ],
        }
        for pattern in alias_patterns.get(label, []):
            alias_match = re.search(pattern, markdown or "", re.I)
            if alias_match:
                return float(alias_match.group(1))
        return None

    def _extract_360_role_scores(self, markdown: str):
        scores = {}
        for label in ("上级", "同事", "下级", "自评"):
            score = self._extract_360_score_metric(markdown, label)
            if score is None:
                patterns = [
                    rf"{re.escape(label)}(?:评分|总分)?\s*[：:]\s*\**([0-5](?:\.\d{{1,2}})?)",
                    rf"{re.escape(label)}(?:评分|总分)?\s+\**([0-5](?:\.\d{{1,2}})?)",
                    rf"{re.escape(label)}是\s*\**([0-5](?:\.\d{{1,2}})?)",
                ]
                for pattern in patterns:
                    match = re.search(pattern, markdown or "", re.I)
                    if match:
                        score = float(match.group(1))
                        break
            if score is not None:
                scores[label] = score
        return scores

    def _extract_360_capability_scores(self, markdown: str):
        blocked = {"他评总分", "自评", "上级", "同事", "下级", "排名", "下级与上级差距"}
        rows = []
        patterns = [
            r"能力项[：:]\s*([^—\-\n：:]{2,24})[—\-]+他评\s*([0-5](?:\.\d{1,2})?)(?:[^\n]{0,24}?自评\s*([0-5](?:\.\d{1,2})?))?",
            r"^\s*(?:[-*]\s*)?([^：:\n—\-]{2,24})[：:]\s*他评\s*([0-5](?:\.\d{1,2})?)(?:[^\n]{0,24}?自评\s*([0-5](?:\.\d{1,2})?))?",
            r"^\s*(?:[-*]\s*)?([^：:\n—\-]{2,24})[—\-]+他评\s*([0-5](?:\.\d{1,2})?)(?:[^\n]{0,24}?自评\s*([0-5](?:\.\d{1,2})?))?",
        ]
        for pattern in patterns:
            for match in re.finditer(pattern, markdown or "", re.M):
                name = re.sub(r"[*`]+", "", match.group(1)).strip()
                if not name or name in blocked or len(name) > 24:
                    continue
                score = float(match.group(2))
                self_score = match.group(3)
                rows.append({"name": name, "score": score, "selfScore": float(self_score) if self_score else None})
        unique = {}
        for row in rows:
            unique.setdefault(row["name"], row)
        return list(unique.values())

    def _score_row_html(self, label: str, score: float, color: str):
        width = max(0, min(100, score / 5 * 100))
        return (
            "<div class=\"score-row\">"
            f"<span>{html_lib.escape(label)}</span>"
            f"<div class=\"score-track\"><i style=\"width:{width:.1f}%;background:{html_lib.escape(color)}\"></i></div>"
            f"<strong>{score:.2f}</strong>"
            "</div>"
        )

    def _chip_card_html(self, name: str, score: str, text: str, style: str):
        return (
            f"<div class=\"chip-card {style}\">"
            f"<b>{html_lib.escape(name)}</b><span>{html_lib.escape(score)}</span><p>{html_lib.escape(text)}</p>"
            "</div>"
        )

    def _quadrant_cell_html(self, title: str, style: str, items):
        tags = "".join(f"<span>{html_lib.escape(item)}</span>" for item in items)
        return f"<div class=\"quadrant-cell {style}\"><h4>{html_lib.escape(title)}</h4><div class=\"quadrant-tags\">{tags}</div></div>"

    def _report_360_html_document(self, report):
        title = report.get("title") or "360报告解读"
        markdown = self._clean_report_content(report.get("content") or report.get("mdContent", ""))
        first_title = ""
        first_heading = re.search(r"^#\s+(.+)$", markdown, re.M)
        if first_heading:
            first_title = first_heading.group(1).strip()
        summary = self._first_markdown_section_text(markdown, "一句话总结")
        body = self._report_markdown_body_html(markdown, skip_first_title=first_title)
        total_score = self._extract_360_score_metric(markdown, "他评总分")
        rank = self._extract_360_text_metric(markdown, "排名", limit=24)
        role_scores = self._extract_360_role_scores(markdown)
        capability_scores = self._extract_360_capability_scores(markdown)
        metrics = []
        if total_score is not None:
            metrics.append(("他评总分", f"{total_score:.2f}", "从报告正文识别"))
        if rank:
            metrics.append(("排名", rank, "从报告正文识别"))
        if role_scores.get("自评") is not None:
            metrics.append(("自评", f"{role_scores['自评']:.2f}", "从报告正文识别"))
        role_metric_values = [f"{role_scores[label]:.2f}" for label in ("上级", "同事", "下级") if role_scores.get(label) is not None]
        if role_metric_values:
            metrics.append(("上级 / 同事 / 下级", " / ".join(role_metric_values), "从报告正文识别"))
        if not metrics:
            metrics.append(("关键指标", "未识别", "请确认正文包含他评总分、排名或角色评分"))
        metrics_html = "".join(
            "<article class=\"metric-card\">"
            f"<span>{html_lib.escape(name)}</span><strong>{html_lib.escape(value)}</strong><em>{html_lib.escape(note)}</em>"
            "</article>"
            for name, value, note in metrics
        )
        score_colors = {"上级": "#f59e0b", "同事": "#2563eb", "下级": "#0f9f8f", "自评": "#e85d75"}
        role_html = "".join(
            self._score_row_html(label, role_scores[label], score_colors[label])
            for label in ("上级", "同事", "下级", "自评")
            if role_scores.get(label) is not None
        )
        role_html = role_html or "<p class=\"muted-note\">未从正文识别到上级、同事、下级或自评分数。</p>"
        ranked_capabilities = sorted(capability_scores, key=lambda item: item["score"], reverse=True)
        strength_html = "".join(
            self._chip_card_html(item["name"], f"{item['score']:.2f}", "正文识别的高分能力项", "positive")
            for item in ranked_capabilities[:3]
        )
        development_html = "".join(
            self._chip_card_html(
                item["name"],
                f"{item['score']:.2f}" + (f" / 自评{item['selfScore']:.2f}" if item.get("selfScore") is not None else ""),
                "正文识别的发展关注项",
                "caution",
            )
            for item in sorted(capability_scores, key=lambda item: item["score"])[:3]
        )
        capability_html = (
            f"<div class=\"chip-grid\">{strength_html}</div><div style=\"height:10px\"></div><div class=\"chip-grid\">{development_html}</div>"
            if strength_html or development_html
            else "<p class=\"muted-note\">未从正文识别到能力项分数，详细内容请查看下方报告正文。</p>"
        )
        css = (
            ":root{--bg:#f5f7fb;--surface:#fff;--ink:#18202f;--muted:#657084;--line:#dfe5ef;--blue:#2457d6;--blue-soft:#eaf1ff;--teal:#0f9f8f;--teal-soft:#e9f8f5;--amber:#c97800;--amber-soft:#fff3dd;--rose:#c84662;--rose-soft:#fff0f3;--slate:#253044;--shadow:0 18px 48px rgba(24,32,47,.10)}"
            "*{box-sizing:border-box}body{margin:0;background:var(--bg);color:var(--ink);font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','PingFang SC','Microsoft YaHei',Arial,sans-serif;font-size:15px;line-height:1.72}.report-page{max-width:1180px;margin:0 auto;padding:28px 28px 64px}"
            ".hero{display:grid;grid-template-columns:minmax(0,1.15fr) minmax(360px,.85fr);gap:18px;align-items:stretch;margin-bottom:18px}.hero-main,.hero-panel,.report-section,.insight-panel,.quadrant-panel{border:1px solid var(--line);border-radius:18px;background:var(--surface);box-shadow:var(--shadow)}"
            ".hero-main{padding:28px;background:linear-gradient(135deg,#131a28,#24324b 58%,#1e4d6f);color:#fff}.kicker{margin:0 0 14px;color:#9edbd4;font-size:13px;font-weight:800}h1{margin:0;font-size:34px;line-height:1.18;letter-spacing:0}.subtitle{margin:10px 0 0;max-width:64ch;color:rgba(255,255,255,.78)}"
            ".summary-box{margin-top:22px;padding:18px;border-radius:14px;background:rgba(255,255,255,.10);border:1px solid rgba(255,255,255,.18)}.summary-box b{display:block;margin-bottom:6px;color:#b7f0e8}.summary-box p{margin:0;color:rgba(255,255,255,.90)}.hero-panel{display:grid;gap:12px;padding:18px}"
            ".metric-card{padding:14px;border:1px solid var(--line);border-radius:14px;background:#fbfcff}.metric-card span{display:block;color:var(--muted);font-size:13px;font-weight:700}.metric-card strong{display:block;margin-top:4px;color:var(--blue);font-size:26px;line-height:1.15}.metric-card em{display:block;margin-top:6px;color:var(--muted);font-style:normal;font-size:12px}"
            ".insight-grid{display:grid;grid-template-columns:1fr 1fr;gap:18px;margin:18px 0}.insight-panel{padding:20px}.insight-panel h2,.report-section h2{margin:0 0 14px;font-size:22px;line-height:1.35}.score-row{display:grid;grid-template-columns:56px minmax(0,1fr) 46px;gap:10px;align-items:center;margin:12px 0}.score-row span{color:var(--muted);font-weight:700}.score-row strong{text-align:right}.score-track{height:12px;border-radius:999px;background:#edf1f7;overflow:hidden}.score-track i{display:block;height:100%;border-radius:inherit}"
            ".chip-grid{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:10px}.chip-card{min-height:128px;padding:14px;border-radius:14px;border:1px solid var(--line)}.chip-card b{display:block;font-size:15px}.chip-card span{display:inline-flex;margin:8px 0;padding:2px 8px;border-radius:999px;font-size:12px;font-weight:800}.chip-card p{margin:0;color:var(--muted);font-size:13px}.chip-card.positive{background:var(--teal-soft);border-color:#bfe8e0}.chip-card.positive span{color:#08766b;background:rgba(15,159,143,.12)}.chip-card.caution{background:var(--amber-soft);border-color:#f3d29b}.chip-card.caution span{color:#9b5700;background:rgba(201,120,0,.12)}"
            ".muted-note{margin:0;color:var(--muted);font-size:13px}"
            ".quadrant-panel{margin:18px 0;padding:22px}.quadrant-head{display:flex;justify-content:space-between;gap:16px;align-items:flex-end;margin-bottom:14px}.quadrant-head h2{margin:0;font-size:22px}.quadrant-head p{margin:0;color:var(--muted);font-size:13px}.quadrant-wrap{display:grid;grid-template-columns:42px 1fr;gap:10px;align-items:stretch}.axis-y{writing-mode:vertical-rl;transform:rotate(180deg);display:grid;place-items:center;color:var(--muted);font-size:12px;font-weight:800}.quadrant-grid{position:relative;display:grid;grid-template-columns:1fr 1fr;gap:10px}.quadrant-grid:before,.quadrant-grid:after{content:'';position:absolute;background:rgba(37,48,68,.18)}.quadrant-grid:before{width:1px;top:0;bottom:0;left:50%}.quadrant-grid:after{height:1px;left:0;right:0;top:50%}"
            ".quadrant-cell{min-height:154px;padding:16px;border-radius:14px;border:1px solid var(--line);background:#fbfcff}.quadrant-cell h4{margin:0 0 12px;font-size:15px}.q-good{background:var(--teal-soft)}.q-watch{background:var(--rose-soft)}.q-dev{background:var(--amber-soft)}.q-muted{background:#f2f5fa}.quadrant-tags{display:flex;gap:8px;flex-wrap:wrap}.quadrant-tags span{padding:6px 9px;border-radius:999px;background:rgba(255,255,255,.72);border:1px solid rgba(24,32,47,.10);font-size:12px;font-weight:750}.axis-x{margin:8px 0 0 52px;display:flex;justify-content:space-between;color:var(--muted);font-size:12px;font-weight:800}"
            ".report-section{margin-top:18px;padding:24px}.report-section h3{margin:24px 0 10px;font-size:18px;color:var(--blue)}.report-section h4{margin:18px 0 8px;font-size:15px;color:var(--slate)}p{margin:10px 0}ul,ol{margin:10px 0 0 22px;padding:0}li{margin:6px 0}strong{color:#111827}code{padding:2px 6px;border-radius:6px;background:#edf1f7;color:#334155}blockquote{margin:14px 0;padding:14px 16px;border:1px solid #c9d8ff;border-radius:12px;background:var(--blue-soft);color:#1c3f97;font-weight:650}.table-wrap{overflow-x:auto;margin:14px 0;border:1px solid var(--line);border-radius:14px}table{width:100%;min-width:760px;border-collapse:collapse;background:#fff}th,td{padding:12px 14px;border-bottom:1px solid var(--line);text-align:left;vertical-align:top}th{background:#f3f6fb;color:#344054;font-size:13px}td{color:#344054}tr:last-child td{border-bottom:0}hr{border:0;border-top:1px solid var(--line);margin:18px 0}"
            "@media(max-width:860px){.report-page{padding:16px}.hero,.insight-grid{grid-template-columns:1fr}.chip-grid{grid-template-columns:1fr}h1{font-size:28px}}@media print{body{background:#fff}.report-page{max-width:none;padding:0}.hero-main,.hero-panel,.report-section,.insight-panel,.quadrant-panel{box-shadow:none;break-inside:avoid}}"
        )
        return (
            "<!doctype html><html lang=\"zh-CN\"><head><meta charset=\"utf-8\">"
            "<meta name=\"viewport\" content=\"width=device-width, initial-scale=1\">"
            f"<title>{html_lib.escape(title)}</title><style>{css}</style></head><body><main class=\"report-page\">"
            "<section class=\"hero\"><div class=\"hero-main\">"
            "<p class=\"kicker\">360°评估解读 · HRBP 一对一沟通版</p>"
            f"<h1>{html_lib.escape(title)}</h1>"
            "<p class=\"subtitle\">基于360个人报告标准版、核心版及开放性反馈整理。页面保留原始 MD 框架，并将关键分数、角色差异和行动重点前置呈现。</p>"
            f"<div class=\"summary-box\"><b>一句话总结</b><p>{self._inline_markdown_to_html(summary)}</p></div>"
            f"</div><aside class=\"hero-panel\" aria-label=\"关键指标\">{metrics_html}</aside></section>"
            f"<section class=\"insight-grid\"><div class=\"insight-panel\"><h2>角色评分对比</h2>{role_html}</div>"
            f"<div class=\"insight-panel\"><h2>优势与发展抓手</h2>{capability_html}</div></section>"
            f"{body}</main></body></html>"
        )

    def _normalize_html_document(self, content: str, title: str = "生成报告"):
        html = self._clean_report_content(content)
        if re.search(r"<!doctype\s+html|<html[\s>]", html, re.I):
            return html
        return (
            "<!doctype html><html lang=\"zh-CN\"><head><meta charset=\"utf-8\">"
            "<meta name=\"viewport\" content=\"width=device-width, initial-scale=1\">"
            f"<title>{html_lib.escape(title)}</title>"
            "<style>"
            "body{margin:0;background:#f5f7fb;color:#172033;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','PingFang SC','Microsoft YaHei',Arial,sans-serif;line-height:1.75;}"
            "main{max-width:920px;margin:0 auto;padding:48px 28px 72px;background:#fff;min-height:100vh;box-shadow:0 24px 64px rgba(16,24,40,.08);}"
            "h1{font-size:30px;line-height:1.25;margin:0 0 18px;color:#101217;}h2{font-size:22px;margin:32px 0 12px;color:#101217;}h3{font-size:18px;margin:24px 0 10px;color:#101217;}"
            "p,li{font-size:15px;color:#344054;}ul{padding-left:22px;}blockquote{margin:16px 0;padding:12px 16px;border-left:4px solid #2457d6;background:#f4f7ff;color:#344054;}"
            ".meta{margin:0 0 28px;padding:12px 14px;border:1px solid #e4e9f2;border-radius:8px;background:#f8fafc;color:#667085;font-size:13px;}"
            "</style></head><body>"
            f"<main><h1>{html_lib.escape(title)}</h1>{html}</main>"
            "</body></html>"
        )

    def _report_html_document(self, report):
        title = report.get("title") or report.get("reportTypeName") or "生成报告"
        if report.get("contentFormat") == "html":
            return self._normalize_html_document(report.get("content", ""), title)
        if report.get("reportType") == "360":
            return self._report_360_html_document(report)
        body = self._markdown_to_html_fragment(self._clean_report_content(report.get("content", "")))
        meta = (
            f"<div class=\"meta\">报告类型：{html_lib.escape(report.get('reportTypeName', ''))}"
            f"　生成时间：{html_lib.escape(report.get('updatedAt', ''))}"
            f"　材料来源：{html_lib.escape(report.get('intro', ''))}</div>"
        )
        return self._normalize_html_document(meta + body, title)

    def _resolve_report_path(self, path_value):
        if not path_value:
            return None
        path = Path(path_value)
        if not path.is_absolute():
            path = self.data_dir / path
        return path

    def _report_path_for_storage(self, path_value):
        path = self._resolve_report_path(path_value)
        if not path:
            return ""
        try:
            return str(path.resolve().relative_to(self.data_dir.resolve()))
        except ValueError:
            return str(path)

    def _report_record_for_storage(self, report):
        stored = dict(report)
        for key in ("mdPath", "htmlPath"):
            if stored.get(key):
                stored[key] = self._report_path_for_storage(stored[key])
        return stored

    def _ensure_report_markdown_file(self, report):
        if not report.get("id"):
            return report
        path_value = report.get("mdPath")
        path = self._resolve_report_path(path_value)
        if not path or not path.exists():
            self.report_markdown_dir.mkdir(parents=True, exist_ok=True)
            path = self.report_markdown_dir / self._safe_report_filename(report)
            path.write_text(self._report_markdown_document(report), encoding="utf-8")
        return {**report, "mdPath": str(path)}

    def _with_report_html_content(self, report):
        if report.get("reportType") == "360":
            return {**report, "htmlContent": self._report_html_document(report)}
        html_path = self._resolve_report_path(report.get("htmlPath", ""))
        if html_path and html_path.exists() and html_path.is_file():
            try:
                return {**report, "htmlContent": html_path.read_text(encoding="utf-8")}
            except UnicodeDecodeError:
                return {**report, "htmlContent": html_path.read_text(encoding="utf-8", errors="ignore")}
        return report

    def _with_report_markdown_content(self, report):
        report = self._ensure_report_markdown_file(report)
        md_path = self._resolve_report_path(report.get("mdPath", ""))
        if md_path and md_path.exists() and md_path.is_file():
            try:
                return {**report, "mdContent": md_path.read_text(encoding="utf-8")}
            except UnicodeDecodeError:
                return {**report, "mdContent": md_path.read_text(encoding="utf-8", errors="ignore")}
        return report

    def _normalized_report_record(self, report):
        if not isinstance(report, dict):
            return {}
        content = self._clean_report_content(report.get("content", ""))
        report_type = report.get("reportType") or "talent-review"
        preset = REPORT_PRESETS.get(report_type) or REPORT_PRESETS["talent-review"]
        content_format = report.get("contentFormat") or self._report_content_format(content)
        intro = report.get("intro") or ""
        if content_format == "html" and re.search(r"```|<!doctype\s+html|<html[\s>]", intro, re.I):
            intro = ""
        title = report.get("title") or self._report_title(content, preset["name"])
        if report_type == "360":
            person_name = self._extract_360_report_person_name(content, report.get("instruction", ""))
            title = f"{person_name}360报告解读" if person_name else "360报告解读"
            intro = "参考材料来源"
        return {
            **report,
            "title": title,
            "intro": intro or self._report_intro(content) or preset["description"],
            "content": content,
            "contentFormat": content_format,
            "reportType": report_type,
            "reportTypeName": report.get("reportTypeName") or preset["name"],
            "source": report.get("source", "ai"),
        }

    def generated_reports(self):
        history_path = self._first_existing_path([self.report_history_path, *[folder / "generated_reports.json" for folder in self.legacy_report_dirs]])
        records = self._read_json(history_path, [])
        if isinstance(records, dict):
            records = records.get("reports", [])
        records = [self._normalized_report_record(report) for report in records if isinstance(report, dict)]
        if not records:
            latest_path = self._first_existing_path([self.generated_report_path, *[folder / "generated_report.json" for folder in self.legacy_report_dirs]])
            legacy = self._read_json(latest_path, {"content": "", "updatedAt": "", "source": "none"})
            if legacy.get("content"):
                record = self._normalized_report_record({
                    "id": "legacy-latest",
                    "title": self._report_title(legacy.get("content", ""), "人才盘点报告"),
                    "intro": self._report_intro(legacy.get("content", "")),
                    "content": legacy.get("content", ""),
                    "contentFormat": legacy.get("contentFormat", self._report_content_format(legacy.get("content", ""))),
                    "instruction": legacy.get("instruction", ""),
                    "reportType": legacy.get("reportType", "talent-review"),
                    "reportTypeName": legacy.get("reportTypeName", REPORT_PRESETS["talent-review"]["name"]),
                    "updatedAt": legacy.get("updatedAt", ""),
                    "source": legacy.get("source", "ai"),
                })
                records = [record]
        return sorted([self._ensure_report_markdown_file(report) for report in records], key=lambda item: item.get("updatedAt", ""), reverse=True)

    def generated_report(self, report_id=None):
        reports = self.generated_reports()
        if report_id:
            for report in reports:
                if report.get("id") == report_id:
                    report = self._with_report_markdown_content(report)
                    if report.get("reportType") == "360":
                        return {**report, "htmlContent": self._report_html_document(report)}
                    return self._with_report_html_content(report)
            return {"content": "", "updatedAt": "", "source": "none"}
        if not reports:
            return {"content": "", "updatedAt": "", "source": "none"}
        report = self._with_report_markdown_content(reports[0])
        if report.get("reportType") == "360":
            return {**report, "htmlContent": self._report_html_document(report)}
        return self._with_report_html_content(report)

    def generated_report_list(self):
        return [
            {
                **{key: report.get(key, "") for key in ("id", "title", "intro", "reportType", "reportTypeName", "updatedAt", "source", "contentFormat", "mdPath", "htmlPath", "htmlGeneratedAt")},
                "hasHtml": bool(report.get("htmlPath") and self._resolve_report_path(report.get("htmlPath", "")).exists()),
            }
            for report in self.generated_reports()
        ]

    def save_generated_report(self, content: str, instruction: str, report_type: str = "talent-review"):
        payload = self._report_record(content, instruction, report_type)
        try:
            self.generated_report_path.parent.mkdir(parents=True, exist_ok=True)
            payload = self._ensure_report_markdown_file(payload)
            self.generated_report_path.write_text(json.dumps(self._report_record_for_storage(payload), ensure_ascii=False, indent=2), encoding="utf-8")
            existing = self._read_json(self.report_history_path, [])
            if isinstance(existing, dict):
                existing = existing.get("reports", [])
            reports = [payload] + [report for report in existing if report.get("id") != payload["id"]]
            self.report_history_path.write_text(json.dumps([self._report_record_for_storage(report) for report in reports[:50]], ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as e:
            payload["_saveError"] = f"报告未持久化: {e}"
        return payload

    def delete_generated_report(self, report_id: str):
        report_id = str(report_id or "").strip()
        if not report_id:
            raise ValueError("缺少报告 ID。")
        existing = self._read_json(self.report_history_path, [])
        if isinstance(existing, dict):
            existing = existing.get("reports", [])
        normalized = [self._normalized_report_record(report) for report in existing if isinstance(report, dict)]
        target = next((report for report in normalized if report.get("id") == report_id), None)
        if not target:
            raise FileNotFoundError("未找到要删除的报告。")
        remaining = [report for report in normalized if report.get("id") != report_id]
        md_path = self._resolve_report_path(target.get("mdPath", ""))
        if md_path and md_path.exists() and md_path.is_file():
            md_path.unlink()
        if self.report_markdown_dir.exists():
            for path in self.report_markdown_dir.glob(f"*{report_id[:12]}.md"):
                if path.is_file():
                    path.unlink()
        html_path = self._resolve_report_path(target.get("htmlPath", ""))
        if html_path and html_path.exists() and html_path.is_file():
            html_path.unlink()
        if self.report_html_dir.exists():
            for path in self.report_html_dir.glob(f"*{report_id[:12]}.html"):
                if path.is_file():
                    path.unlink()
        self.report_history_path.write_text(json.dumps([self._report_record_for_storage(report) for report in remaining[:50]], ensure_ascii=False, indent=2), encoding="utf-8")
        if remaining:
            latest = self._ensure_report_markdown_file(remaining[0])
            self.generated_report_path.write_text(json.dumps(self._report_record_for_storage(latest), ensure_ascii=False, indent=2), encoding="utf-8")
        elif self.generated_report_path.exists():
            self.generated_report_path.unlink()
        return {"deleted": True, "id": report_id, "remaining": len(remaining)}

    def generate_report_html(self, report_id: str):
        report_id = str(report_id or "").strip()
        if not report_id:
            raise ValueError("缺少报告 ID。")
        existing = self._read_json(self.report_history_path, [])
        if isinstance(existing, dict):
            existing = existing.get("reports", [])
        records = [self._normalized_report_record(report) for report in existing if isinstance(report, dict)]
        index = next((idx for idx, report in enumerate(records) if report.get("id") == report_id), -1)
        if index < 0:
            raise FileNotFoundError("未找到要生成 HTML 的报告。")
        report = self._ensure_report_markdown_file(records[index])
        self.report_html_dir.mkdir(parents=True, exist_ok=True)
        path = self._resolve_report_path(report.get("htmlPath", "")) if report.get("htmlPath") else self.report_html_dir / self._safe_report_html_filename(report)
        try:
            path.resolve().relative_to(self.report_html_dir.resolve())
        except ValueError:
            path = self.report_html_dir / self._safe_report_html_filename(report)
        path.write_text(self._report_html_document(report), encoding="utf-8")
        updated = {
            **report,
            "htmlPath": str(path),
            "htmlGeneratedAt": datetime.now(LOCAL_TZ).isoformat(timespec="seconds"),
        }
        records[index] = updated
        self.report_history_path.write_text(json.dumps([self._report_record_for_storage(report) for report in records[:50]], ensure_ascii=False, indent=2), encoding="utf-8")
        latest = self._read_json(self.generated_report_path, {}) if self.generated_report_path.exists() else {}
        if latest.get("id") == report_id:
            self.generated_report_path.write_text(json.dumps(self._report_record_for_storage(updated), ensure_ascii=False, indent=2), encoding="utf-8")
        return self._with_report_html_content(updated)

    def ai_config(self):
        if not hasattr(self, "_runtime_ai_keys"):
            self._runtime_ai_keys = {}
        saved = self._read_json(
            self.ai_config_path,
            {
                "apiKey": "",
                "baseUrl": "https://api.openai.com/v1",
                "model": "",
            },
        )
        if not isinstance(saved, dict):
            saved = {}
        secrets = self._read_json(self.ai_secrets_path, {}) if self.ai_secrets_path.exists() else {}
        if not isinstance(secrets, dict):
            secrets = {}

        def group_config(name, legacy=False):
            raw = saved.get(name, {}) if isinstance(saved.get(name, {}), dict) else {}
            secret = secrets.get(name, {}) if isinstance(secrets.get(name, {}), dict) else {}
            if legacy:
                raw = {
                    "apiKey": raw.get("apiKey", saved.get("apiKey", "")),
                    "baseUrl": raw.get("baseUrl", saved.get("baseUrl", "https://api.openai.com/v1")),
                    "model": raw.get("model", saved.get("model", "")),
                }
                secret = {
                    "apiKey": secret.get("apiKey", secrets.get("apiKey", "")) or raw.get("apiKey", ""),
                }
            env_name = "HROBOT_AI_API_KEY" if name == "multimodal" else "HROBOT_IMAGE_API_KEY"
            return {
                "apiKey": os.environ.get(env_name) or secret.get("apiKey", "") or self._runtime_ai_keys.get(name, "") or "",
                "baseUrl": raw.get("baseUrl", "https://api.openai.com/v1"),
                "model": raw.get("model", ""),
            }

        return {
            "multimodal": group_config("multimodal", legacy=True),
            "image": group_config("image"),
            "updatedAt": saved.get("updatedAt", ""),
        }

    def ai_config_status(self):
        config = self.ai_config()
        def status_for(group):
            return {
                "configured": bool(group.get("apiKey") and group.get("baseUrl") and group.get("model")),
                "baseUrl": group.get("baseUrl", ""),
                "model": group.get("model", ""),
            }

        return {
            "multimodal": status_for(config["multimodal"]),
            "image": status_for(config["image"]),
            "updatedAt": config.get("updatedAt", ""),
        }

    def mcp_config(self):
        saved = self._read_json(
            self.mcp_config_path,
            {
                "url": "",
                "headers": {},
                "authHeaderName": "x-api-key",
            },
        )
        if not isinstance(saved, dict):
            saved = {}
        headers = saved.get("headers", {}) if isinstance(saved.get("headers", {}), dict) else {}
        auth_header_name = str(saved.get("authHeaderName") or "x-api-key").strip() or "x-api-key"
        api_key = os.environ.get("HROBOT_MCP_API_KEY") or headers.get(auth_header_name, "") or headers.get("x-api-key", "")
        return {
            "url": os.environ.get("HROBOT_MCP_URL") or saved.get("url", ""),
            "authHeaderName": auth_header_name,
            "headers": {
                **{key: value for key, value in headers.items() if value},
                **({auth_header_name: api_key} if api_key else {}),
            },
        }

    def mcp_config_status(self):
        config = self.mcp_config()
        auth_header_name = config.get("authHeaderName") or "x-api-key"
        api_key = config.get("headers", {}).get(auth_header_name, "")
        return {
            "url": config.get("url", ""),
            "authHeaderName": auth_header_name,
            "authConfigured": bool(api_key),
            "configured": bool(config.get("url")),
        }

    def save_mcp_config(self, config):
        current = self.mcp_config()
        auth_header_name = str(config.get("authHeaderName") or current.get("authHeaderName") or "x-api-key").strip() or "x-api-key"
        url = str(config.get("url") or current.get("url") or "").strip()
        incoming_key = str(config.get("apiKey") or "").strip()
        headers = dict(current.get("headers", {}))
        headers = {key: value for key, value in headers.items() if key != current.get("authHeaderName")}
        if incoming_key:
            headers[auth_header_name] = incoming_key
        elif current.get("headers", {}).get(auth_header_name):
            headers[auth_header_name] = current["headers"][auth_header_name]
        payload = {
            "url": url,
            "authHeaderName": auth_header_name,
            "headers": headers,
            "updatedAt": datetime.now(LOCAL_TZ).isoformat(timespec="seconds"),
        }
        self.mcp_config_path.parent.mkdir(parents=True, exist_ok=True)
        self.mcp_config_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return self.mcp_config_status()

    def data_source_config(self):
        saved = self._read_json(self.data_source_config_path, {}) if self.data_source_config_path.exists() else {}
        if not isinstance(saved, dict):
            saved = {}
        local_folder = str(saved.get("localFolder") or "").strip()
        return {
            "localFolder": local_folder,
            "localFolderExists": bool(local_folder and Path(local_folder).expanduser().exists()),
            "updatedAt": str(saved.get("updatedAt") or ""),
            "mcp": self.mcp_config_status(),
        }

    def save_data_source_config(self, config):
        current = self.data_source_config()
        local_folder = str(config.get("localFolder") or current.get("localFolder") or "").strip()
        payload = {
            "localFolder": local_folder,
            "updatedAt": datetime.now(LOCAL_TZ).isoformat(timespec="seconds"),
        }
        self.data_source_config_path.parent.mkdir(parents=True, exist_ok=True)
        self.data_source_config_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return self.data_source_config()

    def _preview_table_headers(self, path: Path):
        suffix = path.suffix.lower()
        try:
            if suffix in {".xlsx", ".xlsm"}:
                from openpyxl import load_workbook
                workbook = load_workbook(path, read_only=True, data_only=True)
                sheet = workbook.active
                headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(max_row=1), [])]
                workbook.close()
                return [item for item in headers if item][:30]
            if suffix == ".csv":
                with path.open("r", encoding="utf-8-sig", newline="") as handle:
                    row = next(csv.reader(handle), [])
                return [str(item or "").strip() for item in row if str(item or "").strip()][:30]
            if suffix == ".json":
                data = self._read_json(path, [])
                sample = data[0] if isinstance(data, list) and data else data
                return list(sample.keys())[:30] if isinstance(sample, dict) else []
        except Exception as error:
            _agent_debug_log(
                "server.py:_preview_table_headers",
                "failed to preview local source file",
                {"path": str(path), "errorType": type(error).__name__, "error": str(error)},
                "DATA_SOURCE_SCAN",
            )
        return []

    def _classify_local_source_file(self, path: Path, headers):
        text = " ".join([path.name, *headers]).lower()
        rules = [
            ("talent_review", ["九宫格", "人才盘点", "校准", "潜能", "无成长"]),
            ("feedback_360", ["360", "他评", "自评", "上级", "同事", "下级"]),
            ("performance", ["绩效", "performance", "年度评价", "考核"]),
            ("employee_profile", ["员工", "工号", "姓名", "组织", "职级", "职位", "入职"]),
            ("permission_scope", ["权限", "hrbp", "可见", "scope", "permission"]),
            ("project_evaluation", ["项目", "复盘", "评价", "里程碑"]),
            ("report_material", ["报告", "材料", "访谈", "纪要", "md", "markdown"]),
        ]
        for kind, keywords in rules:
            if any(keyword.lower() in text for keyword in keywords):
                return kind
        return "unknown"

    def _summarize_local_data_scan(self, files, folder: Path):
        type_labels = {
            "talent_review": "人才盘点",
            "feedback_360": "干部360",
            "performance": "绩效",
            "employee_profile": "员工主档",
            "permission_scope": "权限范围",
            "project_evaluation": "项目评价",
            "report_material": "报告材料",
            "unknown": "待确认",
        }
        type_counts = {}
        extension_counts = {}
        folder_counts = {}
        duplicate_names = {}
        duplicate_size_names = {}
        for item in files:
            detected_type = item.get("detectedType") or "unknown"
            type_counts[detected_type] = type_counts.get(detected_type, 0) + 1
            extension = item.get("extension") or ""
            extension_counts[extension] = extension_counts.get(extension, 0) + 1
            parts = Path(item.get("relativePath") or item.get("name") or "").parts
            top_folder = parts[0] if len(parts) > 1 else "根目录"
            folder_counts[top_folder] = folder_counts.get(top_folder, 0) + 1
            duplicate_names.setdefault(str(item.get("name") or "").lower(), []).append(item)
            duplicate_size_names.setdefault((str(item.get("name") or "").lower(), item.get("size")), []).append(item)

        known_count = sum(count for kind, count in type_counts.items() if kind != "unknown")
        unknown_count = type_counts.get("unknown", 0)
        readable_types = [
            f"{type_labels.get(kind, kind)} {count} 个"
            for kind, count in sorted(type_counts.items(), key=lambda pair: (-pair[1], pair[0]))
            if count
        ]
        top_folders = [
            f"{name}（{count}）"
            for name, count in sorted(folder_counts.items(), key=lambda pair: (-pair[1], pair[0]))[:6]
        ]
        duplicate_name_groups = [
            {"name": items[0].get("name"), "count": len(items)}
            for items in duplicate_names.values()
            if len(items) > 1 and items[0].get("name")
        ][:8]
        likely_duplicate_groups = [
            {"name": items[0].get("name"), "count": len(items)}
            for items in duplicate_size_names.values()
            if len(items) > 1 and items[0].get("name")
        ][:8]

        opportunities = []
        if type_counts.get("employee_profile"):
            opportunities.append("员工画像与当前组织/职级/岗位分布")
        if type_counts.get("talent_review"):
            opportunities.append("年度人才盘点、九宫格分布、校准差异和人才池分析")
        if type_counts.get("feedback_360"):
            opportunities.append("干部360反馈、能力项强弱和上下级视角差异")
        if type_counts.get("performance"):
            opportunities.append("绩效趋势、绩效与九宫格/潜力交叉分析")
        if type_counts.get("project_evaluation"):
            opportunities.append("项目经历、项目评价和关键岗位人才复盘")
        if type_counts.get("permission_scope"):
            opportunities.append("HRBP 权限范围与可见数据边界校验")
        if not opportunities and files:
            opportunities.append("需要先确认文件类型和字段映射，再决定可做的分析")

        structure_issues = []
        if unknown_count:
            structure_issues.append(f"有 {unknown_count} 个文件暂未识别类型，可能需要补充命名规范或字段映射。")
        if folder_counts.get("根目录", 0) > max(5, len(files) * 0.4):
            structure_issues.append("较多文件直接堆在根目录，后续按项目类型或年份维护会比较困难。")
        if duplicate_name_groups:
            names = "、".join(item["name"] for item in duplicate_name_groups[:3])
            structure_issues.append(f"存在同名文件，例如 {names}，需要确认是否为重复版本或不同项目。")
        if likely_duplicate_groups:
            names = "、".join(item["name"] for item in likely_duplicate_groups[:3])
            structure_issues.append(f"存在疑似重复文件，例如 {names}，文件名和大小相同。")
        if len(folder_counts) > 12:
            structure_issues.append("子目录较分散，建议收敛到稳定的数据分类目录。")
        if not structure_issues:
            structure_issues.append("未发现明显结构问题，建议继续补充字段映射和年度口径说明。")

        recommendations = [
            "建议固定一级目录：员工主档、人才盘点、干部360、绩效、项目评价、权限配置、报告材料。",
            "建议文件名包含年份/项目/数据类型，例如 2025_人才盘点_校准结果.xlsx。",
            "建议为核心表保留工号、姓名、组织路径、年份/项目节点，方便和 MCP 实时档案合并。",
        ]
        if unknown_count:
            recommendations.append("对待确认文件做一次字段映射确认，确认后保存为可复用规则。")
        if duplicate_name_groups or likely_duplicate_groups:
            recommendations.append("清理重复文件或增加版本号，避免 AI 和报告生成引用过期数据。")

        overview = (
            f"本次扫描 {folder}，识别到 {len(files)} 个可用数据文件。"
            f"其中 {known_count} 个可初步归类，{unknown_count} 个待确认。"
            f"主要内容包括：{'、'.join(readable_types) if readable_types else '暂无明确分类'}。"
        )
        return {
            "overview": overview,
            "typeCounts": type_counts,
            "typeBreakdown": [{"type": kind, "label": type_labels.get(kind, kind), "count": count} for kind, count in sorted(type_counts.items())],
            "extensionCounts": extension_counts,
            "topFolders": top_folders,
            "analysisOpportunities": opportunities,
            "structureIssues": structure_issues,
            "recommendations": recommendations,
            "duplicateNames": duplicate_name_groups,
            "likelyDuplicates": likely_duplicate_groups,
        }

    def scan_local_data_sources(self, local_folder=None):
        folder_text = str(local_folder or self.data_source_config().get("localFolder") or "").strip()
        if not folder_text:
            return {"error": "请先配置本地数据目录。", "files": []}
        folder = Path(folder_text).expanduser()
        if not folder.exists() or not folder.is_dir():
            return {"error": f"本地数据目录不存在：{folder}", "files": []}
        supported = {".xlsx", ".xlsm", ".csv", ".json", ".md", ".txt", ".pdf", ".docx"}
        files = []
        for path in sorted(folder.rglob("*"), key=lambda item: str(item))[:1000]:
            if not path.is_file() or path.suffix.lower() not in supported:
                continue
            headers = self._preview_table_headers(path)
            files.append(
                {
                    "name": path.name,
                    "path": str(path),
                    "relativePath": str(path.relative_to(folder)),
                    "extension": path.suffix.lower(),
                    "size": path.stat().st_size,
                    "updatedAt": datetime.fromtimestamp(path.stat().st_mtime, LOCAL_TZ).isoformat(timespec="seconds"),
                    "headers": headers,
                    "detectedType": self._classify_local_source_file(path, headers),
                }
            )
        summary = self._summarize_local_data_scan(files, folder)
        return {
            "localFolder": str(folder),
            "fileCount": len(files),
            "files": files[:300],
            "summary": summary,
            "truncated": len(files) > 300,
            "scannedAt": datetime.now(LOCAL_TZ).isoformat(timespec="seconds"),
        }

    def update_config(self):
        saved = self._read_json(self.update_config_path, {}) if self.update_config_path.exists() else {}
        if not isinstance(saved, dict):
            saved = {}
        return {
            "sourcePath": str(saved.get("sourcePath") or ""),
            "updatedAt": str(saved.get("updatedAt") or ""),
        }

    def save_update_config(self, config):
        current = self.update_config()
        source_path = str(config.get("sourcePath") or current.get("sourcePath") or "").strip()
        payload = {
            "sourcePath": source_path,
            "updatedAt": datetime.now(LOCAL_TZ).isoformat(timespec="seconds"),
        }
        self.update_config_path.parent.mkdir(parents=True, exist_ok=True)
        self.update_config_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return self.update_config()

    def _is_update_url(self, source):
        parsed = urlparse(str(source or "").strip())
        return parsed.scheme in {"http", "https"}

    def _read_update_url_json(self, url):
        request = Request(
            url,
            headers={
                "Accept": "application/json, text/plain;q=0.9, */*;q=0.1",
                "User-Agent": "Hrobot-Updater/0.1",
            },
        )
        with urlopen(request, timeout=20) as response:
            content_type = response.headers.get("Content-Type", "")
            raw = response.read(1024 * 1024)
        text = raw.decode("utf-8-sig")
        stripped = text.lstrip()
        if stripped.startswith("<"):
            raise ValueError("更新源返回的是网页，不是 release.json。企业微信分享页通常不能作为自动更新直链，请使用可直接访问的 release.json URL。")
        try:
            return json.loads(text)
        except json.JSONDecodeError as error:
            raise ValueError(f"更新源不是有效 JSON：{error}") from error

    def _github_release_api_url(self, source):
        source = str(source or "").strip()
        if not source:
            return GITHUB_RELEASE_API_URL
        parsed = urlparse(source)
        if parsed.netloc.lower() == "api.github.com":
            parts = [part for part in parsed.path.split("/") if part]
            if len(parts) >= 4 and parts[0] == "repos" and parts[3] == "releases":
                return source
        if parsed.netloc.lower() != "github.com":
            return ""
        parts = [part for part in parsed.path.split("/") if part]
        if len(parts) >= 4 and parts[2] == "releases":
            owner, repo = parts[0], parts[1]
            tag = parts[3] if parts[3] != "latest" else "latest"
            if tag == "latest":
                return f"https://api.github.com/repos/{owner}/{repo}/releases/latest"
            return f"https://api.github.com/repos/{owner}/{repo}/releases/tags/{quote(tag)}"
        if len(parts) >= 2:
            return f"https://api.github.com/repos/{parts[0]}/{parts[1]}/releases/latest"
        return ""

    def _asset_platform(self, asset_name):
        name = str(asset_name or "").lower()
        if not name:
            return ""
        if name.endswith(".exe") or "win" in name or "windows" in name:
            return "windows"
        if "mac" in name or "darwin" in name or name.endswith(".dmg"):
            return "mac"
        return ""

    def _asset_architecture(self, asset_name):
        name = str(asset_name or "").lower()
        if "universal" in name:
            return "universal"
        if "arm64" in name or "aarch64" in name or "apple-silicon" in name:
            return "arm64"
        if "x64" in name or "x86_64" in name or "intel" in name:
            return "x64"
        return ""

    def _asset_score_for_platform(self, asset, platform):
        name = str(asset.get("name") or "")
        lower = name.lower()
        score = 0
        if self._asset_platform(name) == platform:
            score += 100
        if platform == "windows" and lower.endswith(".exe"):
            score += 40
        if platform == "mac":
            if lower.endswith(".dmg"):
                score += 50
            elif lower.endswith(".zip"):
                score += 20
            architecture = self._asset_architecture(name)
            if architecture == "universal":
                score += 40
            elif architecture == app_architecture():
                score += 35
            elif architecture:
                score -= 80
        if "source code" in lower:
            score -= 100
        if lower.startswith("hrobot"):
            score += 10
        return score

    def _github_release_to_update_release(self, release):
        if not isinstance(release, dict):
            raise ValueError("GitHub Release returned an invalid payload.")
        version = str(release.get("tag_name") or release.get("name") or "").strip().lstrip("vV")
        assets = release.get("assets") if isinstance(release.get("assets"), list) else []
        platform = app_platform()
        candidates = [
            asset for asset in assets
            if isinstance(asset, dict)
            and asset.get("browser_download_url")
            and self._asset_score_for_platform(asset, platform) > 0
        ]
        selected = max(candidates, key=lambda item: self._asset_score_for_platform(item, platform), default=None)
        package_map = {}
        for asset in assets:
            if not isinstance(asset, dict) or not asset.get("browser_download_url"):
                continue
            asset_platform = self._asset_platform(asset.get("name"))
            if asset_platform:
                architecture = self._asset_architecture(asset.get("name"))
                package_key = f"{asset_platform}-{architecture}" if architecture else asset_platform
                package_map[package_key] = {
                    "name": str(asset.get("name") or ""),
                    "url": str(asset.get("browser_download_url") or ""),
                }
        if not version:
            raise ValueError("GitHub Release is missing a version tag.")
        if not selected:
            available = ", ".join(str(asset.get("name") or "") for asset in assets if isinstance(asset, dict) and asset.get("name"))
            raise ValueError(f"No release asset is available for {platform}. Assets: {available or 'none'}")
        return {
            "app": "Hrobot",
            "version": version,
            "installer": str(selected.get("name") or "HrobotSetup.exe"),
            "installerUrl": str(selected.get("browser_download_url") or ""),
            "notes": str(release.get("body") or ""),
            "publishedAt": str(release.get("published_at") or ""),
            "releasePage": str(release.get("html_url") or GITHUB_RELEASE_PAGE_URL),
            "sourceKind": "github",
            "packages": package_map,
        }

    def _download_update_url(self, url, destination: Path):
        request = Request(url, headers={"User-Agent": "Hrobot-Updater/0.1"})
        with urlopen(request, timeout=120) as response:
            destination.parent.mkdir(parents=True, exist_ok=True)
            with open(destination, "wb") as output:
                shutil.copyfileobj(response, output)
        return destination

    def _update_manifest_path(self, source_path=""):
        configured = str(source_path or self.update_config().get("sourcePath") or "").strip().strip('"')
        github_api = self._github_release_api_url(configured)
        if not configured or github_api:
            return github_api or GITHUB_RELEASE_API_URL
        if not configured:
            raise ValueError("请先填写企业微信网盘同步目录或 release.json 路径。")
        if self._is_update_url(configured):
            return configured
        path = Path(configured).expanduser()
        if path.is_dir():
            path = path / "release.json"
        if path.name.lower() != "release.json":
            raise ValueError("更新源需指向包含 release.json 的目录，或直接指向 release.json 文件。")
        if not path.exists():
            raise FileNotFoundError(f"未找到更新清单：{path}")
        return path

    def _read_update_release(self, source_path=""):
        manifest_ref = self._update_manifest_path(source_path)
        is_url = self._is_update_url(manifest_ref)
        release = self._read_update_url_json(manifest_ref) if is_url else self._read_json(manifest_ref, {})
        if not isinstance(release, dict):
            raise ValueError("release.json 格式不正确。")
        if self._github_release_api_url(manifest_ref):
            release = self._github_release_to_update_release(release)
        version = str(release.get("version") or "").strip()
        installer_name = str(
            release.get("installerUrl")
            or release.get("installer")
            or release.get("file")
            or release.get("package")
            or "HrobotSetup.exe"
        ).strip()
        if not version:
            raise ValueError("release.json 缺少 version。")
        if not installer_name:
            raise ValueError("release.json 缺少 installer 文件名。")
        if is_url:
            installer_ref = installer_name if self._is_update_url(installer_name) else urljoin(manifest_ref, installer_name)
            installer_suffix = Path(urlparse(installer_ref).path).suffix.lower()
            manifest_parent = manifest_ref.rsplit("/", 1)[0] + "/"
        else:
            installer_ref = Path(installer_name)
            if not installer_ref.is_absolute():
                installer_ref = manifest_ref.parent / installer_ref
            installer_suffix = installer_ref.suffix.lower()
            manifest_parent = str(manifest_ref.parent)
        if app_platform() == "windows" and installer_suffix != ".exe":
            raise ValueError("当前 Windows 更新只支持 .exe 安装包。")
        current = app_version_payload()
        return {
            "current": current,
            "latest": {
                "version": version,
                "installer": Path(urlparse(str(installer_ref)).path).name if is_url else installer_ref.name,
                "installerPath": str(installer_ref),
                "installerSourceType": "url" if is_url else "file",
                "notes": str(release.get("notes") or release.get("changelog") or ""),
                "publishedAt": str(release.get("publishedAt") or ""),
                "releasePage": str(release.get("releasePage") or GITHUB_RELEASE_PAGE_URL),
                "platform": app_platform(),
                "canAutoInstall": (
                    (app_platform() == "windows" and installer_suffix == ".exe")
                    or (
                        app_platform() == "mac"
                        and installer_suffix == ".dmg"
                        and getattr(sys, "frozen", False)
                    )
                ),
            },
            "manifestPath": str(manifest_ref),
            "sourcePath": str(release.get("releasePage") or GITHUB_RELEASE_PAGE_URL) if release.get("sourceKind") == "github" else manifest_parent,
            "sourceType": "github" if release.get("sourceKind") == "github" else ("url" if is_url else "file"),
            "packages": release.get("packages") if isinstance(release.get("packages"), dict) else {},
            "updateAvailable": is_newer_version(version, current["version"]),
        }

    def update_status(self, source_path=""):
        config = self.update_config()
        payload = {
            "app": app_version_payload(),
            "config": config,
            "checkedAt": datetime.now(LOCAL_TZ).isoformat(timespec="seconds"),
            "configured": True,
            "releasePage": GITHUB_RELEASE_PAGE_URL,
            "updateAvailable": False,
        }
        source = source_path or GITHUB_RELEASE_PAGE_URL
        try:
            release = self._read_update_release(source)
        except HTTPError as error:
            if self._github_release_api_url(source) and error.code == 404:
                return {
                    **payload,
                    "noRelease": True,
                    "sourcePath": GITHUB_RELEASE_PAGE_URL,
                    "sourceType": "github",
                    "latest": {
                        "releasePage": GITHUB_RELEASE_PAGE_URL,
                        "platform": app_platform(),
                        "notes": "",
                    },
                }
            raise
        return {**payload, **release, "configured": True}

    def install_update(self, source_path=""):
        status = self.update_status(source_path)
        latest = status.get("latest") or {}
        installer_path = Path(latest.get("installerPath") or "")
        if not status.get("updateAvailable"):
            return {**status, "started": False, "message": "当前已经是最新版本。"}
        if not latest.get("canAutoInstall"):
            return {**status, "started": False, "message": "当前系统不支持自动覆盖安装，请打开 GitHub Release 下载对应版本。"}
        source_type = latest.get("installerSourceType") or "file"
        if source_type == "file" and not installer_path.exists():
            raise FileNotFoundError(f"未找到安装包：{installer_path}")
        temp_dir = Path(tempfile.gettempdir()) / "hrobot-update"
        temp_dir.mkdir(parents=True, exist_ok=True)
        local_installer = temp_dir / (latest.get("installer") or "HrobotSetup.exe")
        if source_type == "url":
            self._download_update_url(latest.get("installerPath") or "", local_installer)
        else:
            shutil.copy2(installer_path, local_installer)
        if app_platform() == "mac":
            executable = Path(sys.executable).resolve()
            app_bundle = next((parent for parent in executable.parents if parent.suffix == ".app"), None)
            if not app_bundle:
                raise ValueError("当前程序不是 macOS 应用包，无法执行覆盖更新。")
            updater_path = temp_dir / "install-hrobot-macos-update.sh"
            log_path = temp_dir / "macos-update.log"
            script = f"""#!/bin/bash
set -euo pipefail
exec >> {shlex.quote(str(log_path))} 2>&1
DMG={shlex.quote(str(local_installer))}
APP={shlex.quote(str(app_bundle))}
PID={os.getpid()}
MOUNT_DIR=\"$(mktemp -d /tmp/hrobot-update.XXXXXX)\"
cleanup() {{
  /usr/bin/hdiutil detach \"$MOUNT_DIR\" -quiet >/dev/null 2>&1 || true
  /bin/rmdir \"$MOUNT_DIR\" >/dev/null 2>&1 || true
}}
trap cleanup EXIT
/bin/sleep 2
/usr/bin/hdiutil attach \"$DMG\" -nobrowse -readonly -mountpoint \"$MOUNT_DIR\"
SOURCE_APP=\"$MOUNT_DIR/Hrobot.app\"
test -d \"$SOURCE_APP\"
NEW_APP=\"${{APP}}.new\"
/bin/rm -rf \"$NEW_APP\"
/usr/bin/ditto \"$SOURCE_APP\" \"$NEW_APP\"
/bin/rm -rf \"$APP\"
/bin/mv \"$NEW_APP\" \"$APP\"
cleanup
trap - EXIT
/bin/kill \"$PID\" >/dev/null 2>&1 || true
/bin/sleep 1
/usr/bin/open \"$APP\"
"""
            updater_path.write_text(script, encoding="utf-8")
            updater_path.chmod(0o700)
            subprocess.Popen(
                ["/bin/bash", str(updater_path)],
                cwd=str(temp_dir),
                stdin=subprocess.DEVNULL,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                start_new_session=True,
            )
            return {
                **status,
                "started": True,
                "localInstaller": str(local_installer),
                "message": "macOS 更新包已下载，Hrobot 将完成覆盖安装并重新打开。",
            }
        kwargs = {
            "cwd": str(local_installer.parent),
            "stdin": subprocess.DEVNULL,
            "stdout": subprocess.DEVNULL,
            "stderr": subprocess.DEVNULL,
        }
        if os.name == "nt":
            kwargs["creationflags"] = subprocess.DETACHED_PROCESS | subprocess.CREATE_NEW_PROCESS_GROUP
        subprocess.Popen([str(local_installer)], **kwargs)
        return {
            **status,
            "started": True,
            "localInstaller": str(local_installer),
            "message": "更新安装包已启动。安装完成后会重新打开 Hrobot。",
        }

    def save_ai_config(self, config):
        current = self.ai_config()
        timestamp = datetime.now(timezone.utc).isoformat()
        secrets = self._read_json(self.ai_secrets_path, {}) if self.ai_secrets_path.exists() else {}
        if not isinstance(secrets, dict):
            secrets = {}
        saved_raw = self._read_json(self.ai_config_path, {}) if self.ai_config_path.exists() else {}
        if not isinstance(saved_raw, dict):
            saved_raw = {}
        secrets_changed = False

        legacy_api_key = saved_raw.get("apiKey", "")
        if legacy_api_key and not secrets.get("multimodal"):
            secrets["multimodal"] = {"apiKey": legacy_api_key}
            secrets_changed = True

        def merge_group(name):
            nonlocal secrets_changed
            incoming = config.get(name, {}) if isinstance(config.get(name, {}), dict) else {}
            existing = current.get(name, {})
            api_key = incoming.get("apiKey")
            if api_key:
                self._runtime_ai_keys[name] = api_key
                secrets[name] = {"apiKey": api_key}
                secrets_changed = True
            return {
                "apiKey": "",
                "baseUrl": incoming.get("baseUrl") or existing.get("baseUrl", "https://api.openai.com/v1"),
                "model": incoming.get("model") or existing.get("model", ""),
            }

        if any(key in config for key in ("apiKey", "baseUrl", "model")):
            config = {
                "multimodal": {
                    "apiKey": config.get("apiKey"),
                    "baseUrl": config.get("baseUrl"),
                    "model": config.get("model"),
                }
            }
        payload = {
            "multimodal": merge_group("multimodal"),
            "image": merge_group("image"),
            "updatedAt": timestamp,
        }
        self.ai_config_path.parent.mkdir(parents=True, exist_ok=True)
        self.ai_config_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        if secrets_changed:
            self.ai_secrets_path.parent.mkdir(parents=True, exist_ok=True)
            self.ai_secrets_path.write_text(json.dumps(secrets, ensure_ascii=False, indent=2), encoding="utf-8")
        return self.ai_config_status()

    def _first_query_value(self, query, name):
        values = query.get(name, [])
        if not values:
            return ""
        return str(values[0] or "").strip()

    def intelligence_config(self):
        default = {
            "autoEnabled": True,
            "runAt": "10:00",
            "channel": "all",
            "source": "all",
            "maxPerQuery": 3,
            "wechatFulltextLimit": 1,
            "allowUnverifiedWechat": False,
            "updatedAt": "",
        }
        saved = self._read_json(self.intelligence_config_path, {})
        if not isinstance(saved, dict):
            saved = {}
        config = {**default, **saved}
        config["autoEnabled"] = bool(config.get("autoEnabled"))
        config["allowUnverifiedWechat"] = bool(config.get("allowUnverifiedWechat"))
        config["runAt"] = str(config.get("runAt") or "10:00")[:5]
        if config["channel"] not in {"all", "ai_hr", "game_org"}:
            config["channel"] = "all"
        if config["source"] not in {"all", "bing", "wechat"}:
            config["source"] = "all"
        try:
            config["maxPerQuery"] = max(1, min(10, int(config.get("maxPerQuery") or 3)))
        except (TypeError, ValueError):
            config["maxPerQuery"] = 3
        try:
            config["wechatFulltextLimit"] = max(0, min(5, int(config.get("wechatFulltextLimit", 1))))
        except (TypeError, ValueError):
            config["wechatFulltextLimit"] = 1
        return config

    def save_intelligence_config(self, config):
        current = self.intelligence_config()
        incoming = config if isinstance(config, dict) else {}
        payload = {**current}
        for key in ("autoEnabled", "runAt", "channel", "source", "maxPerQuery", "wechatFulltextLimit", "allowUnverifiedWechat"):
            if key in incoming:
                payload[key] = incoming[key]
        payload["updatedAt"] = datetime.now(LOCAL_TZ).isoformat(timespec="seconds")
        self.intelligence_config_path.parent.mkdir(parents=True, exist_ok=True)
        self.intelligence_config_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return self.intelligence_config()

    def intelligence_update_status(self):
        status = self._read_json(
            self.intelligence_update_status_path,
            {"running": False, "lastStartedAt": "", "lastFinishedAt": "", "lastAutoDate": "", "lastTrigger": "", "ok": None, "message": ""},
        )
        return status if isinstance(status, dict) else {}

    def _save_intelligence_update_status(self, status):
        self.intelligence_update_status_path.parent.mkdir(parents=True, exist_ok=True)
        self.intelligence_update_status_path.write_text(json.dumps(status, ensure_ascii=False, indent=2), encoding="utf-8")

    def _run_intelligence_update_script(self, script_path, args):
        if not getattr(sys, "frozen", False):
            return subprocess.run(
                [sys.executable, str(script_path), *args],
                cwd=str(ROOT),
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=900,
            )

        spec = importlib.util.spec_from_file_location("hrobot_update_intelligence", script_path)
        if spec is None or spec.loader is None:
            raise RuntimeError(f"Cannot load intelligence script: {script_path}")
        module = importlib.util.module_from_spec(spec)
        old_argv = sys.argv[:]
        stdout = io.StringIO()
        stderr = io.StringIO()
        return_code = 0
        try:
            sys.argv = [str(script_path), *args]
            with contextlib.redirect_stdout(stdout), contextlib.redirect_stderr(stderr):
                spec.loader.exec_module(module)
                return_code = int(module.main() or 0)
        except SystemExit as exc:
            return_code = int(exc.code or 0) if isinstance(exc.code, int) else 1
        finally:
            sys.argv = old_argv
        return subprocess.CompletedProcess(
            [str(script_path), *args],
            return_code,
            stdout.getvalue(),
            stderr.getvalue(),
        )

    def update_intelligence_now(self, trigger="manual", target_date=None):
        if not INTELLIGENCE_UPDATE_LOCK.acquire(blocking=False):
            return {"ok": False, "running": True, "message": "情报更新正在运行，请稍后再试。", "status": self.intelligence_update_status()}
        started_at = datetime.now(LOCAL_TZ).isoformat(timespec="seconds")
        status = {
            **self.intelligence_update_status(),
            "running": True,
            "lastStartedAt": started_at,
            "lastTrigger": trigger,
            "ok": None,
            "message": "运行中",
        }
        self._save_intelligence_update_status(status)
        try:
            config = self.intelligence_config()
            script_path = ROOT / "scripts" / "update_intelligence.py"
            date_arg = target_date or (datetime.now(LOCAL_TZ).date() - timedelta(days=1)).isoformat()
            args = [
                "--date",
                date_arg,
                "--channel",
                config["channel"],
                "--source",
                config["source"],
                "--max-per-query",
                str(config["maxPerQuery"]),
                "--wechat-fulltext-limit",
                str(config["wechatFulltextLimit"]),
            ]
            if config.get("allowUnverifiedWechat"):
                args.append("--allow-unverified-wechat")
            completed = self._run_intelligence_update_script(script_path, args)
            finished_at = datetime.now(LOCAL_TZ).isoformat(timespec="seconds")
            ok = completed.returncode == 0
            message = (completed.stdout or completed.stderr or "").strip()
            if not ok:
                _agent_debug_log(
                    "server.py:update_intelligence_now",
                    "intelligence script failed",
                    {
                        "returncode": completed.returncode,
                        "stdoutPreview": (completed.stdout or "")[:500],
                        "stderrPreview": (completed.stderr or "")[:500],
                    },
                    "H5",
                )
            status = {
                **status,
                "running": False,
                "lastFinishedAt": finished_at,
                "lastTrigger": trigger,
                "ok": ok,
                "message": message or ("更新完成" if ok else "更新失败"),
                "returnCode": completed.returncode,
            }
            if trigger == "auto":
                status["lastAutoDate"] = datetime.now(LOCAL_TZ).date().isoformat()
            self._save_intelligence_update_status(status)
            return {"ok": ok, "message": status["message"], "status": status}
        except subprocess.TimeoutExpired:
            status = {
                **status,
                "running": False,
                "lastFinishedAt": datetime.now(LOCAL_TZ).isoformat(timespec="seconds"),
                "lastTrigger": trigger,
                "ok": False,
                "message": "情报更新超时，请稍后重试或检查网络。",
            }
            if trigger == "auto":
                status["lastAutoDate"] = datetime.now(LOCAL_TZ).date().isoformat()
            self._save_intelligence_update_status(status)
            return {"ok": False, "message": status["message"], "status": status}
        except Exception as error:
            status = {
                **status,
                "running": False,
                "lastFinishedAt": datetime.now(LOCAL_TZ).isoformat(timespec="seconds"),
                "lastTrigger": trigger,
                "ok": False,
                "message": f"情报更新异常：{error}",
            }
            if trigger == "auto":
                status["lastAutoDate"] = datetime.now(LOCAL_TZ).date().isoformat()
            self._save_intelligence_update_status(status)
            return {"ok": False, "message": status["message"], "status": status}
        finally:
            INTELLIGENCE_UPDATE_LOCK.release()

    def _read_intelligence_file(self, path: Path):
        payload = self._read_json(path, {"updated_at": "", "items": []})
        if not isinstance(payload, dict):
            payload = {"updated_at": "", "items": []}
        items = payload.get("items", [])
        if not isinstance(items, list):
            items = []
        return {
            "updated_at": payload.get("updated_at", ""),
            "items": items,
        }

    def _filter_intelligence_items(self, items, query):
        channel = self._first_query_value(query, "channel")
        category = self._first_query_value(query, "category")
        date = self._first_query_value(query, "date")
        date_from = self._first_query_value(query, "from")
        date_to = self._first_query_value(query, "to")
        search = self._first_query_value(query, "search").lower()

        def matches(item):
            if channel and item.get("channel") != channel:
                return False
            if category and item.get("category") != category:
                return False
            published_at = str(item.get("published_at", ""))
            if date and published_at != date:
                return False
            if date_from and published_at < date_from:
                return False
            if date_to and published_at > date_to:
                return False
            if search:
                keywords = item.get("keywords", [])
                if not isinstance(keywords, list):
                    keywords = []
                haystack = " ".join(
                    [
                        str(item.get("title", "")),
                        str(item.get("summary", "")),
                        str(item.get("hrbp_takeaway", "")),
                        str(item.get("source", "")),
                        " ".join(str(keyword) for keyword in keywords),
                    ]
                ).lower()
                if search not in haystack:
                    return False
            return True

        return [item for item in items if isinstance(item, dict) and matches(item)]

    def intelligence(self, query=None):
        query = query or {}
        use_history = self._first_query_value(query, "history").lower() in {"1", "true", "yes", "on"}
        if not use_history:
            use_history = any(self._first_query_value(query, key) for key in ("date", "from", "to", "channel", "category", "search"))
            if use_history and not any(self._first_query_value(query, key) for key in ("date", "from", "to")):
                use_history = False
        payload = self._read_intelligence_file(self.intelligence_history_path if use_history else self.intelligence_path)
        items = self._filter_intelligence_items(payload["items"], query)
        return {
            "ok": True,
            "scope": "history" if use_history else "current",
            "updated_at": payload["updated_at"],
            "items": items,
            "total": len(payload["items"]),
            "count": len(items),
        }

    def design_posters(self):
        payload = self._read_json(self.design_history_path, {"updatedAt": "", "items": []})
        if not isinstance(payload, dict):
            payload = {"updatedAt": "", "items": []}
        items = payload.get("items", [])
        if not isinstance(items, list):
            items = []
        return {
            "updatedAt": payload.get("updatedAt", ""),
            "items": items,
        }

    @staticmethod
    def _normalize_home_memo_records(records):
        normalized = []
        seen = set()
        for item in records or []:
            if not isinstance(item, dict):
                continue
            date = str(item.get("date", "")).strip()
            text = str(item.get("text", "")).strip()
            if not date or not text:
                continue
            if date in seen:
                continue
            seen.add(date)
            normalized.append({"date": date, "text": text})
        normalized.sort(key=lambda item: item["date"])
        return normalized

    def home_memos(self):
        payload = self._read_json(self.home_memo_path, DEFAULT_HOME_MEMO)
        if not isinstance(payload, dict):
            payload = dict(DEFAULT_HOME_MEMO)
        return {
            "updatedAt": payload.get("updatedAt", ""),
            "records": self._normalize_home_memo_records(payload.get("records", [])),
        }

    def save_home_memos(self, records):
        payload = {
            "updatedAt": datetime.now(timezone.utc).isoformat(),
            "records": self._normalize_home_memo_records(records),
        }
        self.home_memo_path.parent.mkdir(parents=True, exist_ok=True)
        self.home_memo_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return payload

    def save_home_memo(self, date, text):
        date = str(date or "").strip()
        text = str(text or "").strip()
        if not date:
            raise ValueError("缺少备忘日期。")
        if not text:
            raise ValueError("缺少备忘内容。")
        records = [item for item in self.home_memos()["records"] if item.get("date") != date]
        records.append({"date": date, "text": text})
        return self.save_home_memos(records)

    def delete_home_memo(self, date):
        date = str(date or "").strip()
        if not date:
            raise ValueError("缺少备忘日期。")
        records = [item for item in self.home_memos()["records"] if item.get("date") != date]
        return self.save_home_memos(records)

    def _home_avatar_entry(self, src, name=None, builtin=False):
        src = str(src or "").strip().replace("\\", "/")
        if not src:
            return None
        label = name or Path(src).stem.replace("-", " ").replace("_", " ").strip() or "Avatar"
        return {
            "id": src,
            "name": label,
            "src": src,
            "builtin": bool(builtin),
        }

    def _trim_home_avatar_asset(self, path):
        if path.suffix.lower() not in HOME_AVATAR_TRIMMABLE_EXTENSIONS:
            return None
        try:
            from PIL import Image
        except ImportError:
            return None
        try:
            source_mtime = int(path.stat().st_mtime)
        except OSError:
            return None
        trimmed_name = f"{path.stem}-trimmed-{source_mtime}.png"
        trimmed_path = self.home_avatar_trim_dir / trimmed_name
        if not trimmed_path.exists():
            try:
                image = Image.open(path).convert("RGBA")
                alpha = image.getchannel("A")
                bbox = alpha.getbbox()
                if not bbox:
                    return None
                # Fully opaque generated avatars often carry white canvas padding. Convert
                # near-white pixels to transparent before finding the visual content box.
                if alpha.getextrema() == (255, 255):
                    pixels = image.load()
                    for y in range(image.height):
                        for x in range(image.width):
                            red, green, blue, opacity = pixels[x, y]
                            if opacity and red > 244 and green > 244 and blue > 244:
                                pixels[x, y] = (red, green, blue, 0)
                    bbox = image.getchannel("A").getbbox()
                    if not bbox:
                        return None
                trimmed = image.crop(bbox)
                self.home_avatar_trim_dir.mkdir(parents=True, exist_ok=True)
                trimmed.save(trimmed_path)
            except Exception:
                return None
        try:
            return trimmed_path.resolve().relative_to(ROOT.resolve()).as_posix()
        except ValueError:
            return None

    def _home_avatar_options(self):
        options = [self._home_avatar_entry("assets/brand/hrobot-buddy-avatar.svg", "HRobot Buddy", True)]
        if self.home_avatar_dir.exists():
            for path in sorted(self.home_avatar_dir.iterdir(), key=lambda item: item.name.lower()):
                if path == self.home_avatar_trim_dir:
                    continue
                if not path.is_file() or path.suffix.lower() not in HOME_AVATAR_EXTENSIONS:
                    continue
                try:
                    src = path.resolve().relative_to(ROOT.resolve()).as_posix()
                except ValueError:
                    continue
                entry = self._home_avatar_entry(src)
                if entry:
                    options.append(entry)
        seen = set()
        unique_options = []
        for item in options:
            if item["src"] in seen:
                continue
            seen.add(item["src"])
            source_path = ROOT / item["src"]
            display_src = self._trim_home_avatar_asset(source_path) if not item.get("builtin") else None
            item["displaySrc"] = display_src or item["src"]
            item["trimmed"] = bool(display_src)
            item["url"] = quote(item["displaySrc"], safe="/:@?&=%+~#")
            unique_options.append(item)
        return unique_options

    def home_avatar_config(self):
        self.home_avatar_dir.mkdir(parents=True, exist_ok=True)
        saved = self._read_json(self.home_avatar_config_path, DEFAULT_HOME_AVATAR_CONFIG)
        if not isinstance(saved, dict):
            saved = dict(DEFAULT_HOME_AVATAR_CONFIG)
        options = self._home_avatar_options()
        selected_src = str(saved.get("selectedSrc") or DEFAULT_HOME_AVATAR_CONFIG["selectedSrc"]).strip().replace("\\", "/")
        available = {item["src"] for item in options}
        if selected_src not in available:
            selected_src = DEFAULT_HOME_AVATAR_CONFIG["selectedSrc"]
        selected = next((item for item in options if item["src"] == selected_src), options[0])
        return {
            "updatedAt": saved.get("updatedAt", ""),
            "selectedSrc": selected["src"],
            "selectedUrl": selected["url"],
            "selectedName": selected["name"],
            "avatarFolder": self.home_avatar_dir.relative_to(ROOT).as_posix(),
            "options": options,
        }

    def save_home_avatar_config(self, selected_src):
        selected_src = str(selected_src or "").strip().replace("\\", "/")
        options = self._home_avatar_options()
        available = {item["src"] for item in options}
        if selected_src not in available:
            raise ValueError("Avatar is not in the available list.")
        payload = {
            "updatedAt": datetime.now(timezone.utc).isoformat(),
            "selectedSrc": selected_src,
        }
        self.home_avatar_config_path.parent.mkdir(parents=True, exist_ok=True)
        self.home_avatar_config_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return self.home_avatar_config()

    def _ensure_design_reference_readme(self):
        if self.design_reference_readme_path.exists():
            return
        content = (
            "# Design References\n\n"
            "Place logo images, mascot images, poster templates, and other reference assets in this folder.\n"
            "The current prompt builder will expose file names to the model as text hints.\n"
            "Future upgrades can attach these files as true multimodal references.\n"
        )
        self.design_reference_dir.mkdir(parents=True, exist_ok=True)
        self.design_reference_readme_path.write_text(content, encoding="utf-8")

    def design_reference_assets(self):
        self.design_reference_dir.mkdir(parents=True, exist_ok=True)
        self._ensure_design_reference_readme()
        items = []
        for path in sorted(self.design_reference_dir.iterdir(), key=lambda item: item.name.lower()):
            if not path.is_file() or path.name.lower() == "readme.md":
                continue
            items.append(
                {
                    "name": path.name,
                    "path": f"/data/design_center/references/{path.name}",
                    "size": path.stat().st_size,
                    "updatedAt": datetime.fromtimestamp(path.stat().st_mtime, timezone.utc).isoformat(),
                }
            )
        return items

    def design_prompt_config(self):
        self.design_dir.mkdir(parents=True, exist_ok=True)
        self.design_reference_dir.mkdir(parents=True, exist_ok=True)
        self._ensure_design_reference_readme()
        saved = self._read_json(self.design_prompt_config_path, {})
        if not isinstance(saved, dict):
            saved = {}
        config = {**DEFAULT_DESIGN_PROMPT_CONFIG, **{key: str(value) for key, value in saved.items() if isinstance(value, (str, int, float))}}
        return {
            **config,
            "configPath": str(self.design_prompt_config_path),
            "referenceFolder": str(self.design_reference_dir),
            "referenceFiles": self.design_reference_assets(),
        }

    def save_design_prompt_config(self, config):
        payload = {
            key: str((config or {}).get(key, DEFAULT_DESIGN_PROMPT_CONFIG[key]) or "").strip()
            for key in DEFAULT_DESIGN_PROMPT_CONFIG
        }
        self.design_prompt_config_path.parent.mkdir(parents=True, exist_ok=True)
        self.design_reference_dir.mkdir(parents=True, exist_ok=True)
        self._ensure_design_reference_readme()
        self.design_prompt_config_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return self.design_prompt_config()

    def save_design_poster(self, poster):
        timestamp = datetime.now(timezone.utc).isoformat()
        poster_id = datetime.now().strftime("%Y%m%d%H%M%S%f")
        image_name = f"poster-{poster_id}.png"
        self.design_poster_dir.mkdir(parents=True, exist_ok=True)
        image_path = self.design_poster_dir / image_name

        if poster.get("b64_json"):
            image_path.write_bytes(base64.b64decode(poster["b64_json"]))
        elif poster.get("imageBytes"):
            image_path.write_bytes(poster["imageBytes"])
        else:
            raise ValueError("No image data returned by image model")

        item = {
            "id": poster_id,
            "posterType": poster.get("posterType", ""),
            "style": poster.get("style", ""),
            "size": poster.get("size", ""),
            "prompt": poster.get("prompt", ""),
            "model": poster.get("model", ""),
            "imagePath": f"/data/design_center/posters/{image_name}",
            "createdAt": timestamp,
        }
        history = self.design_posters()
        payload = {
            "updatedAt": timestamp,
            "items": [item] + history["items"],
        }
        self.design_history_path.parent.mkdir(parents=True, exist_ok=True)
        self.design_history_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return item



class Handler(SimpleHTTPRequestHandler):
    store = DataStore(DATA_DIR)
    allow_remote_clients = False

    def handle_one_request(self):
        try:
            super().handle_one_request()
        except Exception as error:
            _agent_debug_log(
                "server.py:handle_one_request",
                "unhandled request exception",
                {"errorType": type(error).__name__, "error": str(error), "path": self.path},
                "H5",
            )
            raise

    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(ROOT), **kwargs)

    def _should_disable_cache(self):
        path = urlparse(getattr(self, "path", "")).path
        return path in {"", "/", "/index.html"} or path.startswith("/static/")

    def _static_no_cache_path(self):
        path = urlparse(getattr(self, "path", "")).path
        if path in {"", "/"}:
            file_path = (ROOT / "index.html").resolve()
        elif self._should_disable_cache():
            file_path = Path(self.translate_path(path)).resolve()
        else:
            return None
        root = ROOT.resolve()
        if not file_path.is_relative_to(root) or not file_path.is_file():
            return None
        return file_path

    def _send_static_no_cache(self, include_body=True):
        file_path = self._static_no_cache_path()
        if not file_path:
            return False
        stat = file_path.stat()
        self.send_response(200)
        self.send_header("Content-Type", self.guess_type(str(file_path)))
        self.send_header("Content-Length", str(stat.st_size))
        self.send_header("Last-Modified", self.date_time_string(stat.st_mtime))
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        self.end_headers()
        if include_body:
            with file_path.open("rb") as handle:
                self.copyfile(handle, self.wfile)
        return True

    def do_HEAD(self):
        if not self._allow_request():
            return
        if self._send_static_no_cache(include_body=False):
            return
        super().do_HEAD()

    def _client_is_local(self):
        host = (self.client_address[0] if self.client_address else "") or ""
        return host.startswith("127.") or host == "::1" or host == "localhost"

    def _allow_request(self):
        if self.allow_remote_clients or self._client_is_local():
            return True
        body = "Remote access is disabled. Restart with --allow-remote-clients only on a trusted network."
        encoded = body.encode("utf-8")
        self.send_response(403)
        self.send_header("Content-Type", "text/plain; charset=utf-8")
        self.send_header("Content-Length", str(len(encoded)))
        self.end_headers()
        self.wfile.write(encoded)
        return False

    def _send_json(self, payload, status=200):
        try:
            body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        except (TypeError, UnicodeEncodeError) as error:
            _agent_debug_log(
                "server.py:_send_json",
                "json.dumps failed",
                {"errorType": type(error).__name__, "error": str(error), "path": self.path, "status": status},
                "H3",
            )
            raise
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_file(self, path: Path, content_type: str):
        body = path.read_bytes()
        filename = quote(path.name)
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{filename}")
        self.end_headers()
        self.wfile.write(body)

    def _read_request_json(self):
        length = int(self.headers.get("Content-Length", "0"))
        raw = self.rfile.read(length)
        try:
            text = raw.decode("utf-8") or "{}"
        except UnicodeDecodeError as error:
            _agent_debug_log(
                "server.py:_read_request_json",
                "request body UTF-8 decode failed",
                {"path": self.path, "length": length, "error": str(error)},
                "H2",
            )
            return None
        try:
            return json.loads(text)
        except json.JSONDecodeError as error:
            _agent_debug_log(
                "server.py:_read_request_json",
                "request JSON parse failed",
                {"path": self.path, "preview": text[:200], "error": str(error)},
                "H2",
            )
            return None

    def _read_multipart_file(self):
        files = self._read_multipart_files()
        return files[0] if files else (None, None)

    def _read_multipart_files(self):
        form = self._read_multipart_form()
        return [(item.get("filename") or "upload", item.get("content") or b"") for item in form.get("files", [])]

    def _read_multipart_form(self):
        content_type = self.headers.get("Content-Type", "")
        match = re.search(r"boundary=(.+)", content_type)
        if not match:
            return {"fields": {}, "files": []}
        boundary = match.group(1).strip('"').encode("utf-8")
        length = int(self.headers.get("Content-Length", "0"))
        body = self.rfile.read(length)
        fields = {}
        files = []
        for part in body.split(b"--" + boundary):
            if b"\r\n\r\n" not in part:
                continue
            raw_headers, content = part.split(b"\r\n\r\n", 1)
            headers = raw_headers.decode("utf-8", errors="replace")
            if content.endswith(b"\r\n"):
                content = content[:-2]
            name_match = re.search(r'name="([^"]*)"', headers)
            field_name = name_match.group(1) if name_match else ""
            filename_match = re.search(r'filename="([^"]*)"', headers)
            if filename_match:
                filename = filename_match.group(1) or "upload"
                if filename and not filename.isascii():
                    _agent_debug_log(
                        "server.py:_read_multipart_files",
                        "non-ascii multipart filename",
                        {"filename": filename, "path": self.path, "headersPreview": headers[:300]},
                        "H2",
                    )
                files.append({"field": field_name, "filename": filename, "content": content})
            elif field_name:
                fields[field_name] = content.decode("utf-8-sig", errors="replace")
        return {"fields": fields, "files": files}

    @staticmethod
    def _ai_request_payload(model, messages):
        return {
            "model": model,
            "messages": messages,
        }

    @staticmethod
    def _ai_request_payload_with_tools(model, messages, tools):
        payload = Handler._ai_request_payload(model, messages)
        if tools:
            payload["tools"] = tools
            payload["tool_choice"] = "auto"
        return payload

    @staticmethod
    def _safe_tool_name(index, name):
        safe = re.sub(r"[^a-zA-Z0-9_-]+", "_", str(name or f"tool_{index}")).strip("_")
        safe = safe or f"tool_{index}"
        return f"mcp_{index}_{safe}"[:64]

    @staticmethod
    def _parse_mcp_response_body(body, content_type=""):
        text = body.decode("utf-8", errors="replace").strip()
        if not text:
            return {}
        if "text/event-stream" in (content_type or "") or text.startswith(("event:", "data:")):
            data_lines = []
            for line in text.splitlines():
                if line.startswith("data:"):
                    value = line[5:].strip()
                    if value and value != "[DONE]":
                        data_lines.append(value)
            for value in data_lines:
                try:
                    return json.loads(value)
                except json.JSONDecodeError:
                    continue
            return {"error": {"message": text[:1000]}}
        return json.loads(text)

    @staticmethod
    def _mcp_result_text(result):
        content = result.get("content") if isinstance(result, dict) else None
        if isinstance(content, list):
            parts = []
            for item in content:
                if not isinstance(item, dict):
                    continue
                if item.get("type") == "text":
                    parts.append(str(item.get("text", "")))
                elif item.get("type"):
                    parts.append(json.dumps(item, ensure_ascii=False))
            if parts:
                return "\n".join(part for part in parts if part)
        return json.dumps(result, ensure_ascii=False)

    def _mcp_request(self, method, params=None, request_id=None):
        config = self.store.mcp_config()
        url = (config.get("url") or "").strip()
        headers = config.get("headers") or {}
        if not url or not headers.get("x-api-key"):
            raise RuntimeError("MCP 未配置，请设置 HROBOT_MCP_URL 和 HROBOT_MCP_API_KEY，或填写 data/mcp_config.json。")
        payload = {
            "jsonrpc": "2.0",
            "method": method,
        }
        if params is not None:
            payload["params"] = params
        if request_id is not None:
            payload["id"] = request_id
        request_headers = {
            **headers,
            "Accept": "application/json, text/event-stream",
            "Content-Type": "application/json",
        }
        session_id = getattr(self, "_mcp_session_id", "")
        if session_id:
            request_headers["mcp-session-id"] = session_id
        request = Request(
            url,
            data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            headers=request_headers,
            method="POST",
        )
        with urlopen(request, timeout=120) as response:
            response_session_id = response.headers.get("mcp-session-id")
            if response_session_id:
                self._mcp_session_id = response_session_id
            data = self._parse_mcp_response_body(response.read(), response.headers.get("Content-Type", ""))
        if isinstance(data, dict) and data.get("error"):
            error = data["error"]
            message = error.get("message") if isinstance(error, dict) else str(error)
            raise RuntimeError(message or "MCP 请求失败")
        return data.get("result", data) if isinstance(data, dict) else data

    def _mcp_initialize(self):
        if getattr(self, "_mcp_initialized", False):
            return
        self._mcp_request(
            "initialize",
            {
                "protocolVersion": "2025-03-26",
                "capabilities": {},
                "clientInfo": {"name": "HRobot Web AI Chat", "version": "1.0"},
            },
            request_id=1,
        )
        try:
            self._mcp_request("notifications/initialized", request_id=None)
        except Exception:
            pass
        self._mcp_initialized = True

    def _mcp_tools_for_ai(self):
        try:
            self._mcp_initialize()
            result = self._mcp_request("tools/list", {}, request_id=2)
        except Exception as error:
            _agent_debug_log(
                "server.py:_mcp_tools_for_ai",
                "MCP tools/list failed",
                {"errorType": type(error).__name__, "error": str(error)},
                "MCP1",
            )
            return [], {}
        tools = result.get("tools", []) if isinstance(result, dict) else []
        ai_tools = []
        tool_map = {}
        for index, tool in enumerate(tools[:32], start=1):
            if not isinstance(tool, dict) or not tool.get("name"):
                continue
            ai_name = self._safe_tool_name(index, tool.get("name"))
            schema = tool.get("inputSchema") if isinstance(tool.get("inputSchema"), dict) else {"type": "object", "properties": {}}
            ai_tools.append(
                {
                    "type": "function",
                    "function": {
                        "name": ai_name,
                        "description": tool.get("description") or f"HRobot MCP tool: {tool.get('name')}",
                        "parameters": schema,
                    },
                }
            )
            tool_map[ai_name] = tool.get("name")
        return ai_tools, tool_map

    def _mcp_call_tool(self, name, arguments):
        self._mcp_initialize()
        result = self._mcp_request(
            "tools/call",
            {
                "name": name,
                "arguments": arguments if isinstance(arguments, dict) else {},
            },
            request_id=int(time.time() * 1000) % 1000000000,
        )
        return self._mcp_result_text(result if isinstance(result, dict) else {"result": result})

    @staticmethod
    def _extract_profile_names_from_text(text):
        text = str(text or "")
        blocked = {
            "最近", "一次", "晋升", "时间", "任职", "记录", "绩效", "结果", "季度", "个人",
            "档案", "信息", "查询", "查看", "分析", "当前", "系统", "数据", "人才", "盘点",
        }
        candidates = []
        patterns = [
            r"([\u4e00-\u9fff]{2,4})的(?:个人档案|档案|任职|履历|绩效|晋升|人才|盘点|最近)",
            r"([\u4e00-\u9fff]{2,4})(?:近|最近|过去|本年|今年|去年)\S{0,12}(?:绩效|任职|履历|晋升|档案|盘点)",
            r"(?:查一下|查询|查看|看看|看一下|分析|关于)\s*([\u4e00-\u9fff]{2,4})",
            r"员工档案[：:]\s*([\u4e00-\u9fff]{2,4})",
            r"姓名[：:]\s*([\u4e00-\u9fff]{2,4})",
        ]
        for pattern in patterns:
            for match in re.finditer(pattern, text):
                name = match.group(1).strip().rstrip("的")
                if name and name not in blocked and not any(word in name for word in blocked):
                    candidates.append(name)
        return list(dict.fromkeys(candidates))[:3]

    def _mcp_profile_context_for_question(self, question, history):
        related_terms = ("档案", "任职", "履历", "绩效", "季度", "晋升", "职级", "部门", "人才盘点", "九宫格")
        question_text = str(question or "")
        history_text = "\n".join(str(item.get("content", "")) for item in (history or [])[-4:] if isinstance(item, dict))
        if not any(term in question_text for term in related_terms):
            return ""
        names = self._extract_profile_names_from_text(question_text) or self._extract_profile_names_from_text(history_text)
        if not names:
            return ""
        chunks = []
        for name in names:
            try:
                result = self._mcp_call_tool("talent-profile-search", {"names": [name]})
                chunks.append(f"## MCP 人才档案查询：{name}\n{result}")
            except Exception as error:
                chunks.append(f"## MCP 人才档案查询：{name}\nMCP 查询失败：{error}")
                _agent_debug_log(
                    "server.py:_mcp_profile_context_for_question",
                    "MCP profile prefetch failed",
                    {"name": name, "errorType": type(error).__name__, "error": str(error)},
                    "MCP3",
                )
        context = "\n\n".join(chunks)
        if len(context) > 120000:
            context = context[:120000] + "\n...MCP 查询结果过长，后续内容已截断。"
        return context

    @staticmethod
    def _extract_talent_search_arguments(text):
        text = str(text or "").strip()
        arguments = {}
        work_ids = []
        for pattern in (r"(?:工号|员工工号|workId|employeeId)[：:\s]*([0-9]{1,12})", r"\b([0-9]{3,12})\b"):
            for match in re.finditer(pattern, text, re.I):
                work_ids.append(match.group(1))
        if work_ids:
            arguments["workIds"] = list(dict.fromkeys(work_ids))[:5]

        names = Handler._extract_profile_names_from_text(text)
        if names:
            arguments["names"] = names

        dept_matches = []
        dept_pattern = r"([\u4e00-\u9fffA-Za-z0-9_/·（）()]{2,32}(?:中心|工作室|项目组|项目部|事业部|技术部|运营部|美术部|行政部|部门|部))"
        for match in re.finditer(dept_pattern, text):
            value = match.group(1).strip(" ，,。；;：:")
            if value and value not in {"部门", "二级部门"} and not any(token in value for token in ("哪个", "哪些", "所有", "多少")):
                dept_matches.append(value)
        if dept_matches and not names:
            arguments["deptPaths"] = list(dict.fromkeys(dept_matches))[:5]

        level_match = re.search(r"\b([PMT][_-]?\d+(?:-\d+)?)\b", text, re.I)
        if level_match:
            arguments["jobLevel"] = level_match.group(1).replace("_", "").upper()
            arguments["jobLevelOperator"] = "EQ"

        if "在职" in text and "离职" not in text:
            arguments["onJob"] = True
        elif "离职" in text:
            arguments["onJob"] = False
        return arguments

    def _forced_mcp_talent_search(self, query, history=None):
        history_text = "\n".join(str(item.get("content", "")) for item in (history or [])[-4:] if isinstance(item, dict))
        arguments = self._extract_talent_search_arguments(query)
        if not arguments:
            arguments = self._extract_talent_search_arguments(history_text)
        if not arguments:
            return {
                "error": "请在输入框里写清楚要检索的姓名、工号或部门，例如：江晓伟近四个季度绩效，或 工号 2219。",
            }
        try:
            result = self._mcp_call_tool("talent-profile-search", arguments)
        except Exception as error:
            _agent_debug_log(
                "server.py:_forced_mcp_talent_search",
                "MCP forced talent search failed",
                {"arguments": arguments, "errorType": type(error).__name__, "error": str(error)},
                "MCP4",
            )
            return {"error": f"MCP 人才检索失败：{error}", "arguments": arguments}
        if len(result) > 120000:
            result = result[:120000] + "\n...MCP 返回内容过长，后续内容已截断。"
        return {
            "message": result or "MCP 未返回内容。",
            "arguments": arguments,
        }

    @staticmethod
    def _clean_text(value):
        if value is None:
            return ""
        text = str(value).strip()
        return "" if text.lower() in {"", "n/a", "none", "null", "-"} else text

    @staticmethod
    def _first_clean(*values):
        for value in values:
            text = Handler._clean_text(value)
            if text:
                return text
        return ""

    @staticmethod
    def _extract_person_profile_query(text):
        text = str(text or "").strip()
        work_id_match = re.search(r"(?:工号|员工工号|employeeId|workId)[:：\s]*([0-9]{1,12})|\b([0-9]{3,12})\b", text, re.I)
        names = []
        blocked = {"人员", "员工", "档案", "信息", "资料", "查询", "查看", "分析", "人才", "绩效", "晋升", "履历"}
        patterns = [
            r"([\u4e00-\u9fff]{2,4})的(?:人员信息|个人信息|人员档案|个人档案|档案|资料|履历)",
            r"(?:查询|查看|看看|生成|打开)\s*([\u4e00-\u9fff]{2,4})(?:的)?(?:人员信息|个人信息|人员档案|个人档案|档案|资料|履历)?",
            r"(?:姓名|人员|员工)[:：\s]*([\u4e00-\u9fff]{2,4})",
        ]
        for pattern in patterns:
            for match in re.finditer(pattern, text):
                name = Handler._clean_text(match.group(1)).rstrip("的")
                if name and name not in blocked and not any(token in name for token in blocked):
                    names.append(name)
        if not names:
            compact = re.sub(r"(人员信息|个人信息|人员档案|个人档案|档案|资料|履历|查询|查看|生成|打开|的)", " ", text)
            for token in re.findall(r"[\u4e00-\u9fff]{2,4}", compact):
                if token not in blocked:
                    names.append(token)
        result = {}
        if work_id_match:
            result["workIds"] = [work_id_match.group(1) or work_id_match.group(2)]
        if names:
            result["names"] = list(dict.fromkeys(names))[:3]
        return result

    @staticmethod
    def _profile_matches_query(person, arguments):
        profile = person.get("profile") if isinstance(person.get("profile"), dict) else {}
        employee_id = Handler._clean_text(person.get("employeeId") or profile.get("employeeId"))
        name = Handler._clean_text(person.get("name") or profile.get("name"))
        work_ids = {str(item) for item in arguments.get("workIds", [])}
        names = {str(item) for item in arguments.get("names", [])}
        if work_ids and employee_id in work_ids:
            return True
        if names and name in names:
            return True
        return False

    @staticmethod
    def _normalize_profile_tags(tags):
        if isinstance(tags, dict):
            values = []
            for key, value in tags.items():
                if isinstance(value, list):
                    values.extend(Handler._clean_text(item) for item in value)
                elif value:
                    label = Handler._clean_text(key)
                    text = Handler._clean_text(value)
                    values.append(f"{label}：{text}" if label and text else text or label)
            return [item for item in values if item]
        if isinstance(tags, list):
            return [Handler._clean_text(item) for item in tags if Handler._clean_text(item)]
        text = Handler._clean_text(tags)
        return [text] if text else []

    @staticmethod
    def _normalize_recent_performance(profile):
        recent = profile.get("recentPerformance") or []
        rows = []
        if isinstance(recent, list):
            for item in recent:
                if isinstance(item, dict):
                    period = Handler._first_clean(item.get("period"), item.get("name"), item.get("quarter"))
                    value = Handler._first_clean(item.get("value"), item.get("managerRating"), item.get("rating"))
                else:
                    text = Handler._clean_text(item)
                    if ":" in text:
                        period, value = [part.strip() for part in text.split(":", 1)]
                    elif "：" in text:
                        period, value = [part.strip() for part in text.split("：", 1)]
                    else:
                        period, value = text, ""
                if period or value:
                    rows.append({"period": period, "value": value or "-"})
        if not rows:
            for item in (profile.get("performanceHistory") or [])[:6]:
                if not isinstance(item, dict):
                    continue
                rows.append(
                    {
                        "period": Handler._first_clean(item.get("period"), item.get("quarter")),
                        "value": Handler._first_clean(item.get("managerRating"), item.get("rating"), item.get("selfRating"), "-"),
                    }
                )
        return rows[:6]

    @staticmethod
    def _normalize_talent_reviews(profile):
        rows = []
        for item in profile.get("talentReviewHistory") or []:
            if not isinstance(item, dict):
                continue
            period = Handler._first_clean(item.get("period"), item.get("year"))
            value = Handler._first_clean(item.get("value"), item.get("grid"), item.get("nineBox"))
            if period or value:
                rows.append({"period": period, "value": value or "-"})
        return rows[:6]

    @staticmethod
    def _normalize_work_history(profile):
        rows = []
        for item in profile.get("workHistory") or []:
            if isinstance(item, dict):
                project = Handler._first_clean(item.get("project"), item.get("name"), item.get("title"))
                date = Handler._first_clean(item.get("date"), item.get("period"), item.get("startDate"))
            else:
                text = Handler._clean_text(item)
                if "：" in text:
                    project, date = [part.strip() for part in text.split("：", 1)]
                elif ":" in text:
                    project, date = [part.strip() for part in text.split(":", 1)]
                else:
                    project, date = text, ""
            if project:
                rows.append({"project": project, "date": date})
        return rows[:12]

    @staticmethod
    def _normalize_performance_comments(profile):
        rows = []
        for item in profile.get("performanceHistory") or []:
            if not isinstance(item, dict):
                continue
            period = Handler._first_clean(item.get("period"), item.get("quarter"))
            manager = Handler._clean_text(item.get("managerComment"))
            employee = Handler._clean_text(item.get("employeeSummary"))
            rating = Handler._first_clean(item.get("managerRating"), item.get("rating"), item.get("selfRating"))
            if period or manager or employee:
                rows.append({"period": period, "rating": rating, "managerComment": manager, "employeeSummary": employee})
        return rows[:4]

    @staticmethod
    def _person_profile_validation(profile_card):
        required = {
            "姓名": profile_card["basicInfo"].get("name"),
            "工号": profile_card["basicInfo"].get("employeeId"),
            "部门": profile_card["basicInfo"].get("departmentPath"),
            "职位": profile_card["basicInfo"].get("title"),
            "职级": profile_card["basicInfo"].get("level"),
            "序列": profile_card["basicInfo"].get("sequence"),
            "绩效记录": profile_card.get("performance"),
            "人才盘点": profile_card.get("talentReviews"),
            "晋升路径": profile_card.get("promotionHistory"),
            "项目履历": profile_card.get("workHistory"),
        }
        matched = [key for key, value in required.items() if bool(value)]
        missing = [key for key, value in required.items() if not bool(value)]
        return {
            "matchedFields": matched,
            "missingFields": missing,
            "completeness": round(len(matched) / max(len(required), 1), 2),
        }

    def _person_profile_card_from_record(self, person, mcp_status=None):
        profile = person.get("profile") if isinstance(person.get("profile"), dict) else {}
        source = {**profile, **person}
        name = self._first_clean(source.get("name"), profile.get("name"))
        level = self._first_clean(source.get("level"), source.get("levelRaw"))
        title = self._first_clean(source.get("title"))
        recent_performance = self._normalize_recent_performance(source)
        talent_reviews = self._normalize_talent_reviews(source)
        promotion_history = [
            {
                "date": self._first_clean(item.get("date"), item.get("time")),
                "level": self._first_clean(item.get("level"), item.get("jobLevel")),
                "reason": self._first_clean(item.get("reason"), item.get("remark")),
            }
            for item in (source.get("promotionHistory") or [])
            if isinstance(item, dict)
        ][:12]
        work_history = self._normalize_work_history(source)
        comments = self._normalize_performance_comments(source)
        latest_review = talent_reviews[0]["value"] if talent_reviews else ""
        latest_perf = recent_performance[0]["value"] if recent_performance else ""
        profile_card = {
            "type": "person_profile_card",
            "id": self._first_clean(source.get("employeeId"), name),
            "basicInfo": {
                "name": name,
                "employeeId": self._first_clean(source.get("employeeId")),
                "gender": self._first_clean(source.get("gender")),
                "age": self._first_clean(source.get("age")),
                "departmentPath": self._first_clean(source.get("departmentPath"), source.get("departmentPathRaw"), source.get("department")),
                "title": title,
                "level": level,
                "sequence": self._first_clean(source.get("sequence")),
                "status": self._first_clean(source.get("status")),
                "employmentType": self._first_clean(source.get("employmentType")),
                "hireDate": self._first_clean(source.get("hireDate")),
                "tenure": self._first_clean(source.get("keyAttribute"), source.get("tenure"), source.get("司龄")),
                "highestEducation": self._first_clean(source.get("highestEducation")),
                "graduationSchool": self._first_clean(source.get("graduationSchool")),
                "dataUpdatedAt": self._first_clean(source.get("dataUpdatedAt")),
            },
            "profileTags": self._normalize_profile_tags(source.get("profileTags")),
            "performance": recent_performance,
            "talentReviews": talent_reviews,
            "promotionHistory": promotion_history,
            "workHistory": work_history,
            "comments": comments,
            "aiSummary": [
                {
                    "title": "综合评价",
                    "content": f"{name or '该员工'}{('，' + level) if level else ''}{('，' + title) if title else ''}。当前档案已匹配基础信息、绩效、盘点、晋升和履历字段，建议结合业务负责人反馈继续复核。",
                },
                {
                    "title": "绩效与盘点",
                    "content": f"最近绩效为{latest_perf or '暂无记录'}，最近人才盘点为{latest_review or '暂无记录'}。如需判断波动原因，应继续追问具体季度表现和项目背景。",
                },
                {
                    "title": "发展建议",
                    "content": "建议围绕关键项目交付、跨团队协作、可量化绩效指标和下一职级要求做结构化复盘。",
                },
            ],
            "source": {
                "profileSource": "HRobot MCP 人才档案快照",
                "mcpStatus": mcp_status or {"checked": False, "ok": False, "message": "未触发实时 MCP 校验。"},
            },
        }
        profile_card["validation"] = self._person_profile_validation(profile_card)
        return profile_card

    def _build_person_profile_card(self, query, history=None):
        history_text = "\n".join(str(item.get("content", "")) for item in (history or [])[-4:] if isinstance(item, dict))
        arguments = self._extract_person_profile_query(query) or self._extract_person_profile_query(history_text)
        if not arguments:
            arguments = self._extract_talent_search_arguments(query) or self._extract_talent_search_arguments(history_text)
        if not arguments:
            return {"error": "请在输入框里写清楚要生成档案的姓名或工号，例如：梁显耀的人员信息，或 工号 2219。"}

        mcp_status = {"checked": True, "ok": False, "message": "MCP 未返回可用内容。"}
        try:
            mcp_result = self._mcp_call_tool("talent-profile-search", arguments)
            mcp_status = {
                "checked": True,
                "ok": bool(self._clean_text(mcp_result)),
                "message": "实时 MCP 校验通过。" if self._clean_text(mcp_result) else "MCP 返回为空。",
                "preview": self._clean_text(mcp_result)[:600],
            }
        except Exception as error:
            mcp_status = {"checked": True, "ok": False, "message": f"实时 MCP 校验失败：{error}"}
            _agent_debug_log(
                "server.py:_build_person_profile_card",
                "MCP profile card validation failed",
                {"arguments": arguments, "errorType": type(error).__name__, "error": str(error)},
                "MCP_PROFILE_CARD",
            )

        searchable_people = []
        seen_profile_ids = set()
        for person in self.store.profiles():
            if not isinstance(person, dict):
                continue
            profile = person.get("profile") if isinstance(person.get("profile"), dict) else person
            key = self._first_clean(person.get("employeeId"), profile.get("employeeId"), person.get("name"), profile.get("name"))
            if key in seen_profile_ids:
                continue
            seen_profile_ids.add(key)
            searchable_people.append(person)
        candidates = [person for person in searchable_people if self._profile_matches_query(person, arguments)]
        if not candidates and arguments.get("names"):
            names = set(arguments.get("names") or [])
            candidates = [
                person
                for person in searchable_people
                if isinstance(person, dict) and any(name and name in self._clean_text(person.get("name")) for name in names)
            ]
        if not candidates:
            return {
                "error": "未能在本地 HRobot 人才档案快照中匹配到该人员。请确认姓名或工号，或先同步/导入 MCP 人才档案快照。",
                "arguments": arguments,
                "mcpStatus": mcp_status,
            }
        card = self._person_profile_card_from_record(candidates[0], mcp_status=mcp_status)
        return {
            "message": f"已生成{card['basicInfo'].get('name') or '该员工'}的人员档案，右侧“人员档案”中可查看完整信息。",
            "arguments": arguments,
            "profile": card,
            "matchCount": len(candidates),
        }

    @staticmethod
    def _image_request_payload(model, prompt, size):
        return {
            "model": model,
            "prompt": prompt,
            "size": size or "1024x1024",
            "n": 1,
        }

    def _call_ai_model(self, messages, enable_mcp=False, timeout=300):
        config = self.store.ai_config()["multimodal"]
        if not (config.get("apiKey") and config.get("baseUrl") and config.get("model")):
            return {
                "configured": False,
                "message": "请先在 09 设置页配置模型 Base URL、模型名，并通过本次运行输入 API Key，或设置 HROBOT_AI_API_KEY 环境变量。"
            }

        endpoint = config["baseUrl"].rstrip("/") + "/chat/completions"
        tools, tool_map = self._mcp_tools_for_ai() if enable_mcp else ([], {})
        if tools:
            messages = [
                *messages[:-1],
                {
                    "role": "system",
                    "content": "当用户问题需要查询实时 HR 数据、通讯录、部门或人才档案时，优先调用可用的 HRobot MCP 工具；工具结果为空或报错时，请明确说明限制。",
                },
                messages[-1],
            ]
        for _ in range(4):
            payload = self._ai_request_payload_with_tools(config["model"], messages, tools)
            try:
                request_body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
            except (TypeError, UnicodeEncodeError) as error:
                _agent_debug_log(
                    "server.py:_call_ai_model",
                    "AI payload json.dumps failed",
                    {"errorType": type(error).__name__, "error": str(error)},
                    "H3",
                )
                return {"configured": True, "error": f"AI 请求序列化失败: {error}"}
            request = Request(
                endpoint,
                data=request_body,
                headers={
                    "Authorization": f"Bearer {config['apiKey']}",
                    "Content-Type": "application/json",
                },
                method="POST",
            )
            try:
                with urlopen(request, timeout=timeout) as response:
                    data = json.loads(response.read().decode("utf-8"))
            except HTTPError as error:
                detail = error.read().decode("utf-8", errors="replace")
                _agent_debug_log(
                    "server.py:_call_ai_model",
                    "AI HTTPError",
                    {"code": error.code, "detailPreview": detail[:500]},
                    "H4",
                )
                return {"configured": True, "error": f"模型接口返回 {error.code}: {detail}"}
            except URLError as error:
                _agent_debug_log(
                    "server.py:_call_ai_model",
                    "AI URLError",
                    {"reason": str(error.reason)},
                    "H4",
                )
                return {"configured": True, "error": f"无法连接模型接口: {error.reason}"}

            message = data.get("choices", [{}])[0].get("message", {})
            tool_calls = message.get("tool_calls") or []
            if not tool_calls:
                content = message.get("content", "")
                return {"configured": True, "message": content or "模型未返回内容。", "raw": data}
            messages.append(
                {
                    "role": "assistant",
                    "content": message.get("content") or "",
                    "tool_calls": tool_calls,
                }
            )
            for tool_call in tool_calls:
                function = tool_call.get("function", {}) if isinstance(tool_call, dict) else {}
                ai_tool_name = function.get("name", "")
                mcp_tool_name = tool_map.get(ai_tool_name)
                try:
                    arguments = json.loads(function.get("arguments") or "{}")
                except json.JSONDecodeError:
                    arguments = {}
                try:
                    result_text = self._mcp_call_tool(mcp_tool_name, arguments) if mcp_tool_name else f"未知工具：{ai_tool_name}"
                except Exception as error:
                    result_text = f"MCP 工具调用失败：{error}"
                    _agent_debug_log(
                        "server.py:_call_ai_model",
                        "MCP tools/call failed",
                        {"tool": mcp_tool_name, "errorType": type(error).__name__, "error": str(error)},
                        "MCP2",
                    )
                messages.append(
                    {
                        "role": "tool",
                        "tool_call_id": tool_call.get("id", ""),
                        "name": ai_tool_name,
                        "content": result_text[:120000],
                    }
                )
        return {"configured": True, "message": "模型连续调用工具但未生成最终回复，请缩小问题范围后重试。"}

    def _build_intelligence_key_summary(self, query):
        channel = self.store._first_query_value(query, "channel")
        if channel not in {"", "ai_hr", "game_org"}:
            channel = ""
        history_payload = self.store._read_intelligence_file(self.store.intelligence_history_path)
        current_payload = self.store._read_intelligence_file(self.store.intelligence_path)
        items = history_payload["items"] or current_payload["items"]
        items = [item for item in items if isinstance(item, dict)]
        if channel:
            items = [item for item in items if item.get("channel") == channel]

        def item_date(item):
            raw = str(item.get("published_at") or item.get("collected_at") or "")[:10]
            try:
                return datetime.fromisoformat(raw).date()
            except (TypeError, ValueError):
                return None

        dated_items = [(item, item_date(item)) for item in items]
        dated_items = [(item, date_value) for item, date_value in dated_items if date_value]
        if not dated_items:
            return {
                "ok": True,
                "configured": True,
                "message": "暂无可用于分析的上周情报。",
                "item_count": 0,
                "date_range": "",
            }

        latest_date = max(date_value for _, date_value in dated_items)
        start_date = latest_date - timedelta(days=6)
        week_items = [
            item for item, date_value in dated_items
            if start_date <= date_value <= latest_date
        ]
        week_items.sort(
            key=lambda item: (
                int(item.get("importance") or 0),
                str(item.get("published_at") or ""),
            ),
            reverse=True,
        )
        compact_items = [
            {
                "channel": item.get("channel", ""),
                "date": item.get("published_at", ""),
                "source": item.get("source", ""),
                "title": item.get("title", ""),
                "summary": item.get("summary", ""),
                "hrbp_takeaway": item.get("hrbp_takeaway", ""),
                "importance": item.get("importance", ""),
                "confidence": item.get("confidence", ""),
                "keywords": item.get("keywords", [])[:5] if isinstance(item.get("keywords"), list) else [],
            }
            for item in week_items[:30]
        ]
        messages = [
            {
                "role": "system",
                "content": (
                    "你是资深 HRBP 情报分析师，擅长把一周新闻提炼成可供 HRBP 判断的关键信号。"
                    "不要复述新闻列表，不要逐条改写标题；要给出分析结论、趋势判断和对组织人才工作的含义。"
                ),
            },
            {
                "role": "user",
                "content": (
                    f"请基于以下 {len(compact_items)} 条情报，总结 {start_date.isoformat()} 至 {latest_date.isoformat()} 的关键情报。\n"
                    "输出要求：\n"
                    "1. 先用一句话给出本周总判断。\n"
                    "2. 给出 3-5 条关键信号，每条包含：结论、依据、HRBP 含义。\n"
                    "3. 最后给出 2 个下周应继续观察的问题。\n"
                    "4. 使用简洁中文，可以用项目符号；不要输出 JSON，不要写空泛套话，总长度控制在 600 字以内。\n\n"
                    f"{json.dumps(compact_items, ensure_ascii=False)}"
                ),
            },
        ]
        result = self._call_ai_model(messages, timeout=60)
        return {
            "ok": not bool(result.get("error")),
            "configured": result.get("configured", True),
            "message": result.get("message", ""),
            "error": result.get("error", ""),
            "item_count": len(compact_items),
            "date_range": f"{start_date.isoformat()} 至 {latest_date.isoformat()}",
            "generated_at": datetime.now(LOCAL_TZ).isoformat(timespec="seconds"),
        }

    def _call_image_model(self, prompt, size="1024x1024"):
        config = self.store.ai_config()["image"]
        if not (config.get("apiKey") and config.get("baseUrl") and config.get("model")):
            return {
                "configured": False,
                "message": "请先在 09 设置页配置图片模型，或设置 HROBOT_IMAGE_API_KEY 环境变量。",
            }
        endpoint = config["baseUrl"].rstrip("/") + "/images/generations"
        payload = self._image_request_payload(config["model"], prompt, size)
        request = Request(
            endpoint,
            data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            headers={
                "Authorization": f"Bearer {config['apiKey']}",
                "Content-Type": "application/json",
            },
            method="POST",
        )
        try:
            with urlopen(request, timeout=300) as response:
                data = json.loads(response.read().decode("utf-8"))
        except HTTPError as error:
            detail = error.read().decode("utf-8", errors="replace")
            return {"configured": True, "error": f"图像模型接口返回 {error.code}: {detail}"}
        except URLError as error:
            return {"configured": True, "error": f"无法连接图像模型接口: {error.reason}"}

        first = (data.get("data") or [{}])[0]
        if first.get("b64_json"):
            return {
                "configured": True,
                "model": config["model"],
                "b64_json": first["b64_json"],
                "raw": data,
            }
        if first.get("url"):
            try:
                with urlopen(first["url"], timeout=120) as image_response:
                    return {
                        "configured": True,
                        "model": config["model"],
                        "imageBytes": image_response.read(),
                        "raw": data,
                    }
            except URLError as error:
                return {"configured": True, "error": f"无法下载生成图片: {error.reason}"}
        return {"configured": True, "error": "图像模型未返回 b64_json 或 url。", "raw": data}

    def _build_ai_messages(self, question, history):
        context = self.store.analysis_context()
        context_text = json.dumps(context, ensure_ascii=False)
        if len(context_text) > 120000:
            context_text = context_text[:120000] + "\n...内容过长，后续数据已截断。"
        system = (
            "你是资深 HRBP 和人才盘点顾问。请基于系统提供的人才盘点、九宫格、档案和校准数据回答问题。"
            "不得编造不存在的数据；涉及个人评价时要谨慎、具体、可追溯，并优先服务 HR 复盘、校准和发展建议场景。"
        )
        messages = [
            {"role": "system", "content": system},
            {"role": "system", "content": f"可用数据如下：\n{context_text}"},
        ]
        for item in history[-12:]:
            role = item.get("role")
            content = item.get("content")
            if role in {"user", "assistant"} and content:
                messages.append({"role": role, "content": content})
        messages.append({"role": "user", "content": question})
        return messages

    def _build_report_messages(self, instruction, department=None, report_type=None, selected_assets=None):
        context = self.store.analysis_context(department=department)
        context_text = json.dumps(context, ensure_ascii=False)
        if len(context_text) > 100000:
            context_text = context_text[:100000] + "\n...内容过长，后续数据已截断。"
        preset = REPORT_PRESETS.get(report_type or "") or REPORT_PRESETS["talent-review"]
        target_name = self.store._extract_360_report_person_name("", instruction) if report_type == "360" else ""
        selected_assets = selected_assets if isinstance(selected_assets, dict) else None
        selected_abilities = selected_assets.get("abilities", []) if selected_assets is not None else None
        selected_skills = selected_assets.get("skills", []) if selected_assets is not None else None
        selected_materials = selected_assets.get("materials", []) if selected_assets is not None else None
        assets_text = json.dumps(
            self.store.report_asset_context(
                report_type,
                target_name=target_name,
                selected_abilities=selected_abilities,
                selected_skills=selected_skills,
                selected_materials=selected_materials,
            ),
            ensure_ascii=False,
        )
        if len(assets_text) > 80000:
            assets_text = assets_text[:80000] + "\n...内容过长，后续材料已截断。"
        system = (
            "你是资深 HRBP 和组织人才分析顾问，负责基于真实材料生成可直接用于人才复盘的报告。"
            "请严格依据系统提供的人才盘点、档案、能力中心、导入 skill 和材料，不得编造评分、姓名、反馈原文或结论。"
            f"本次报告类型为：{preset['name']}。请优先遵循能力中心中所选能力的说明、结构、口径和注意事项。"
        )
        if department:
            system += f"\n请只分析部门或组织范围：{department}，不要扩展到无关人员。"
        if report_type == "360" and target_name:
            system += f"\n本次 360 报告对象为：{target_name}。只使用该对象的 360 材料生成报告；如已有标准版与核心版即可生成，不要求必须存在详细版。"
        user_instruction = instruction or f"请生成一份{preset['name']}，优先使用已选择的能力、skill 和材料，并结合 2026 人才盘点数据。"
        markdown_instruction = (
            "输出格式要求：只返回 Markdown 正文，不要返回 HTML，不要把内容放进代码块。"
            "请使用清晰的一级到三级标题、短段落、项目符号和必要的表格。"
        )
        system = f"{system}\n{markdown_instruction}"
        user_instruction = f"{user_instruction}\n\n{markdown_instruction}"
        return [
            {"role": "system", "content": system},
            {"role": "system", "content": f"人才盘点与档案数据：\n{context_text}"},
            {"role": "system", "content": f"能力中心、导入 skill 和分析材料：\n{assets_text}"},
            {"role": "user", "content": user_instruction},
        ]

    def _build_design_prompt(self, poster_type, style, size, requirement):
        config = self.store.design_prompt_config()
        reference_files = [item.get("name", "") for item in config.get("referenceFiles", []) if item.get("name")]
        reference_summary = "无" if not reference_files else "；".join(reference_files[:12])
        prompt = config.get("template", DEFAULT_DESIGN_PROMPT_CONFIG["template"])
        values = {
            "basePrompt": config.get("basePrompt", ""),
            "brandRequirements": config.get("brandRequirements", ""),
            "customRequirements": config.get("customRequirements", "") or "无额外固定限定。",
            "referenceInstructions": config.get("referenceInstructions", ""),
            "referenceSummary": reference_summary,
            "posterType": poster_type,
            "style": style,
            "size": size,
            "requirement": requirement,
        }
        try:
            return prompt.format(**values)
        except (KeyError, ValueError):
            return "\n".join(
                [
                    values["basePrompt"],
                    values["brandRequirements"],
                    values["customRequirements"],
                    values["referenceInstructions"],
                    f"可用参考素材：{values['referenceSummary']}",
                    f"海报类型：{poster_type}",
                    f"风格：{style}",
                    f"尺寸：{size}",
                    f"需求：{requirement}",
                ]
            )

    def _ai_summarize_data_source_scan(self, scan_payload):
        if scan_payload.get("error"):
            return None
        if not self.store.ai_config_status().get("multimodal", {}).get("configured"):
            return None
        bounded = {
            "localFolder": scan_payload.get("localFolder", ""),
            "fileCount": scan_payload.get("fileCount", 0),
            "summary": scan_payload.get("summary", {}),
            "files": [
                {
                    "relativePath": item.get("relativePath", ""),
                    "extension": item.get("extension", ""),
                    "detectedType": item.get("detectedType", ""),
                    "headers": (item.get("headers") or [])[:12],
                    "size": item.get("size", 0),
                }
                for item in (scan_payload.get("files") or [])[:80]
            ],
        }
        messages = [
            {
                "role": "system",
                "content": (
                    "你是资深 HR 数据产品顾问。请基于本地文件夹扫描元数据，给 HR 用户一段简洁、可执行的数据盘点总结。"
                    "不要逐条列文件。请覆盖：大概有哪些内容、文件数量、可做哪些分析或信息呈现、数据结构/文件夹治理问题、建议项。"
                    "不要编造扫描结果之外的具体文件内容。"
                ),
            },
            {
                "role": "user",
                "content": (
                    "请用中文输出 Markdown，结构为：总体判断、内容概况、可做分析、结构问题、建议项。\n\n"
                    f"扫描元数据：\n{json.dumps(bounded, ensure_ascii=False)}"
                ),
            },
        ]
        try:
            result = self._call_ai_model(messages, timeout=12)
            message = str(result.get("message") or "").strip() if isinstance(result, dict) else ""
            return message or None
        except Exception as error:
            _agent_debug_log(
                "server.py:_ai_summarize_data_source_scan",
                "AI scan summary failed",
                {"errorType": type(error).__name__, "error": str(error)},
                "DATA_SOURCE_SCAN",
            )
            return None

    def do_GET(self):
        if not self._allow_request():
            return
        parsed = urlparse(self.path)
        path = parsed.path
        query = parse_qs(parsed.query)
        if self._send_static_no_cache():
            return
        if path == "/api/people":
            self._send_json({"people": self.store.people()})
            return
        if path == "/api/review-results":
            self._send_json({"results": self.store.review_results()})
            return
        if path == "/api/profiles":
            self._send_json({"profiles": self.store.profiles()})
            return
        if path == "/api/people/context":
            department = (query.get("department") or [""])[0].strip() or None
            self._send_json(self.store.people_data().analysis_context(department=department))
            return
        if path == "/api/overrides":
            self._send_json(self.store.overrides())
            return
        if path == "/api/profile-notes":
            self._send_json(self.store.profile_notes())
            return
        if path == "/api/ai/config":
            self._send_json(self.store.ai_config_status())
            return
        if path == "/api/data-sources/config":
            self._send_json(self.store.data_source_config())
            return
        if path == "/api/server/status":
            self._send_json(server_status_payload())
            return
        if path == "/api/app/update":
            try:
                self._send_json(self.store.update_status(self.store._first_query_value(query, "sourcePath")))
            except Exception as error:
                self._send_json(
                    {
                        "app": app_version_payload(),
                        "config": self.store.update_config(),
                        "configured": True,
                        "releasePage": GITHUB_RELEASE_PAGE_URL,
                        "updateAvailable": False,
                        "error": str(error),
                    },
                    status=400,
                )
            return
        if path == "/api/home-memos":
            self._send_json(self.store.home_memos())
            return
        if path == "/api/home-avatar":
            self._send_json(self.store.home_avatar_config())
            return
        if path == "/api/intelligence/key-summary":
            self._send_json(self._build_intelligence_key_summary(query))
            return
        if path == "/api/intelligence":
            self._send_json(self.store.intelligence(query))
            return
        if path == "/api/intelligence/config":
            self._send_json({"config": self.store.intelligence_config(), "status": self.store.intelligence_update_status()})
            return
        if path == "/api/design/posters":
            self._send_json(self.store.design_posters())
            return
        if path == "/api/design/prompt-config":
            self._send_json(self.store.design_prompt_config())
            return
        if path == "/api/agent-projects":
            self._send_json(self.store.agent_projects())
            return
        if path == "/api/ai/context":
            self._send_json(self.store.analysis_context())
            return
        if path == "/api/talent-pools":
            self._send_json(self.store.talent_pools())
            return
        if path == "/api/report/assets":
            self._send_json(self.store.report_assets())
            return
        if path == "/api/report/abilities":
            self._send_json(self.store.report_abilities())
            return
        if path == "/api/report/presets":
            self._send_json({"presets": self.store.report_presets()})
            return
        if path == "/api/report/list":
            self._send_json({"reports": self.store.generated_report_list()})
            return
        if path == "/api/import/sources":
            self._send_json(self.store.import_sources())
            return
        if path == "/api/report/latest":
            if query.get("list"):
                self._send_json({"reports": self.store.generated_report_list()})
                return
            self._send_json(self.store.generated_report((query.get("id") or [None])[0]))
            return
        if path == "/api/export/calibrated-excel":
            try:
                output_path = self.store.export_calibrated_excel()
                self._send_file(output_path, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except FileNotFoundError as error:
                self._send_json({"error": str(error)}, status=404)
            return
        if path == "/api/export/calibration-differences":
            try:
                output_path = self.store.export_calibration_differences()
                self._send_file(output_path, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except FileNotFoundError as error:
                self._send_json({"error": str(error)}, status=404)
            return
        super().do_GET()

    def do_DELETE(self):
        if not self._allow_request():
            return
        path = urlparse(self.path).path
        if path == "/api/report":
            payload = self._read_request_json()
            if payload is None:
                self._send_json({"error": "Invalid JSON"}, status=400)
                return
            try:
                self._send_json(self.store.delete_generated_report(payload.get("id", "")))
            except FileNotFoundError as error:
                self._send_json({"error": str(error)}, status=404)
            except ValueError as error:
                self._send_json({"error": str(error)}, status=400)
            return
        if path == "/api/import/file":
            payload = self._read_request_json()
            if payload is None:
                self._send_json({"error": "Invalid JSON"}, status=400)
                return
            try:
                self._send_json(self.store.delete_imported_file(payload.get("type", ""), payload.get("filename", "")))
            except FileNotFoundError as error:
                self._send_json({"error": str(error)}, status=404)
            except ValueError as error:
                self._send_json({"error": str(error)}, status=400)
            return
        self._send_json({"error": "Not found"}, status=404)

    def do_POST(self):
        if not self._allow_request():
            return
        path = urlparse(self.path).path
        if path == "/api/import/review-excel":
            filename, content = self._read_multipart_file()
            if not content:
                self._send_json({"error": "请上传人才盘点 Excel 文件。"}, status=400)
                return
            temp_dir = self.store.data_dir / "uploads"
            temp_dir.mkdir(parents=True, exist_ok=True)
            temp_path = temp_dir / (Path(filename).name or "review.xlsx")
            temp_path.write_bytes(content)
            try:
                self._send_json(self.store.import_review_excel(temp_path, filename))
            except Exception as error:
                self._send_json({"error": str(error)}, status=400)
            return
        if path == "/api/import/profiles-json":
            filename, content = self._read_multipart_file()
            if not content:
                self._send_json({"error": "请上传人才档案 JSON 文件。"}, status=400)
                return
            try:
                data = json.loads(content.decode("utf-8-sig"))
                self._send_json(self.store.import_profiles_json(data, filename))
            except Exception as error:
                self._send_json({"error": str(error)}, status=400)
            return
        if path == "/api/import/employee-roster-excel":
            files = self._read_multipart_files()
            if not files:
                self._send_json({"error": "请上传员工花名册 Excel 文件。"}, status=400)
                return
            temp_dir = self.store.data_dir / "uploads"
            temp_dir.mkdir(parents=True, exist_ok=True)
            temp_files = []
            try:
                for index, (filename, content) in enumerate(files, start=1):
                    if not content:
                        continue
                    temp_path = temp_dir / (Path(filename).name or f"employee_roster_{index}.xlsx")
                    temp_path.write_bytes(content)
                    temp_files.append((temp_path, filename))
                self._send_json(self.store.import_employee_roster_excels(temp_files))
            except Exception as error:
                self._send_json({"error": str(error)}, status=400)
            return
        if path in {"/api/report/upload-skill", "/api/report/upload-material"}:
            filename, content = self._read_multipart_file()
            if not content:
                self._send_json({"error": "请先选择要上传的文件。"}, status=400)
                return
            kind = "skill" if path.endswith("upload-skill") else "material"
            try:
                self._send_json(self.store.import_report_asset(kind, filename, content))
            except Exception as error:
                self._send_json({"error": str(error)}, status=400)
            return
        if path == "/api/report/ability":
            form = self._read_multipart_form()
            fields = form.get("fields", {})
            zip_file = next((item for item in form.get("files", []) if item.get("content")), None)
            try:
                self._send_json(self.store.save_report_ability(
                    fields.get("name", ""),
                    description_md=fields.get("descriptionMd", ""),
                    zip_filename=zip_file.get("filename", "") if zip_file else "",
                    zip_content=zip_file.get("content") if zip_file else None,
                    ability_id=fields.get("id", ""),
                ))
            except Exception as error:
                self._send_json({"error": str(error)}, status=400)
            return
        if path == "/api/agent-projects/upload":
            filename, content = self._read_multipart_file()
            if not content:
                self._send_json({"error": "请上传独立 Web 项目 zip。"}, status=400)
                return
            try:
                self._send_json(self.store.import_agent_project_zip(filename, content, ai_provider=self._call_ai_model))
            except Exception as error:
                self._send_json({"error": str(error)}, status=400)
            return
        payload = self._read_request_json()
        if payload is None:
            self._send_json({"error": "Invalid JSON"}, status=400)
            return
        if path == "/api/overrides":
            changes = payload.get("changes", [])
            self._send_json(self.store.save_overrides(changes))
            return
        if path == "/api/profile-notes":
            try:
                self._send_json(self.store.save_profile_note(payload))
            except ValueError as error:
                self._send_json({"error": str(error)}, status=400)
            return
        if path == "/api/ai/config":
            self._send_json(self.store.save_ai_config(payload))
            return
        if path == "/api/data-sources/config":
            mcp_payload = payload.get("mcp") if isinstance(payload.get("mcp"), dict) else {}
            if mcp_payload:
                self.store.save_mcp_config(mcp_payload)
            self._send_json(self.store.save_data_source_config(payload))
            return
        if path == "/api/data-sources/scan":
            scan_payload = self.store.scan_local_data_sources(payload.get("localFolder"))
            ai_summary = self._ai_summarize_data_source_scan(scan_payload)
            if ai_summary:
                scan_payload["aiSummary"] = ai_summary
            self._send_json(scan_payload)
            return
        if path == "/api/server/restart":
            status = server_status_payload()
            self._send_json({**status, "restarting": True, "message": "服务器正在重启，请稍候刷新状态。"})
            schedule_server_restart()
            return
        if path == "/api/app/update/config":
            self._send_json({"config": self.store.save_update_config(payload), "app": app_version_payload()})
            return
        if path == "/api/app/update/check":
            try:
                self._send_json(self.store.update_status(payload.get("sourcePath") or ""))
            except Exception as error:
                self._send_json(
                    {
                        "app": app_version_payload(),
                        "config": self.store.update_config(),
                        "updateAvailable": False,
                        "error": str(error),
                    },
                    status=400,
                )
            return
        if path == "/api/app/update/install":
            try:
                self._send_json(self.store.install_update(payload.get("sourcePath") or ""))
            except Exception as error:
                self._send_json({"error": str(error), "app": app_version_payload(), "config": self.store.update_config()}, status=400)
            return
        if path == "/api/home-memos":
            self._send_json(self.store.save_home_memo(payload.get("date", ""), payload.get("text", "")))
            return
        if path == "/api/home-memos/delete":
            self._send_json(self.store.delete_home_memo(payload.get("date", "")))
            return
        if path == "/api/home-memos/migrate":
            self._send_json(self.store.save_home_memos(payload.get("records", [])))
            return
        if path == "/api/home-avatar":
            try:
                self._send_json(self.store.save_home_avatar_config(payload.get("selectedSrc", "")))
            except ValueError as error:
                self._send_json({"error": str(error)}, status=400)
            return
        if path == "/api/intelligence/config":
            self._send_json({"config": self.store.save_intelligence_config(payload), "status": self.store.intelligence_update_status()})
            return
        if path == "/api/design/prompt-config":
            self._send_json(self.store.save_design_prompt_config(payload))
            return
        if path == "/api/agent-projects/delete":
            try:
                self._send_json(self.store.delete_agent_project(payload.get("id", "")))
            except FileNotFoundError as error:
                self._send_json({"error": str(error)}, status=404)
            except ValueError as error:
                self._send_json({"error": str(error)}, status=400)
            except Exception as error:
                self._send_json({"error": str(error)}, status=500)
            return
        if path == "/api/agent-projects/analyze":
            try:
                self._send_json(self.store.analyze_agent_project(payload.get("id", ""), ai_provider=self._call_ai_model, force_ai=bool(payload.get("forceAi", True))))
            except FileNotFoundError as error:
                self._send_json({"error": str(error)}, status=404)
            except ValueError as error:
                self._send_json({"error": str(error)}, status=400)
            except Exception as error:
                self._send_json({"error": str(error)}, status=500)
            return
        if path == "/api/agent-projects/open":
            try:
                self._send_json(self.store.open_agent_project(payload.get("id", "")))
            except FileNotFoundError as error:
                self._send_json({"error": str(error)}, status=404)
            except ValueError as error:
                self._send_json({"error": str(error)}, status=400)
            except Exception as error:
                self._send_json({"error": str(error)}, status=500)
            return
        if path == "/api/intelligence/update":
            self._send_json(self.store.update_intelligence_now(trigger="manual", target_date=(payload.get("date") or "").strip() or None))
            return
        if path == "/api/talent-pools":
            self._send_json(self.store.save_talent_pools(payload.get("pools", [])))
            return
        if path == "/api/ai/chat":
            question = (payload.get("message") or "").strip()
            if not question:
                self._send_json({"error": "请输入问题。"}, status=400)
                return
            history = payload.get("history", [])
            messages = self._build_ai_messages(question, history)
            mcp_context = self._mcp_profile_context_for_question(question, history)
            if mcp_context:
                messages.insert(
                    -1,
                    {
                        "role": "system",
                        "content": (
                            "以下是 HRobot MCP 实时查询结果。回答涉及该员工档案、任职、晋升、绩效、盘点等问题时，"
                            "必须优先依据这段 MCP 结果；不要因为本地九宫格上下文缺字段就判断查不到。\n\n"
                            f"{mcp_context}"
                        ),
                    },
                )
            self._send_json(self._call_ai_model(messages, enable_mcp=True))
            return
        if path == "/api/mcp/person-profile-card":
            query = (payload.get("message") or payload.get("query") or "").strip()
            if not query:
                self._send_json({"error": "请输入要生成档案的姓名或工号。"}, status=400)
                return
            result = self._build_person_profile_card(query, payload.get("history", []))
            self._send_json(result, status=404 if result.get("error") else 200)
            return
        if path == "/api/ai/image/test":
            prompt = (payload.get("prompt") or "HRobot orange simple poster test").strip()
            result = self._call_image_model(prompt, payload.get("size") or "1024x1024")
            self._send_json(
                {
                    "ok": not bool(result.get("error")) and result.get("configured") is not False,
                    "configured": result.get("configured", True),
                    "model": result.get("model", self.store.ai_config()["image"].get("model", "")),
                    "message": result.get("message", ""),
                    "error": result.get("error", ""),
                    "hasImage": bool(result.get("b64_json") or result.get("imageBytes")),
                },
                status=502 if result.get("error") else 200,
            )
            return
        if path == "/api/design/posters/generate":
            requirement = (payload.get("prompt") or payload.get("requirement") or "").strip()
            if not requirement:
                self._send_json({"error": "请先填写海报需求描述。"}, status=400)
                return
            poster_type = (payload.get("posterType") or "custom").strip()
            style = (payload.get("style") or "modern").strip()
            size = (payload.get("size") or "1024x1024").strip()
            prompt = self._build_design_prompt(poster_type, style, size, requirement)
            result = self._call_image_model(prompt, size)
            if result.get("configured") is False:
                self._send_json(result, status=503)
                return
            if result.get("error"):
                self._send_json(result, status=502)
                return
            try:
                saved = self.store.save_design_poster(
                    {
                        "posterType": poster_type,
                        "style": style,
                        "size": size,
                        "prompt": requirement,
                        "model": result.get("model", self.store.ai_config()["image"].get("model", "")),
                        "b64_json": result.get("b64_json"),
                        "imageBytes": result.get("imageBytes"),
                    }
                )
            except Exception as error:
                self._send_json({"error": f"海报保存失败: {error}"}, status=500)
                return
            self._send_json({"ok": True, "poster": saved})
            return
        if path == "/api/report/generate":
            instruction = (payload.get("instruction") or "").strip()
            department = (payload.get("department") or "").strip() or None
            report_type = (payload.get("reportType") or "talent-review").strip()
            selected_assets = payload.get("selectedAssets")
            if selected_assets is not None and not isinstance(selected_assets, dict):
                self._send_json({"error": "所选报告资料格式不正确。"}, status=400)
                return
            if selected_assets is not None and any(
                not isinstance(selected_assets.get(key, []), list)
                for key in ("abilities", "skills", "materials")
            ):
                self._send_json({"error": "所选报告资料清单格式不正确。"}, status=400)
                return
            if report_type not in REPORT_PRESETS:
                report_type = "talent-review"
            try:
                msgs = self._build_report_messages(
                    instruction,
                    department=department,
                    report_type=report_type,
                    selected_assets=selected_assets,
                )
                result = self._call_ai_model(msgs)
                if result.get("message"):
                    message = result["message"]
                    if self.store._report_content_format(message) == "html":
                        message = self.store._html_to_markdown_text(message)
                        result["message"] = message
                    result["report"] = self.store.save_generated_report(message, instruction, report_type)
                self._send_json(result)
            except Exception as ex:
                import traceback as tb
                _agent_debug_log(
                    "server.py:do_POST:/api/report/generate",
                    "report generation failed",
                    {"errorType": type(ex).__name__, "error": str(ex), "traceback": tb.format_exc()},
                    "report-generate",
                )
                self._send_json({"error": "报告生成异常，请稍后重试或检查服务端日志。"}, status=500)
            return
        if path == "/api/report/ability/delete":
            try:
                self._send_json(self.store.delete_report_ability(payload.get("id", "")))
            except FileNotFoundError as error:
                self._send_json({"error": str(error)}, status=404)
            except ValueError as error:
                self._send_json({"error": str(error)}, status=400)
            return
        if path == "/api/report/html":
            try:
                self._send_json({"report": self.store.generate_report_html(payload.get("id", ""))})
            except FileNotFoundError as error:
                self._send_json({"error": str(error)}, status=404)
            except ValueError as error:
                self._send_json({"error": str(error)}, status=400)
            return
        self._send_json({"error": "Not found"}, status=404)


def start_intelligence_scheduler(store: DataStore):
    def loop():
        while True:
            try:
                config = store.intelligence_config()
                now = datetime.now(LOCAL_TZ)
                run_at = config.get("runAt", "10:00")
                status = store.intelligence_update_status()
                if (
                    config.get("autoEnabled")
                    and now.strftime("%H:%M") == run_at
                    and status.get("lastAutoDate") != now.date().isoformat()
                    and not status.get("running")
                ):
                    store.update_intelligence_now(trigger="auto")
            except Exception as error:
                status = store.intelligence_update_status()
                status.update(
                    {
                        "running": False,
                        "lastFinishedAt": datetime.now(LOCAL_TZ).isoformat(timespec="seconds"),
                        "lastTrigger": "auto",
                        "ok": False,
                        "message": f"自动更新调度异常：{error}",
                    }
                )
                store._save_intelligence_update_status(status)
            time.sleep(30)

    thread = threading.Thread(target=loop, name="intelligence-scheduler", daemon=True)
    thread.start()
    return thread


def main():
    global SERVER_HOST, SERVER_PORT, SERVER_ALLOW_REMOTE_CLIENTS
    parser = argparse.ArgumentParser()
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", default=8767, type=int)
    parser.add_argument("--allow-remote-clients", action="store_true", help="Allow non-local clients. Use only on a trusted network.")
    args = parser.parse_args()
    SERVER_HOST = args.host
    SERVER_PORT = args.port
    SERVER_ALLOW_REMOTE_CLIENTS = args.allow_remote_clients
    Handler.allow_remote_clients = args.allow_remote_clients
    start_intelligence_scheduler(Handler.store)
    server = ReusableThreadingHTTPServer((args.host, args.port), Handler)
    browser_host = "127.0.0.1" if args.host in {"0.0.0.0", "::", ""} else args.host
    browser_url = f"http://{browser_host}:{args.port}/index.html"
    print(f"Talent nine-box app running at {browser_url}")
    if getattr(sys, "frozen", False) and sys.platform == "darwin":
        threading.Timer(1.2, lambda: webbrowser.open(browser_url)).start()
    if args.host not in {"127.0.0.1", "localhost", "::1"} and not args.allow_remote_clients:
        print("Remote clients are blocked by default. Add --allow-remote-clients only on a trusted network.")
    server.serve_forever()


if __name__ == "__main__":
    main()
